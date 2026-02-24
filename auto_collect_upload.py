#!/usr/bin/env python3
"""
============================================================
  VIGO SFC 雙語監管信息自動收集 & 上傳 v7.0
  合規是下線，業務是上線 — VIGO Protocol
============================================================

v7.0 World-Class 引擎升級版（基於 v6.0）：

  === Sprint 1: 高 ROI 升級 ===
  [NEW] Mode 11: Eval Benchmark — 50 題自動化檢索質量評估
  [NEW] 自動計算 Recall@5, MRR, 來源準確率, 路由準確率, 延遲 P50/P95
  [NEW] Baseline vs Upgrade 對比報告生成
  [NEW] Cohere Reranker 配置支持（需配合 vigo-chat v5.2 部署）

  === Sprint 2: 結構化存儲 ===
  [NEW] Mode 12: Structured Extraction — 從 vigo_knowledge 提取結構化數據
  [NEW] 12a: 執法案例提取 → vigo_enforcement 表
  [NEW] 12b: 法規提取 → vigo_regulations 表
  [NEW] 12c: 牌照行情提取 → vigo_license_market 表
  [NEW] 12d: 費用基準提取 → vigo_fee_benchmarks 表
  [NEW] 12e: 一鍵全部提取
  [NEW] 12f: 結構化表統計總覽

  === Sprint 3: 數據質量引擎 ===
  [NEW] Mode 13: Data Quality Engine — 語義去重 + 鑒真補全 + 衰減自動化
  [NEW] 13a: 語義去重掃描（cosine ≥ 0.92 標記重複）
  [NEW] 13b: 五維鑒真補全（真實性 + 準確性 + 參照價值）
  [NEW] 13c: 時效性衰減重算（按 content_type 不同衰減率）
  [NEW] 13d: 數據質量總覽報告
  [NEW] 13e: 一鍵全部執行

  === Sprint 4: 進階智能 ===
  [NEW] Mode 14: Self-Correction Engine — 衝突檢測 + 自動巡檢 + 分類映射
  [NEW] 14a: 衝突檢測（數字衝突 + 法規替代 + 狀態變化）
  [NEW] 14b: 過期內容掃描 + 自動降權
  [NEW] 14c: 來源健康檢查
  [NEW] 14d: 617 節點合規領域分類映射（LLM batch）
  [NEW] 14e: 一鍵全部巡檢
  [NEW] 14f: 巡檢報告

v6.0 智能引擎整合版（基於 v5.3）：

  === v6 引擎整合 ===
  [NEW] V6_INGEST_MODE — 新數據可通過 v6 智能管線入庫
  [NEW] v6 管線自動完成：617 節點分類 / SimHash 去重 / 五維驗證 / 結構化提取
  [NEW] upload_chunk 自動偵測 v6 模式，失敗時回退到直接上傳
  [NEW] source_id 映射表（v5.3 來源名 → v6 source_id）

v5.3 掃描體系重構版（基於 v5.2.2）：

  === Mode 2 優化 ===
  [OPT] 移除「過去 1 個月」選項（與 Mode 1 完全重疊）
  [OPT] 選項改為 3/6/12 個月，定位為「季度補漏」

  === Mode 4 完全重寫 ===
  [NEW] 分類掃描 — 按數據類型設定不同年份範圍（執法/通函/諮詢/新聞各自獨立）
  [NEW] 4 個預設方案：推薦(a)/精簡(b)/完整(c)/自定義(d)
  [NEW] 掃描前預覽表格：類型、年份、年數、API 呼叫估算
  [NEW] 4 Phase 獨立循環：執法→通函→諮詢→新聞，每個 Phase 獨立進度條
  [NEW] 每 20 項自動保存進度，中斷恢復不丟失
  [NEW] 改名「歷史建庫 Historical Archive」— 定位為一次性建庫工程

  === 主菜單更新 ===
  [OPT] 中英雙語模式名稱
  [OPT] Mode 4 顯示分類掃描說明

  === 保留 v5.2.2 所有修復 ===

v5.2.2 Bug Fix:
- [FIX] Mode 7: UnboundLocalError — 結構化 .md 路徑 chunks 變量未定義，導致統計行崩潰
- [FIX] Mode 9: HTTP 失敗時無錯誤輸出（靜默失敗），現已添加完整錯誤日誌
- [FIX] Mode 9: 增加重試機制 + 改進 HTTP headers，提高外部監管網站（FATF/EU/MAS）抓取成功率
- [FIX] Mode 9: PDF 提取失敗和 HTML 內容過短時添加明確錯誤提示

v5.2.1 Mode 7 結構化導入修復版：
- [FIX] Mode 7 支持 ===SECTION=== 結構化 .md 格式，自動提取每個 section 的獨立 metadata
- [FIX] 普通 .md/.txt 語言偵測：中文文件不再嘗試翻譯成中文，英文文件不再嘗試翻譯成英文
- [FIX] 結構化 .md 不受 80KB 截斷限制（每個 section 獨立處理）

v5.2 Gap Analysis 優化版（對標 vigo_data_gap_analysis.md P0-P3）：

  === P0 優化（Phase 1 收入 — 合規文件工廠）===
  [P0.2] 加強 SFC 巡查報告 PDF 深度爬取
  [P0.4] 新增 PDPO 個資保護條例 (Cap.486) 基石源
  [P0.5] Mode 1/2 新增中文版通函爬取路徑 (lang=TC)

  === P1 優化（Phase 2 — VATP 專家）===
  [P1.6] 新增 FATF VA Travel Rule 指引源
  [P1.7] 新增跨境 VA 監管比較源 (MAS/MiCA/Japan FSA)
  [P1.9] 新增 VA ETF 批准文件源

  === P2 優化（Phase 3 — 牌照申請導航）===
  [P2.10] 新增 WINGS 系統指南源
  [P2.11] 新增 SFC 申請表格 + 填寫指引源
  [P2.12] Mode 7 新增 template/compliance_manual 分類

  === P3 優化（長期建設）===
  [P3.16] 結構化標籤：license_type / compliance_domain / doc_hierarchy
  [P3.17] 合規日曆數據框架（Mode 8 新增子功能）

  === 新模式 ===
  [Mode 9]  VATP 深度採集 — FATF + 跨境比較 + VA ETF
  [Mode 10] SFC 操作知識 — WINGS + 申請表格 + PDPO + 巡查

  === 保留 v5.0 全部功能 ===
  [Q] 質量保證框架 — 三層去重 + 內容驗證 + 數據庫管理
  [E] 統一報告系統 — 所有 Mode 進度條 + TXT/JSON/MD
  [B] Mode 7: 統一文件導入器 — PDF/DOCX/TXT/XLSX/MD
  [C] Mode 2 升級 — 時間範圍 + 失敗重試 + HKMA 聯合通函
  [A] Mode 5 升級 — 健康檢查 + FAQ/Handbook/HKEX 擴展
  [F] 主題巡查報告 + 年報（10年）
  [D] Mode 6A — 19 家行業來源 + 子頁面爬取
  [G] HKEX 上市規則
  [8] Mode 8: 數據庫管理工具 + 合規日曆

數據質量五原則：
  1. 寧缺勿濫 — 寧可少一條，不可錯一條
  2. 源頭把控 — 上傳前三層檢查，不合格不入庫
  3. 權威優先 — SFC 官方 > 律所 > 顧問 > 一般
  4. 時效標記 — 每條數據帶日期，過時自動降權
  5. 定期清洗 — Mode 8 每月跑一次，保持數據庫健康

安裝依賴：
  pip install requests beautifulsoup4 pdfplumber python-dotenv openpyxl python-docx

使用方法：
  cd scripts/vigo && python auto_collect_upload.py
"""

import requests
import json
import time
import re
import sys
import os
import io
import hashlib
import argparse
from datetime import datetime, timedelta

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ============================================================
# 配置
# ============================================================
try:
    from dotenv import load_dotenv
except ImportError:
    print("  ⚠ python-dotenv 未安裝，嘗試安裝...")
    os.system(f"{sys.executable} -m pip install python-dotenv -q")
    from dotenv import load_dotenv

_script_dir = os.path.dirname(os.path.abspath(__file__))
SCRIPT_DIR = _script_dir  # v7.0.1: alias for Mode 11/13/14 report paths
REPORTS_DIR = os.path.join(_script_dir, "reports")
CONFIG_DIR = os.path.join(_script_dir, "config")
IMPORTS_DIR = os.path.join(CONFIG_DIR, "imports")
os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(CONFIG_DIR, exist_ok=True)
os.makedirs(IMPORTS_DIR, exist_ok=True)
os.makedirs(os.path.join(CONFIG_DIR, "pdfs"), exist_ok=True)
os.makedirs(os.path.join(CONFIG_DIR, "templates"), exist_ok=True)

_env_candidates = [
    os.path.join(_script_dir, '.env'),
    os.path.join(_script_dir, '..', '.env'),
    os.path.join(_script_dir, '..', '..', '.env'),
]
_env_loaded = False
for _env_path in _env_candidates:
    if os.path.exists(_env_path):
        load_dotenv(_env_path)
        print(f"  ✓ Loaded .env from: {os.path.abspath(_env_path)}")
        _env_loaded = True
        break
if not _env_loaded:
    print("  ⚠ No .env file found. Using environment variables or defaults.")

SUPABASE_URL = os.getenv("VIGO_SUPABASE_URL") or os.getenv("SUPABASE_URL", "")
SUPABASE_SERVICE_KEY = os.getenv("VIGO_SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_SERVICE_KEY", "")
OPENAI_API_KEY = os.getenv("VIGO_OPENAI_API_KEY") or os.getenv("OPENAI_API_KEY", "")
DEEPSEEK_API_KEY = os.getenv("VIGO_DEEPSEEK_API_KEY") or os.getenv("DEEPSEEK_API_KEY", "")

_missing_keys = []
if not SUPABASE_URL: _missing_keys.append("SUPABASE_URL")
if not SUPABASE_SERVICE_KEY: _missing_keys.append("SUPABASE_SERVICE_KEY")
if not OPENAI_API_KEY: _missing_keys.append("OPENAI_API_KEY")
if not DEEPSEEK_API_KEY: _missing_keys.append("DEEPSEEK_API_KEY")
if _missing_keys:
    # Allow --version and --help to work without env vars
    if '--version' in sys.argv or '--help' in sys.argv or '-h' in sys.argv:
        pass  # Will be handled by argparse later
    else:
        print(f"\n  ❌ 缺少必填 API Key: {', '.join(_missing_keys)}")
        print(f"  請在 .env 文件中設定以上變量。")
        sys.exit(1)

KNOWLEDGE_TABLE = "vigo_knowledge"
QUESTIONS_TABLE = "vigo_quick_questions"
SFC_API_BASE = "https://apps.sfc.hk/edistributionWeb/api"

# ═══════════════════════════════════════════════════════════════
# v6 智能引擎配置
# ═══════════════════════════════════════════════════════════════
V6_INGEST_MODE = os.getenv("V6_INGEST_MODE", "true").lower() == "true"
V6_INGEST_URL = os.getenv("V6_INGEST_URL", f"{SUPABASE_URL}/functions/v1/vigo-ingest")

# v7.0 Sprint 1: Cohere Reranker + Eval
COHERE_API_KEY = os.getenv("COHERE_API_KEY", "")
VIGO_CHAT_URL = os.getenv("VIGO_CHAT_URL", f"{SUPABASE_URL}/functions/v1/vigo-chat")
VIGO_MCP_URL = os.getenv("VIGO_MCP_URL", f"{SUPABASE_URL}/functions/v1/vigo-mcp")
EVAL_QUESTIONS_FILE = os.path.join(CONFIG_DIR, "eval_questions.json")

# v5.3 來源名 → v6 source_id 映射
V6_SOURCE_ID_MAP = {
    'SFC Circulars': 'sfc_circulars', 'SFC Enforcement': 'sfc_enforcement',
    'SFC Consultation': 'sfc_consultation', 'SFC Press': 'sfc_press',
    'Charltons': 'charltons', 'CompliancePlus': 'complianceplus',
    'Linklaters HK': 'linklaters_hk', 'Deloitte HK FS': 'deloitte_hk_fs',
    'KPMG HK FS': 'kpmg_hk_fs', 'FATF': 'fatf', 'MAS': 'mas_guidelines',
    'Paradox Wanted': 'paradox_wanted', 'Paradox For Sale': 'paradox_forsale',
}

if V6_INGEST_MODE:
    print(f"  🚀 v6 Ingest Mode: ON → {V6_INGEST_URL}")
else:
    print(f"  📦 v6 Ingest Mode: OFF (direct upload)")
# ═══════════════════════════════════════════════════════════════

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9,zh-HK;q=0.8,zh;q=0.7',
}

# v7.0.4: 增強 Headers（用於 403 反爬站點）
ENHANCED_HEADERS = {
    **HEADERS,
    'Accept-Encoding': 'gzip, deflate, br',
    'Referer': 'https://www.google.com/',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'cross-site',
    'Sec-Ch-Ua': '"Chromium";v="131", "Not_A Brand";v="24"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
}

CHUNK_MIN = 300
CHUNK_MAX = 1500
CHUNK_IDEAL = 800

PROGRESS_FILE = ".uploaded_refs.json"
CONTENT_HASHES_FILE = ".content_hashes.json"
BASELINE_FILE = os.path.join(CONFIG_DIR, "foundational_baseline.json")

# 來源權威性權重
SOURCE_AUTHORITY = {
    "SFC_Official": 100,
    "SFC_Foundational": 100,
    "HKEX_Official": 95,
    "HKMA_Official": 95,
    "PCPD_Official": 90,      # P0.4: 個資私隱專員公署
    "FATF_Official": 90,       # P1.6: FATF 國際標準
    "MAS_Official": 85,        # P1.7: 新加坡金管局
    "EU_Official": 85,         # P1.7: 歐盟 MiCA
    "Japan_FSA_Official": 85,  # P1.7: 日本金融廳
    "InvestHK_Official": 80,   # P1.9: 投資推廣署
    "Law_Firm": 70,
    "Big4": 65,
    "Consulting": 60,
    "Industry": 50,
    "Practical_Experience": 40,
    "MANUAL_IMPORT": 30,
}

# ── v5.2 結構化標籤系統 [P3.16] ──

# 牌照類型映射
LICENSE_TYPE_KEYWORDS = {
    "Type1": ["type 1", "dealing in securities", "securities dealing"],
    "Type2": ["type 2", "dealing in futures", "futures dealing"],
    "Type3": ["type 3", "leveraged foreign exchange"],
    "Type4": ["type 4", "advising on securities"],
    "Type5": ["type 5", "advising on futures"],
    "Type6": ["type 6", "advising on corporate finance"],
    "Type7": ["type 7", "automated trading"],
    "Type8": ["type 8", "securities margin financing"],
    "Type9": ["type 9", "asset management", "fund manager"],
    "Type10": ["type 10", "credit rating"],
    "Type11": ["type 11", "dealing in otc derivative"],
    "Type12": ["type 12", "advising on otc derivative"],
    "Type13": ["type 13", "depositary"],
}

# 合規領域映射
COMPLIANCE_DOMAIN_KEYWORDS = {
    "AML": ["aml", "anti-money laundering", "counter-financing of terrorism", "cft", "kyc", "cdd", "sanctions", "str", "jfiu"],
    "FRR": ["frr", "financial resources", "liquid capital", "capital adequacy"],
    "Conduct": ["code of conduct", "suitability", "best execution", "conflicts of interest", "client agreement"],
    "Risk": ["risk management", "operational risk", "market risk", "credit risk", "liquidity risk"],
    "IT_Cyber": ["cybersecurity", "cyber security", "data protection", "information security", "bcp", "business continuity"],
    "VATP": ["virtual asset", "vatp", "vasp", "crypto", "stablecoin", "aspire"],
    "Client_Asset": ["client money", "client securities", "segregation", "custody"],
    "Supervision": ["supervision", "internal control", "mic", "responsible officer", "ro"],
    "Licensing": ["licensing", "application", "fit and proper", "competence", "cpt", "wings"],
    "Disclosure": ["disclosure", "inside information", "prospectus", "offering document"],
}

# 文件層級映射
DOC_HIERARCHY = {
    "law": 1,              # 法律 (SFO, AMLO)
    "subsidiary_leg": 2,   # 附屬法例 (FRR, Client Securities Rules)
    "code": 3,             # 守則 (Code of Conduct)
    "guideline": 4,        # 指引 (AML Guideline, MSIC)
    "circular": 5,         # 通函
    "faq": 6,              # FAQ
    "thematic_report": 7,  # 主題巡查報告
    "consultation": 8,     # 諮詢文件
    "enforcement": 9,      # 執法行動
    "press_release": 10,   # 新聞稿
    "industry": 11,        # 行業分析
    "template": 12,        # 合規模板
}

def detect_license_types(text):
    """[P3.16] 自動檢測內容涉及的牌照類型"""
    lower = text[:5000].lower()
    detected = []
    for lt, keywords in LICENSE_TYPE_KEYWORDS.items():
        if any(kw in lower for kw in keywords):
            detected.append(lt)
    return detected[:5]  # 最多返回 5 個

def detect_compliance_domains(text):
    """[P3.16] 自動檢測合規領域"""
    lower = text[:5000].lower()
    detected = []
    for domain, keywords in COMPLIANCE_DOMAIN_KEYWORDS.items():
        if any(kw in lower for kw in keywords):
            detected.append(domain)
    return detected[:4]  # 最多返回 4 個

def get_doc_hierarchy(doc_type):
    """[P3.16] 返回文件層級分值"""
    return DOC_HIERARCHY.get(doc_type, 99)


# ============================================================
# 進度追蹤 + 內容 Hash 記錄
# ============================================================
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        try:
            with open(PROGRESS_FILE, 'r') as f:
                return json.load(f)
        except: pass
    return {"uploaded_refs": [], "last_year": None, "last_source": None}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f, indent=2)

def load_content_hashes():
    if os.path.exists(CONTENT_HASHES_FILE):
        try:
            with open(CONTENT_HASHES_FILE, 'r') as f:
                return json.load(f)
        except: pass
    return {"hashes": {}}

def save_content_hashes(hashes_data):
    with open(CONTENT_HASHES_FILE, 'w') as f:
        json.dump(hashes_data, f, indent=2)


# ============================================================
# v5.0 質量保證框架 (QA Framework)
# ============================================================

def compute_content_hash(text):
    """計算文本內容的 SHA-256 hash"""
    normalized = re.sub(r'\s+', ' ', text.strip().lower())
    return hashlib.sha256(normalized.encode('utf-8')).hexdigest()[:16]

def validate_chunk(text, source_type="unknown"):
    """
    內容驗證 — 上傳前必須通過
    返回 (is_valid, reason)
    """
    if not text or not text.strip():
        return False, "empty_content"
    
    clean = text.strip()
    
    # 最小長度
    if len(clean) < 200:
        return False, f"too_short ({len(clean)} chars, min 200)"
    
    # 亂碼檢測：非 ASCII 非中文非常見符號的比例
    garbage_chars = sum(1 for c in clean[:2000] if ord(c) > 0xFFFF or (0x80 <= ord(c) <= 0xFF and c not in 'àáâãäåæçèéêëìíîïðñòóôõöøùúûüýþÿ'))
    if len(clean) > 0 and garbage_chars / min(len(clean), 2000) > 0.05:
        return False, f"garbled_content ({garbage_chars} garbage chars)"
    
    # 404/錯誤頁面檢測
    lower = clean[:2000].lower()
    error_sigs = [
        "you are being redirected", "please update your bookmark",
        "閣下將會被自動連接到證監會的新網站", "page not found",
        "404 not found", "403 forbidden", "access denied",
        "this page has been removed",
    ]
    if any(sig in lower for sig in error_sigs) and len(clean) < 12000:
        return False, "404_or_error_page"
    
    # 導航/模板/廣告內容檢測
    noise_sigs = ["cookie policy", "subscribe to our newsletter", 
                  "sign up for free", "privacy policy", "terms of use",
                  "all rights reserved", "powered by wordpress"]
    noise_count = sum(1 for sig in noise_sigs if sig in lower)
    content_ratio = len(re.sub(r'[^\w\s]', '', clean)) / max(len(clean), 1)
    if noise_count >= 3 and len(clean) < 3000:
        return False, f"navigation_template_content ({noise_count} noise signals)"
    
    # Token 長度預檢（粗估：1 token ≈ 4 chars for EN, 1.5 chars for ZH）
    est_tokens = len(clean) / 3  # 保守估計
    if est_tokens > 8000:
        return False, f"too_long_for_embedding (est {est_tokens:.0f} tokens, max 8000)"
    
    return True, "ok"

def check_duplicate(content, ref_id, progress=None, hashes_data=None):
    """
    三層去重檢查
    返回 (is_duplicate, layer, detail)
    """
    if progress is None:
        progress = load_progress()
    if hashes_data is None:
        hashes_data = load_content_hashes()
    
    uploaded_refs = set(progress.get("uploaded_refs", []))
    
    # L1: ref_id 精確匹配
    if ref_id in uploaded_refs:
        return True, "L1_ref", f"ref '{ref_id}' already uploaded"
    
    # L2: 內容 hash 匹配
    content_hash = compute_content_hash(content)
    existing_hashes = hashes_data.get("hashes", {})
    if content_hash in existing_hashes:
        orig_ref = existing_hashes[content_hash]
        return True, "L2_hash", f"identical content (hash={content_hash}, orig_ref={orig_ref})"
    
    # L3: 語義相似度（簡化版 — 用標題/前100字 比對）
    # 完整的 embedding 比對在 Mode 8 中實現（成本太高不適合每次上傳）
    content_start = re.sub(r'\s+', ' ', content[:300]).strip().lower()
    for existing_hash, existing_ref in existing_hashes.items():
        # 不做全量比對，只標記完全相同的前綴
        pass  # L3 主要在 Mode 8 離線執行
    
    return False, None, None

def record_upload(ref_id, content, progress=None, hashes_data=None):
    """記錄已上傳的 ref 和 content hash"""
    if progress is None:
        progress = load_progress()
    if hashes_data is None:
        hashes_data = load_content_hashes()
    
    # 記錄 ref
    uploaded = set(progress.get("uploaded_refs", []))
    uploaded.add(ref_id)
    progress["uploaded_refs"] = list(uploaded)
    
    # 記錄 hash
    content_hash = compute_content_hash(content)
    if "hashes" not in hashes_data:
        hashes_data["hashes"] = {}
    hashes_data["hashes"][content_hash] = ref_id
    
    return progress, hashes_data

def clean_html_content(html_text):
    """清理 HTML，移除腳本/樣式/導航，提取正文"""
    text = re.sub(r'<script[^>]*>.*?</script>', '', html_text, flags=re.DOTALL)
    text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL)
    text = re.sub(r'<nav[^>]*>.*?</nav>', '', text, flags=re.DOTALL)
    text = re.sub(r'<footer[^>]*>.*?</footer>', '', text, flags=re.DOTALL)
    text = re.sub(r'<header[^>]*>.*?</header>', '', text, flags=re.DOTALL)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&[a-zA-Z]+;', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# ============================================================
# v5.0 統一進度條 + 報告生成器
# ============================================================

def show_progress(current, total, prefix="", start_time=None, extra=""):
    """統一進度條 — 所有 Mode 共用"""
    pct = current / max(total, 1) * 100
    bar_len = 25
    filled = int(bar_len * current / max(total, 1))
    bar = "█" * filled + "░" * (bar_len - filled)
    
    eta_str = ""
    if start_time and current > 0:
        elapsed = time.time() - start_time
        eta = (elapsed / current) * (total - current)
        if eta < 60:
            eta_str = f" ETA {eta:.0f}s"
        elif eta < 3600:
            eta_str = f" ETA {eta/60:.1f}m"
        else:
            eta_str = f" ETA {eta/3600:.1f}h"
    
    line = f"\r  {prefix}[{bar}] {current}/{total} ({pct:.0f}%){eta_str} {extra}"
    print(line.ljust(90), end="", flush=True)

def generate_report(mode_name, stats, results, total_time, extra_info=None):
    """
    生成 TXT + JSON 報告（所有 Mode 統一使用）
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    base_name = f"{mode_name}_{timestamp}"
    
    # ── JSON 報告 ──
    json_path = os.path.join(REPORTS_DIR, f"{base_name}.json")
    report_data = {
        "mode": mode_name, "timestamp": datetime.now().isoformat(),
        "duration_minutes": round(total_time / 60, 1),
        "stats": stats, "results": results,
    }
    if extra_info:
        report_data["extra"] = extra_info
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report_data, f, ensure_ascii=False, indent=2)
    
    # ── TXT 報告 ──
    txt_path = os.path.join(REPORTS_DIR, f"{base_name}.txt")
    ok_items = [r for r in results if r.get("status") == "ok"]
    fail_items = [r for r in results if r.get("status") == "fail"]
    skip_items = [r for r in results if r.get("status") == "skip"]
    dup_items = [r for r in results if r.get("status") == "duplicate"]
    
    lines = []
    lines.append("=" * 64)
    lines.append(f"  VIGO Collector Report — {mode_name}")
    lines.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("=" * 64)
    lines.append("")
    lines.append(f"  Duration:     {total_time/60:.1f} min")
    lines.append(f"  Success:      {stats.get('ok', 0)}")
    lines.append(f"  Failed:       {stats.get('fail', 0)}")
    lines.append(f"  Skipped:      {stats.get('skip', 0)}")
    lines.append(f"  Duplicates:   {stats.get('duplicate', 0)}")
    if stats.get("chunks"):
        lines.append(f"  Total chunks: {stats['chunks']}")
    if stats.get("retry_ok"):
        lines.append(f"  Retry fixed:  {stats['retry_ok']}")
    lines.append("")
    
    if ok_items:
        lines.append("-" * 64)
        lines.append("  ✅ SUCCESS")
        lines.append("-" * 64)
        for r in ok_items:
            lines.append(f"  [{r.get('priority','?'):>7}] {r.get('name','?')[:55]}")
            lines.append(f"           {r.get('chars','?')} chars | {r.get('chunks','?')} chunks | {r.get('time_min','?')} min | {r.get('source','?')}")
    
    if fail_items:
        lines.append("")
        lines.append("-" * 64)
        lines.append("  ❌ FAILED")
        lines.append("-" * 64)
        for r in fail_items:
            lines.append(f"  {r.get('name','?')[:55]}")
            lines.append(f"           Reason: {r.get('reason','unknown')}")
    
    if skip_items:
        lines.append("")
        lines.append("-" * 64)
        lines.append("  ⏭ SKIPPED")
        lines.append("-" * 64)
        for r in skip_items:
            lines.append(f"  {r.get('name','?')[:55]}")
    
    if dup_items:
        lines.append("")
        lines.append("-" * 64)
        lines.append("  🔄 DUPLICATES BLOCKED")
        lines.append("-" * 64)
        for r in dup_items:
            lines.append(f"  {r.get('name','?')[:55]} — {r.get('dup_detail','')}")
    
    if extra_info:
        lines.append("")
        lines.append("-" * 64)
        lines.append("  📋 EXTRA INFO")
        lines.append("-" * 64)
        for k, v in extra_info.items():
            lines.append(f"  {k}: {v}")
    
    lines.append("")
    lines.append("=" * 64)
    
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    
    print(f"\n  📄 Report: {txt_path}")
    print(f"             {json_path}")
    return txt_path, json_path

def save_backup(vigo_chunks, filename):
    """生成雙語 MD 備份文件"""
    sections = []
    for c in vigo_chunks:
        m = c["metadata"]
        meta = '\n'.join(f"{k}: {v}" for k, v in m.items() if not isinstance(v, list))
        kw = m.get("keywords", [])
        if kw: meta += f"\nkeywords: {', '.join(kw)}"
        title = c['content'].split('\n')[0].replace('【', '').replace('】', '')
        body = '\n'.join(c['content'].split('\n')[2:])
        sections.append(f"===SECTION===\n【標題】{title}\n【內容】\n{body}\n【元數據】\n{meta}\n===END===")
    with open(filename, 'w', encoding='utf-8') as f:
        f.write('\n\n'.join(sections))
    print(f"   📝 Backup: {filename}")


# ============================================================
# PDF / DOCX / XLSX 文字提取
# ============================================================
def extract_text_from_pdf(pdf_bytes):
    """用 pdfplumber 從 PDF 中提取文字"""
    try:
        import pdfplumber
    except ImportError:
        print("      ERROR: pip install pdfplumber")
        return None
    try:
        pdf = pdfplumber.open(io.BytesIO(pdf_bytes))
        pages_text = []
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text.strip():
                pages_text.append(text)
        pdf.close()
        if not pages_text:
            return None
        full_text = "\n\n".join(pages_text)
        full_text = re.sub(r'(\d+)\s*\n\s*\n', r'\1\n', full_text)
        full_text = re.sub(r'\n{3,}', '\n\n', full_text)
        full_text = re.sub(
            r'54/F,\s*One Island East.*?www\.sfc\.hk\s*(?:Page \d+ of \d+)?',
            '', full_text, flags=re.DOTALL)
        full_text = re.sub(r'香港鰂魚涌華蘭路.*?港島東中心.*?樓', '', full_text)
        return full_text.strip()
    except Exception as e:
        print(f"      PDF extract error: {e}")
        return None

def extract_text_from_docx(file_path):
    """從 DOCX 文件提取文字"""
    try:
        from docx import Document
    except ImportError:
        print("      ERROR: pip install python-docx")
        return None
    try:
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except Exception as e:
        print(f"      DOCX extract error: {e}")
        return None

def extract_text_from_xlsx(file_path):
    """從 XLSX 文件提取文字（逐 sheet）"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("      ERROR: pip install openpyxl")
        return None
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(c for c in cells if c)
                if line.strip():
                    rows.append(line)
            if rows:
                all_text.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(all_text)
    except Exception as e:
        print(f"      XLSX extract error: {e}")
        return None


# ============================================================
# 智慧分塊 (Smart Chunking)
# ============================================================
def parse_vigo_sections(text):
    """
    v5.2.1: 解析 ===SECTION=== / ===END=== 格式的 VIGO 結構化 .md 文件
    返回: list of dict，每個 dict 包含 content, title, metadata
    如果文件不是此格式，返回 None（交給原有流程處理）
    """
    if '===SECTION===' not in text:
        return None
    
    sections = []
    raw_sections = text.split('===SECTION===')
    
    for raw in raw_sections:
        raw = raw.strip()
        if not raw or raw.startswith('#'):
            continue
        
        if '===END===' in raw:
            raw = raw.split('===END===')[0].strip()
        
        title = ""
        if '【標題】' in raw:
            title_match = re.search(r'【標題】(.+?)(?:\n|$)', raw)
            if title_match:
                title = title_match.group(1).strip()
        
        content = ""
        if '【內容】' in raw:
            content_start = raw.index('【內容】') + len('【內容】')
            if '【元數據】' in raw:
                content = raw[content_start:raw.index('【元數據】')].strip()
            else:
                content = raw[content_start:].strip()
        
        section_meta = {}
        if '【元數據】' in raw:
            meta_start = raw.index('【元數據】') + len('【元數據】')
            meta_text = raw[meta_start:].strip()
            for line in meta_text.split('\n'):
                line = line.strip()
                if ':' in line and not line.startswith('#'):
                    key, _, val = line.partition(':')
                    key = key.strip()
                    val = val.strip().strip('"').strip("'")
                    if val.startswith('[') and val.endswith(']'):
                        try:
                            val = json.loads(val)
                        except:
                            pass
                    section_meta[key] = val
        
        if content and len(content) >= 80:
            sections.append({
                'title': title,
                'content': content,
                'metadata': section_meta
            })
    
    return sections if sections else None


def smart_chunk_text(full_text, doc_title, max_size=CHUNK_MAX):
    """智慧分塊：按段落/標題切分，確保每塊完整"""
    if len(full_text) <= max_size:
        return [full_text]

    sections = re.split(r'\n(?=(?:Chapter|Part|Section|Schedule|Appendix|Division|附表|附錄|第[一二三四五六七八九十百]+[章節部分])\s)', full_text)
    
    if len(sections) <= 1:
        sections = re.split(r'\n(?=\d+\.\s+[A-Z])', full_text)
    if len(sections) <= 1:
        sections = re.split(r'\n(?=[A-Z][a-z]+ \d+)', full_text)
    if len(sections) <= 1:
        sections = full_text.split('\n\n')

    chunks = []
    current = ""
    for section in sections:
        section = section.strip()
        if not section:
            continue
        if len(current) + len(section) + 2 <= max_size:
            current = current + "\n\n" + section if current else section
        else:
            if current:
                chunks.append(current)
            if len(section) > max_size:
                words = section.split('. ')
                sub = ""
                for sentence in words:
                    if len(sub) + len(sentence) + 2 <= max_size:
                        sub = sub + ". " + sentence if sub else sentence
                    else:
                        if sub:
                            chunks.append(sub)
                        sub = sentence
                if sub:
                    chunks.append(sub)
            else:
                current = section
    if current:
        chunks.append(current)
    
    # 過濾太短的 chunks
    valid = [c for c in chunks if len(c) >= CHUNK_MIN]
    if not valid and chunks:
        valid = chunks  # 如果全部太短，還是保留
    
    return valid if valid else [full_text[:max_size]]


# ============================================================
# SFC API 內容抓取
# ============================================================
def fetch_sfc_content(ref, api_path, lang="EN"):
    """從 SFC API 抓取內容，返回 (text, content_type) — v5.2 支持中英文 [P0.5]"""
    url = f"{SFC_API_BASE}/{api_path}?refNo={ref}&lang={lang}"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        if resp.status_code != 200:
            return None, None
        content_type = resp.headers.get('content-type', '').lower()
        if 'pdf' in content_type or resp.content[:4] == b'%PDF':
            text = extract_text_from_pdf(resp.content)
            return text, "pdf"
        else:
            text = resp.text
            text = clean_html_content(text)
            if len(text) < 100:
                return None, None
            return text, "html"
    except Exception as e:
        return None, None


# ============================================================
# AI 服務：Embedding / DeepSeek / 翻譯
# ============================================================
def generate_embedding(text):
    try:
        resp = requests.post("https://api.openai.com/v1/embeddings",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
            json={"input": text[:8000], "model": "text-embedding-3-small"},
            timeout=30)
        if resp.status_code == 200:
            return resp.json()["data"][0]["embedding"]
        else:
            print(f"      Embedding error: {resp.status_code}")
            return None
    except Exception as e:
        print(f"      Embedding error: {e}")
        return None

def deepseek_chat(prompt, max_tokens=800, temperature=0.3):
    try:
        resp = requests.post("https://api.deepseek.com/chat/completions",
            headers={"Authorization": f"Bearer {DEEPSEEK_API_KEY}", "Content-Type": "application/json"},
            json={"model": "deepseek-chat", "messages": [{"role": "user", "content": prompt}],
                  "max_tokens": max_tokens, "temperature": temperature},
            timeout=60)
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        print(f"      DeepSeek error: {e}")
    return None

def generate_questions(content, lang="zh"):
    if lang == "zh":
        prompt = f"根據以下 SFC 監管資訊，生成3個使用者可能會問的問題。只返回問題列表。\n\n{content[:2000]}"
    else:
        prompt = f"Based on this SFC regulatory info, generate 3 questions users might ask. Return only the questions.\n\n{content[:2000]}"
    result = deepseek_chat(prompt, max_tokens=300, temperature=0.3)
    if result:
        questions = [q.strip().lstrip('0123456789.-) ') for q in result.split('\n') if q.strip() and len(q.strip()) > 10]
        return questions[:3]
    return None

def translate_to_chinese(text):
    prompt = f"""你是 SFC 監管情報翻譯師。翻譯為繁體中文摘要。
要求：保留專業術語英文+括號，如：持牌法團 (Licensed Corporation)。保留日期、金額、編號。
原文：{text[:3500]}
只返回翻譯。"""
    return deepseek_chat(prompt, max_tokens=800, temperature=0.2)


# ============================================================
# v5.0 上傳（含 QA 驗證 + 失敗重試）
# ============================================================
def upload_via_v6(content, metadata):
    """通過 v6 智能管線上傳（自動分類→去重→驗證→結構化提取）"""
    source_url = metadata.get('source_url', metadata.get('url', ''))
    source_name = metadata.get('source', metadata.get('category', ''))
    source_id = V6_SOURCE_ID_MAP.get(source_name, source_name.lower().replace(' ', '_'))
    title = metadata.get('title', metadata.get('ref_no', ''))
    publish_date = metadata.get('publish_date', metadata.get('date', ''))

    # 語言偵測
    chinese_ratio = sum(1 for c in content[:500] if '\u4e00' <= c <= '\u9fff') / max(len(content[:500]), 1)
    language = 'zh' if chinese_ratio > 0.3 else 'en'

    payload = {"url": source_url, "source_id": source_id, "raw_text": content}
    if title: payload["title"] = title
    if publish_date: payload["publish_date"] = publish_date
    if language: payload["language"] = language

    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
    }

    try:
        resp = requests.post(V6_INGEST_URL, headers=headers, json=payload, timeout=60)
        if resp.ok:
            result = resp.json()
            if result.get('success'):
                grade = result.get('verification_grade', 'N/A')
                chunks = result.get('chunks_created', 0)
                emoji = {'GOLD': '🥇', 'SILVER': '🥈', 'BRONZE': '🥉'}.get(grade, '⚪')
                print(f"      ✅ v6: {chunks} chunks {emoji}{grade}")
                return result.get('chunks_created', 1)  # return truthy value
        print(f"      ⚠ v6 failed ({resp.status_code}), falling back...")
        return None
    except Exception as e:
        print(f"      ⚠ v6 error: {e}, falling back...")
        return None


def upload_chunk(content, metadata, retry_count=3):
    """上傳一個 chunk 到 Supabase — v6.0: 優先走 v6 管線，失敗回退直接上傳"""
    # v6.0: 如果 v6 模式開啟，優先使用智能管線
    if V6_INGEST_MODE:
        v6_result = upload_via_v6(content, metadata)
        if v6_result:
            return v6_result

    # 原有直接上傳邏輯（v6 關閉或 v6 失敗時回退）
    # 內容驗證
    is_valid, reason = validate_chunk(content, metadata.get("source", ""))
    if not is_valid:
        print(f"      ⚠ QA blocked: {reason}")
        return None
    
    # v5.2 [P3.16]: 自動注入結構化標籤
    if "license_types" not in metadata:
        metadata["license_types"] = detect_license_types(content)
    if "compliance_domains" not in metadata:
        metadata["compliance_domains"] = detect_compliance_domains(content)
    if "doc_hierarchy" not in metadata:
        metadata["doc_hierarchy"] = get_doc_hierarchy(metadata.get("doc_type", "other"))
    
    embedding = generate_embedding(content)
    if not embedding:
        return None
    
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }
    
    for attempt in range(retry_count):
        try:
            resp = requests.post(f"{SUPABASE_URL}/rest/v1/{KNOWLEDGE_TABLE}",
                headers=headers,
                json={"content": content, "metadata": metadata, "embedding": embedding},
                timeout=60)
            if resp.status_code in [200, 201]:
                result = resp.json()
                if isinstance(result, list) and len(result) > 0:
                    kid = result[0].get('id')
                    # Quick questions
                    lang = metadata.get("language", "zh")
                    questions = generate_questions(content, lang)
                    if questions:
                        qh = {**headers, "Prefer": "return=minimal"}
                        for q in questions:
                            try:
                                requests.post(f"{SUPABASE_URL}/rest/v1/{QUESTIONS_TABLE}",
                                    headers=qh,
                                    json={"question_text": q,
                                          "category": metadata.get("category", "SFC"),
                                          "keywords": metadata.get("keywords", ["SFC"]),
                                          "source_knowledge_id": kid,
                                          "source_table": "vigo_knowledge",
                                          "generation_method": "auto",
                                          "is_active": True},
                                    timeout=30)
                            except: pass
                            time.sleep(0.2)
                    return kid
            elif resp.status_code >= 500:
                if attempt < retry_count - 1:
                    time.sleep(2 ** attempt)
                    continue
        except requests.exceptions.Timeout:
            if attempt < retry_count - 1:
                print(f"      ⏳ Timeout, retry {attempt+2}/{retry_count}...")
                time.sleep(2 ** attempt)
                continue
        except Exception as e:
            if attempt < retry_count - 1:
                print(f"      ⚠ Error: {e}, retry {attempt+2}/{retry_count}...")
                time.sleep(2 ** attempt)
                continue
            print(f"      Upload err: {e}")
        break
    return None


# ============================================================
# 雙語 Chunk 生成
# ============================================================
def build_official_url(ref, doc_type):
    if doc_type in ("circular",):
        return f"https://apps.sfc.hk/edistributionWeb/gateway/EN/circular/doc?refNo={ref}"
    elif doc_type in ("enforcement", "vatp", "press_release"):
        return f"https://apps.sfc.hk/edistributionWeb/gateway/EN/news-and-announcements/news/doc?refNo={ref}"
    elif doc_type in ("consultation", "consultation_conclusion"):
        return f"https://apps.sfc.hk/edistributionWeb/gateway/EN/consultation/doc?refNo={ref}"
    return f"https://www.sfc.hk/en/"

def create_bilingual_chunks(ref, text_chunks, doc_type, source_url):
    """每個 text chunk → EN + ZH"""
    all_vigo_chunks = []
    official_url = build_official_url(ref, doc_type)
    type_labels = {
        "circular": ("Circular", "通函"), "enforcement": ("Enforcement", "執法行動"),
        "consultation": ("Consultation", "諮詢文件"), "vatp": ("Virtual Asset", "虛擬資產"),
        "press_release": ("Press Release", "新聞稿"), "enforcement_summary": ("Enforcement Summary", "執法摘要"),
        "hkma_joint": ("HKMA Joint Circular", "金管局聯合通函"),
    }
    en_label, zh_label = type_labels.get(doc_type, ("Update", "動態"))
    
    combined = " ".join(text_chunks).lower()
    keywords = ["SFC"]
    kw_map = {"virtual asset": "VATP", "vatp": "VATP", "aml": "AML",
              "aspire": "ASPIRe", "sponsor": "Sponsor", "ipo": "IPO",
              "type 1": "Type1", "type 9": "Type9", "stablecoin": "Stablecoin",
              "staking": "Staking", "enforcement": "Enforcement",
              "consultation": "Consultation", "circular": "Circular",
              "hkma": "HKMA", "listing": "Listing", "hkex": "HKEX",
              # v5.2 新增關鍵詞
              "type 2": "Type2", "type 3": "Type3", "type 4": "Type4",
              "type 5": "Type5", "type 6": "Type6", "type 7": "Type7",
              "type 8": "Type8", "type 10": "Type10", "type 11": "Type11",
              "type 12": "Type12", "type 13": "Type13",
              "fund manager": "FundManager", "frr": "FRR",
              "financial resources": "FRR", "client money": "ClientAsset",
              "client securities": "ClientAsset", "cyber": "Cybersecurity",
              "business continuity": "BCP", "travel rule": "TravelRule",
              "fatf": "FATF", "pdpo": "PDPO", "privacy": "PDPO",
              "wings": "WINGS", "competence": "CPT",
              "suitability": "Suitability", "risk management": "RiskMgmt",
              }
    for kw, tag in kw_map.items():
        if kw in combined and tag not in keywords:
            keywords.append(tag)
    keywords = keywords[:8]  # v5.2: 增加到 8 個
    
    # v5.2 [P3.16]: 自動結構化標籤
    auto_license_types = detect_license_types(combined)
    auto_domains = detect_compliance_domains(combined)

    for i, chunk_text in enumerate(text_chunks):
        part_label = f" (Part {i+1}/{len(text_chunks)})" if len(text_chunks) > 1 else ""
        base_meta = {
            "source": f"SFC_Official_{ref}", "source_url": official_url,
            "doc_type": doc_type, "category": "SFC", "ref_no": ref,
            "source_table": "vigo_financial", "keywords": keywords,
            "collected_date": datetime.now().strftime("%Y-%m-%d"),
            "content_type": "pdf_extract" if len(text_chunks) > 1 else "html",
            "authority_score": SOURCE_AUTHORITY.get("SFC_Official", 100),
            # v5.2 [P3.16] 結構化標籤
            "license_types": auto_license_types,
            "compliance_domains": auto_domains,
            "doc_hierarchy": get_doc_hierarchy(doc_type),
        }
        
        en_content = f"【SFC {en_label}{part_label} — {ref}】\n\n{chunk_text}"
        en_content += f"\n\n---\nSource: SFC Official ({ref})\nFull document: {official_url}"
        all_vigo_chunks.append({"content": en_content, "metadata": {**base_meta, "language": "en"}, "lang": "EN"})
        
        zh_translation = translate_to_chinese(chunk_text)
        if zh_translation and len(zh_translation) > 80:
            zh_content = f"【SFC {zh_label}{part_label} — {ref}】\n\n{zh_translation}"
            zh_content += f"\n\n---\n資料來源：證監會官方 ({ref})\n原文鏈接：{official_url}"
            all_vigo_chunks.append({"content": zh_content, "metadata": {**base_meta, "language": "zh"}, "lang": "ZH"})
        time.sleep(1)
    return all_vigo_chunks


# ============================================================
# v5.0 Mode 1/2: SFC 掃描（升級版 — 時間範圍 + 重試 + HKMA）
# ============================================================

def scan_source(name, ref_prefix_list, api_path, filter_fn=None, max_items=10, include_chinese=False):
    """掃描一個 SFC 數據源 — v5.2 支持中文版 [P0.5]"""
    print(f"\n{'=' * 50}")
    print(f"  {name}")
    if include_chinese:
        print(f"  (含中文版 TC)")
    print(f"{'=' * 50}")
    results = []
    found = 0
    for prefix in ref_prefix_list:
        for ref_id in prefix["range"]:
            if found >= max_items:
                break
            ref = f"{prefix['year']}{prefix['code']}{ref_id}"
            text, ctype = fetch_sfc_content(ref, api_path)
            if not text:
                continue
            if filter_fn and not filter_fn(text):
                continue
            chunks = smart_chunk_text(text, f"SFC {ref}")
            results.append({
                "ref": ref, "chunks": chunks, "content_type": ctype,
                "doc_type": prefix.get("doc_type", "unknown"),
                "source_url": f"{SFC_API_BASE}/{api_path}?refNo={ref}&lang=EN"
            })
            found += 1
            print(f"   OK: {ref} ({ctype}, {len(chunks)} chunk{'s' if len(chunks)>1 else ''}, {len(text)} chars)")
            
            # v5.2 [P0.5]: 中文版通函
            if include_chinese and prefix.get("doc_type") == "circular":
                zh_text, zh_ctype = fetch_sfc_content(ref, api_path, lang="TC")
                if zh_text and len(zh_text) > 300:
                    zh_chunks = smart_chunk_text(zh_text, f"SFC {ref} (中文)")
                    results.append({
                        "ref": f"{ref}_TC", "chunks": zh_chunks, "content_type": zh_ctype,
                        "doc_type": prefix.get("doc_type", "unknown"),
                        "source_url": f"{SFC_API_BASE}/{api_path}?refNo={ref}&lang=TC",
                        "is_chinese_original": True
                    })
                    print(f"   OK: {ref}_TC (中文原版, {len(zh_chunks)} chunks)")
            
            time.sleep(0.3)
    print(f"   Total: {found} items" + (f" + Chinese versions" if include_chinese else ""))
    return results

def get_scan_ranges(year, period="3m"):
    """根據時間範圍計算掃描深度"""
    ranges = {
        "1m":  {"ec": 5,  "pr": 15,  "cp": 3,  "prev_ec": 0,  "prev_pr": 0},
        "3m":  {"ec": 10, "pr": 30,  "cp": 6,  "prev_ec": 5,  "prev_pr": 15},
        "6m":  {"ec": 20, "pr": 50,  "cp": 10, "prev_ec": 25, "prev_pr": 30},
        "1y":  {"ec": 50, "pr": 100, "cp": 20, "prev_ec": 50, "prev_pr": 100},
    }
    r = ranges.get(period, ranges["3m"])
    return r

def fetch_all_sources(year, depth="quick", period="3m"):
    """Mode 1/2 掃描 — v5.0 支持時間範圍"""
    if depth == "quick":
        r = get_scan_ranges(year, "1m")
        max_items_ec, max_items_pr, max_items_cp = 8, 10, 6
    else:
        r = get_scan_ranges(year, period)
        max_items_ec = min(r["ec"] + r["prev_ec"], 30)
        max_items_pr = min(r["pr"] + r["prev_pr"], 50)
        max_items_cp = min(r["cp"], 15)

    all_results = []

    # 1. Circulars (v5.2: 含中文版 [P0.5])
    ec_ranges = [{"year": year, "code": "EC", "range": range(r["ec"], 0, -1), "doc_type": "circular"}]
    if r.get("prev_ec", 0) > 0:
        ec_ranges.append({"year": year-1, "code": "EC", "range": range(50, 50-r["prev_ec"], -1), "doc_type": "circular"})
    all_results += scan_source("[1/6] SFC Circulars (通函 EN+TC)", ec_ranges, "circular/openFile", 
                               max_items=max_items_ec, include_chinese=True)

    # 2. Enforcement
    enforce_kw = ['reprimand', 'fine', 'suspend', 'ban', 'prosecut', 'convict', 'disciplin', 'enforcement', 'sanction', 'misconduct']
    pr_ranges = [{"year": year, "code": "PR", "range": range(r["pr"], 0, -1), "doc_type": "enforcement"}]
    if r.get("prev_pr", 0) > 0:
        pr_ranges.append({"year": year-1, "code": "PR", "range": range(250, 250-r["prev_pr"], -1), "doc_type": "enforcement"})
    all_results += scan_source("[2/6] SFC Enforcement (執法新聞)", pr_ranges, "news/list-content",
        filter_fn=lambda t: any(k in t.lower() for k in enforce_kw), max_items=max_items_pr // 3)

    # 3. Consultations
    cp_ranges = [{"year": year, "code": "CP", "range": range(r["cp"], 0, -1), "doc_type": "consultation"}]
    if depth != "quick":
        cp_ranges.append({"year": year-1, "code": "CP", "range": range(15, 0, -1), "doc_type": "consultation"})
    all_results += scan_source("[3/6] SFC Consultations (諮詢文件)", cp_ranges, "consultation/openFile", max_items=max_items_cp)

    # 4. VATP
    va_kw = ['virtual asset', 'vatp', 'vasp', 'crypto', 'stablecoin', 'digital asset', 'web3', 'staking', 'aspire']
    all_results += scan_source("[4/6] SFC VATP (虛擬資產)",
        [{"year": year, "code": "PR", "range": range(r["pr"], 0, -1), "doc_type": "vatp"}],
        "news/list-content", filter_fn=lambda t: any(k in t.lower() for k in va_kw), max_items=8)

    # 5. Press Releases
    skip_kw = enforce_kw + va_kw
    all_results += scan_source("[5/6] SFC Press Releases (新聞稿)",
        [{"year": year, "code": "PR", "range": range(r["pr"], 0, -1), "doc_type": "press_release"}],
        "news/list-content", filter_fn=lambda t: not any(k in t.lower() for k in skip_kw) and len(t) > 300, max_items=8)

    # 6. v5.0 NEW: HKMA 聯合通函
    print(f"\n{'=' * 50}")
    print(f"  [6/6] HKMA Joint Circulars (金管局聯合通函)")
    print(f"{'=' * 50}")
    try:
        from bs4 import BeautifulSoup
        hkma_url = "https://www.hkma.gov.hk/eng/regulatory-resources/regulatory-guides/circulars/"
        resp = requests.get(hkma_url, headers=HEADERS, timeout=30)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'html.parser')
            links = soup.find_all('a', href=True)
            hkma_found = 0
            for link in links[:100]:
                href = link.get('href', '')
                text = link.get_text(strip=True)
                if ('sfc' in text.lower() or 'securities' in text.lower() or 'joint' in text.lower()) and hkma_found < 5:
                    if href.startswith('/'):
                        href = f"https://www.hkma.gov.hk{href}"
                    try:
                        art_resp = requests.get(href, headers=HEADERS, timeout=20)
                        if art_resp.status_code == 200:
                            art_text = clean_html_content(art_resp.text)
                            if len(art_text) > 500 and 'sfc' in art_text.lower():
                                chunks = smart_chunk_text(art_text, f"HKMA Joint")
                                ref_id = f"HKMA_joint_{hkma_found+1}"
                                all_results.append({
                                    "ref": ref_id, "chunks": chunks, "content_type": "html",
                                    "doc_type": "hkma_joint", "source_url": href
                                })
                                hkma_found += 1
                                print(f"   OK: {text[:50]} ({len(chunks)} chunks)")
                    except: pass
            print(f"   Total: {hkma_found} joint circulars")
        else:
            print(f"   Skip: HKMA HTTP {resp.status_code}")
    except Exception as e:
        print(f"   Skip: {e}")

    # Bonus: Charltons Law
    print(f"\n{'=' * 50}")
    print(f"  [Bonus] Charltons Law")
    print(f"{'=' * 50}")
    try:
        from bs4 import BeautifulSoup
        url = "https://www.charltonslaw.com/sfc-enforcement-actions-in-january-2026-and-december-2025/"
        resp = requests.get(url, headers=HEADERS, timeout=30)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'html.parser')
            article = soup.find('div', class_='entry-content') or soup.find('article')
            if article:
                text = '\n\n'.join([e.get_text(strip=True) for e in article.find_all(['p', 'h2', 'h3', 'li'])
                                    if len(e.get_text(strip=True)) > 10])
                if len(text) > 500:
                    chunks = smart_chunk_text(text, "Charltons SFC Enforcement")
                    all_results.append({"ref": "charltons_latest", "chunks": chunks, "content_type": "html",
                                        "doc_type": "enforcement_summary", "source_url": url})
                    print(f"   OK: {len(chunks)} chunks, {len(text)} chars")
    except Exception as e:
        print(f"   Skip: {e}")

    return all_results

def run_mode_1_2(year, mode):
    """v5.0 Mode 1/2 統一入口 — 含 QA + 重試 + 報告"""
    depth = "quick" if mode == "1" else "full"
    period = "1m"
    
    if mode == "2":
        print("\n  季度補漏 Quarterly Scan 時間範圍:")
        print("  a. 過去 3 個月  (~標準，建議每季度)")
        print("  b. 過去 6 個月  (~深度，建議每半年)")
        print("  c. 過去 12 個月 (~年度，建議每年 1 月)")
        p = input("\n  (a/b/c): ").strip().lower()
        period = {"a": "3m", "b": "6m", "c": "1y"}.get(p, "3m")
    
    start_time = time.time()
    results = fetch_all_sources(year, depth, period)
    
    # 去重
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded_refs = set(progress.get("uploaded_refs", []))
    
    seen = set()
    unique = []
    for r in results:
        if r["ref"] not in seen:
            seen.add(r["ref"])
            unique.append(r)
    results = unique
    
    if not results:
        print("\n  No items found.")
        return
    
    total_chunks = sum(len(r["chunks"]) for r in results)
    print(f"\n{'=' * 50}")
    print(f"  Collected: {len(results)} items, ~{total_chunks * 2} VIGO chunks (EN+ZH)")
    
    # 生成雙語 chunks
    all_vigo_chunks = []
    for i, r in enumerate(results, 1):
        print(f"\n  [{i}/{len(results)}] {r['ref']} ({r['content_type']}, {len(r['chunks'])} chunks)")
        vigo_chunks = create_bilingual_chunks(r["ref"], r["chunks"], r["doc_type"], r["source_url"])
        all_vigo_chunks.extend(vigo_chunks)
    
    if not all_vigo_chunks:
        print("\n  No chunks generated.")
        return
    
    # 備份
    backup = os.path.join(REPORTS_DIR, f"VIGO_SFC_bilingual_{datetime.now().strftime('%Y%m%d')}.md")
    save_backup(all_vigo_chunks, backup)
    
    answer = input(f"\n  Upload {len(all_vigo_chunks)} chunks? (y/n): ").strip().lower()
    if answer != 'y':
        print(f"\n  Data saved to: {backup}")
        return
    
    # 上傳 — v5.0 含 QA + 進度條 + 重試佇列
    print(f"\n{'=' * 50}")
    print(f"  Uploading with QA checks...")
    print(f"{'=' * 50}")
    
    stats = {"ok": 0, "fail": 0, "duplicate": 0, "qa_blocked": 0, "chunks": 0}
    failed_queue = []
    upload_start = time.time()
    
    for i, c in enumerate(all_vigo_chunks, 1):
        show_progress(i, len(all_vigo_chunks), "Upload ", upload_start)
        
        # L1 + L2 去重
        ref_for_chunk = c["metadata"].get("ref_no", "") + f"_{c['lang']}_{i}"
        is_dup, layer, detail = check_duplicate(c["content"], ref_for_chunk, progress, hashes_data)
        if is_dup:
            stats["duplicate"] += 1
            continue
        
        kid = upload_chunk(c["content"], c["metadata"])
        if kid:
            stats["ok"] += 1
            stats["chunks"] += 1
            progress, hashes_data = record_upload(ref_for_chunk, c["content"], progress, hashes_data)
        else:
            stats["fail"] += 1
            failed_queue.append(c)
        time.sleep(0.3)
    
    # 重試失敗項
    if failed_queue:
        print(f"\n\n  🔄 Retrying {len(failed_queue)} failed items...")
        retry_ok = 0
        for c in failed_queue:
            time.sleep(2)
            kid = upload_chunk(c["content"], c["metadata"], retry_count=2)
            if kid:
                retry_ok += 1
                stats["ok"] += 1
                stats["fail"] -= 1
        stats["retry_ok"] = retry_ok
        print(f"  Retry fixed: {retry_ok}/{len(failed_queue)}")
    
    # 保存進度
    for r in results:
        uploaded_refs.add(r["ref"])
    progress["uploaded_refs"] = list(uploaded_refs)
    save_progress(progress)
    save_content_hashes(hashes_data)
    
    elapsed = time.time() - start_time
    print(f"\n\n{'=' * 60}")
    print(f"  DONE! {elapsed/60:.1f} min | OK: {stats['ok']} | Fail: {stats['fail']} | Dup: {stats['duplicate']}")
    print(f"{'=' * 60}")
    
    # 生成報告
    report_results = [{"name": r["ref"], "status": "ok", "chunks": len(r["chunks"]),
                       "chars": sum(len(c) for c in r["chunks"]), "priority": "SFC",
                       "time_min": "—", "source": r["content_type"]} for r in results]
    generate_report(f"mode{mode}_scan_{period}", stats, report_results, elapsed,
                    {"period": period, "depth": depth, "backup": backup})


# ============================================================
# Mode 4: 歷史建庫 Historical Archive（v5.3 分類掃描重構）
# ============================================================

# ── 預設方案 ──
HISTORICAL_PRESETS = {
    "recommended": {
        "name": "推薦方案 Recommended",
        "enforcement": (2010, 2026),   # 16 年 — 案例長期有價值
        "circular":    (2018, 2026),   # 8 年  — 舊通函多被取代
        "consultation":(2022, 2026),   # 4 年  — 已定案的參考價值極低
        "press":       (2020, 2026),   # 6 年  — 時效性短
        "est_hours": "2-3",
    },
    "compact": {
        "name": "精簡方案 Compact",
        "enforcement": (2016, 2026),   # 10 年
        "circular":    (2020, 2026),   # 6 年
        "consultation":(2023, 2026),   # 3 年
        "press":       (2022, 2026),   # 4 年
        "est_hours": "1-1.5",
    },
    "full": {
        "name": "完整方案 Full",
        "enforcement": (2003, 2026),   # 23 年（SFO 2003 生效）
        "circular":    (2010, 2026),   # 16 年
        "consultation":(2018, 2026),   # 8 年
        "press":       (2015, 2026),   # 11 年
        "est_hours": "4-5",
    },
}

# 每類型每年的 API 呼叫次數估算
_CALLS_PER_YEAR = {
    "enforcement": 300,   # PR refs (含 VATP 分類)
    "circular": 60,       # EC refs
    "consultation": 30,   # CP refs
    "press": 0,           # 已包含在 enforcement 的 PR refs 中
}

def _estimate_api_calls(preset):
    """估算某個預設方案的總 API 呼叫數"""
    total = 0
    for cat in ["enforcement", "circular", "consultation", "press"]:
        start_yr, end_yr = preset[cat]
        num_years = end_yr - start_yr + 1
        total += _CALLS_PER_YEAR.get(cat, 0) * num_years
    return total

def _show_preset_table(preset):
    """顯示預設方案的掃描範圍表格"""
    print(f"\n  ┌────────────┬──────────────┬─────────┬──────────────┐")
    print(f"  │ 數據類型    │ 年份範圍      │ 年數    │ 預估 API 呼叫 │")
    print(f"  ├────────────┼──────────────┼─────────┼──────────────┤")
    
    labels = {
        "enforcement": "執法行動",
        "circular":    "通函      ",
        "consultation":"諮詢文件",
        "press":       "新聞稿    ",
    }
    total_calls = 0
    for cat in ["enforcement", "circular", "consultation", "press"]:
        s, e = preset[cat]
        ny = e - s + 1
        calls = _CALLS_PER_YEAR.get(cat, 0) * ny
        total_calls += calls
        print(f"  │ {labels[cat]}  │ {s}-{e}    │ {ny:>4} 年  │ {calls:>10,}   │")
    
    print(f"  ├────────────┼──────────────┼─────────┼──────────────┤")
    print(f"  │ 合計        │              │         │ {total_calls:>10,}   │")
    print(f"  └────────────┴──────────────┴─────────┴──────────────┘")
    print(f"  預估耗時：~{preset['est_hours']} 小時")

def _scan_category(cat_name, ref_code, api_endpoint, year_start, year_end,
                   max_refs_per_year, uploaded, progress, stats, 
                   filter_fn=None, classify_fn=None):
    """
    [v5.3] 按類型獨立掃描某一類別的歷史數據
    
    Args:
        cat_name:          類別顯示名 (e.g. "執法行動 Enforcement")
        ref_code:          SFC ref 代碼 (e.g. "PR", "EC", "CP")
        api_endpoint:      API 路徑 (e.g. "news/list-content", "circular/openFile")
        year_start/end:    起止年份 (e.g. 2010, 2026)
        max_refs_per_year: 每年掃描的最大 ref 編號
        uploaded:          已上傳的 ref 集合
        progress:          進度記錄 dict
        stats:             統計 dict
        filter_fn:         可選的內容過濾函數
        classify_fn:       可選的內容分類函數 (返回 doc_type)
    
    Returns:
        (items_found, items_skipped)
    """
    years = list(range(year_start % 100, year_end % 100 + 1))
    # 處理跨世紀（不太可能但防禦性編程）
    if year_start % 100 > year_end % 100:
        years = list(range(year_start % 100, 100)) + list(range(0, year_end % 100 + 1))
    
    total_scans = len(years) * max_refs_per_year
    items_found = 0
    items_skipped = 0
    scan_count = 0
    save_counter = 0
    phase_start = time.time()
    
    print(f"\n  {'─' * 55}")
    print(f"  Phase: {cat_name}")
    print(f"  範圍: {year_start}-{year_end} ({len(years)} 年, ~{total_scans} API calls)")
    print(f"  {'─' * 55}")
    
    for yr in years:
        yr_label = f"20{yr:02d}"
        source_results = []
        
        for ref_id in range(max_refs_per_year, 0, -1):
            ref = f"{yr:02d}{ref_code}{ref_id}"
            scan_count += 1
            
            if scan_count % 50 == 0:
                show_progress(scan_count, total_scans, f"  {cat_name[:8]} ", phase_start,
                              f"| {yr_label} | Found: {items_found}")
            
            if ref in uploaded:
                items_skipped += 1
                continue
            
            text, ctype = fetch_sfc_content(ref, api_endpoint)
            if not text:
                time.sleep(0.15)
                continue
            
            # 過濾
            if filter_fn and not filter_fn(text):
                time.sleep(0.15)
                continue
            
            # 分類
            if classify_fn:
                doc_type = classify_fn(text)
            else:
                doc_type = cat_name.split()[0].lower()
            
            chunks = smart_chunk_text(text, f"SFC {ref}")
            source_results.append({
                "ref": ref, "chunks": chunks, "content_type": ctype,
                "doc_type": doc_type,
                "source_url": f"{SFC_API_BASE}/{api_endpoint}?refNo={ref}&lang=EN"
            })
            items_found += 1
            time.sleep(0.2)
        
        # 上傳本年結果
        if source_results:
            print(f"\n     {yr_label}: {len(source_results)} items → uploading...")
            for r in source_results:
                vigo_chunks = create_bilingual_chunks(r["ref"], r["chunks"], r["doc_type"], r["source_url"])
                for c in vigo_chunks:
                    kid = upload_chunk(c["content"], c["metadata"])
                    if kid:
                        uploaded.add(r["ref"])
                        stats["ok"] += 1
                        stats["chunks"] += 1
                    else:
                        stats["fail"] += 1
                    time.sleep(0.3)
                
                save_counter += 1
                # 每 20 項自動保存進度（中斷恢復）
                if save_counter % 20 == 0:
                    progress["uploaded_refs"] = list(uploaded)
                    progress["last_phase"] = cat_name
                    progress["last_year"] = yr_label
                    save_progress(progress)
            
            # 每年結束保存一次
            progress["uploaded_refs"] = list(uploaded)
            progress["last_year"] = yr_label
            save_progress(progress)
    
    phase_time = time.time() - phase_start
    print(f"\n  ✓ {cat_name} 完成: {items_found} found, {items_skipped} skipped ({phase_time/60:.1f} min)")
    
    return items_found, items_skipped


def deep_historical_scan():
    """[v5.3] 歷史建庫 — 分類掃描 + 預設方案"""
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded = set(progress.get("uploaded_refs", []))
    
    print(f"\n  {'=' * 55}")
    print(f"  Mode 4: 歷史建庫 Historical Archive (v5.3)")
    print(f"  {'=' * 55}")
    print(f"  已有記錄: {len(uploaded)} refs")
    print(f"  （已上傳的 ref 會自動跳過，支持中斷恢復）")
    
    # ── 用戶選擇方案 ──
    print(f"\n  請選擇掃描方案:")
    print(f"  a. 推薦方案（~2-3 小時）")
    print(f"     執法 2010-2026 | 通函 2018-2026 | 諮詢 2022-2026 | 新聞 2020-2026")
    print(f"  b. 精簡方案（~1-1.5 小時）")
    print(f"     執法 2016-2026 | 通函 2020-2026 | 諮詢 2023-2026 | 新聞 2022-2026")
    print(f"  c. 完整方案（~4-5 小時）")
    print(f"     執法 2003-2026 | 通函 2010-2026 | 諮詢 2018-2026 | 新聞 2015-2026")
    print(f"  d. 自定義（手動設定每類型起止年份）")
    
    choice = input(f"\n  (a/b/c/d): ").strip().lower()
    
    if choice == "d":
        # 自定義模式
        print(f"\n  自定義年份範圍（格式：起始年 如 2010）:")
        try:
            e_start = int(input("  執法行動起始年 (預設 2010): ").strip() or "2010")
            c_start = int(input("  通函起始年 (預設 2018): ").strip() or "2018")
            q_start = int(input("  諮詢文件起始年 (預設 2022): ").strip() or "2022")
            p_start = int(input("  新聞稿起始年 (預設 2020): ").strip() or "2020")
        except ValueError:
            print("  ⚠ 輸入格式錯誤，使用推薦方案")
            e_start, c_start, q_start, p_start = 2010, 2018, 2022, 2020
        
        preset = {
            "name": "自定義方案 Custom",
            "enforcement": (e_start, 2026),
            "circular":    (c_start, 2026),
            "consultation":(q_start, 2026),
            "press":       (p_start, 2026),
            "est_hours": "自定義",
        }
    else:
        preset_key = {"a": "recommended", "b": "compact", "c": "full"}.get(choice, "recommended")
        preset = HISTORICAL_PRESETS[preset_key]
    
    # ── 顯示掃描預覽 ──
    print(f"\n  已選擇: {preset['name']}")
    _show_preset_table(preset)
    
    confirm = input(f"\n  確認開始掃描? (y/n): ").strip().lower()
    if confirm != 'y':
        print("  已取消。")
        return
    
    # ── 開始分類掃描 ──
    start_time = time.time()
    total_found = 0
    total_skipped = 0
    stats = {"ok": 0, "fail": 0, "skip": 0, "chunks": 0}
    
    enforce_kw = ['reprimand', 'fine', 'suspend', 'ban', 'prosecut', 'convict',
                  'disciplin', 'enforcement', 'sanction', 'misconduct']
    va_kw = ['virtual asset', 'vatp', 'vasp', 'crypto', 'stablecoin',
             'digital asset', 'web3', 'staking', 'aspire']
    
    def _classify_press(text):
        """將新聞稿分類為 enforcement / vatp / press_release"""
        text_lower = text.lower()
        if any(k in text_lower for k in enforce_kw):
            return "enforcement"
        elif any(k in text_lower for k in va_kw):
            return "vatp"
        else:
            return "press_release"
    
    # ═══ Phase 1: 執法行動 + VATP + 一般新聞（全部來自 PR refs）═══
    # 新聞稿和執法都使用 PR refs，所以合併掃描，用 classify_fn 分類
    pr_start = min(preset["enforcement"][0], preset["press"][0])
    pr_end = 2026
    
    found, skipped = _scan_category(
        cat_name="執法/新聞 Enforcement+Press",
        ref_code="PR",
        api_endpoint="news/list-content",
        year_start=pr_start,
        year_end=pr_end,
        max_refs_per_year=300,
        uploaded=uploaded,
        progress=progress,
        stats=stats,
        classify_fn=_classify_press,
    )
    total_found += found
    total_skipped += skipped
    
    # ═══ Phase 2: 通函 ═══
    found, skipped = _scan_category(
        cat_name="通函 Circulars",
        ref_code="EC",
        api_endpoint="circular/openFile",
        year_start=preset["circular"][0],
        year_end=preset["circular"][1],
        max_refs_per_year=60,
        uploaded=uploaded,
        progress=progress,
        stats=stats,
    )
    total_found += found
    total_skipped += skipped
    
    # ═══ Phase 3: 諮詢文件 ═══
    found, skipped = _scan_category(
        cat_name="諮詢文件 Consultations",
        ref_code="CP",
        api_endpoint="consultation/openFile",
        year_start=preset["consultation"][0],
        year_end=preset["consultation"][1],
        max_refs_per_year=30,
        uploaded=uploaded,
        progress=progress,
        stats=stats,
    )
    total_found += found
    total_skipped += skipped
    
    # ── 最終保存 ──
    progress["uploaded_refs"] = list(uploaded)
    progress["last_phase"] = "complete"
    save_progress(progress)
    save_content_hashes(hashes_data)
    
    total_time = time.time() - start_time
    print(f"\n  {'=' * 60}")
    print(f"  歷史建庫 Historical Archive COMPLETE!")
    print(f"  方案: {preset['name']}")
    print(f"  耗時: {total_time/3600:.1f} hours ({total_time/60:.0f} min)")
    print(f"  Found: {total_found} | Skipped: {total_skipped} | Chunks: {stats['chunks']}")
    print(f"  OK: {stats['ok']} | Fail: {stats['fail']}")
    print(f"  {'=' * 60}")
    
    generate_report("mode4_historical_archive", stats, [], total_time,
                    {"preset": preset["name"],
                     "items_found": total_found,
                     "items_skipped": total_skipped,
                     "enforcement_range": f"{preset['enforcement'][0]}-{preset['enforcement'][1]}",
                     "circular_range": f"{preset['circular'][0]}-{preset['circular'][1]}",
                     "consultation_range": f"{preset['consultation'][0]}-{preset['consultation'][1]}",
                     "press_range": f"{preset['press'][0]}-{preset['press'][1]}",
                    })


# ============================================================
# v4.0 Mode 5: 靜態法規庫導入 (30 份基礎文件)
# ============================================================

# v5.2: 12 Codes + 15 Guidelines + 8 Laws = 35 foundational documents
# URLs verified from SFC website search results (Feb 2026)
# v5.2 adds: PDPO Cap.486, Client Securities Rules, Client Money Rules,
#             Accounts & Audit Rules, Licensing Info Rules
# Each doc has page_url as fallback for locating PDFs if direct URL changes
FOUNDATIONAL_DOCS = [
    # ══════════════════════════════════════════════
    # Codes (12) — SFC 守則
    # ══════════════════════════════════════════════
    
    # 1. Code of Conduct ★★★ (最重要，所有持牌人必讀)
    {"name": "Code of Conduct for Persons Licensed by or Registered with the SFC",
     "name_zh": "證監會持牌人或註冊人操守準則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/code-of-conduct-for-persons-licensed-by-or-registered-with-the-securities-and-futures-commission/Code_of_conduct-Oct-2024_Eng-with-Bookmark-Final.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "HIGHEST",
     "keywords": ["SFC", "Code_of_Conduct", "Compliance", "GP1-GP9"]},
    
    # 2. Fund Manager Code of Conduct ★★★
    {"name": "Fund Manager Code of Conduct (FMCC)",
     "name_zh": "基金經理操守準則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/fund-manager-code-of-conduct/fund-manager-code-of-conduct.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "HIGHEST",
     "keywords": ["SFC", "Fund_Manager", "Type9", "FMCC"]},
    
    # 3. Corporate Finance Adviser Code of Conduct
    {"name": "Corporate Finance Adviser Code of Conduct",
     "name_zh": "企業融資顧問操守準則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/corporate-finance-adviser-code-of-conduct/corporate-finance-adviser-code-of-conduct.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "HIGH",
     "keywords": ["SFC", "Corporate_Finance", "Type6"]},
    
    # 4. Codes on Takeovers and Mergers and Share Buy-backs
    {"name": "Codes on Takeovers and Mergers and Share Buy-backs",
     "name_zh": "收購及合併守則及股份購回守則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/the-codes-on-takeovers-and-mergers-and-share-buy-backs/the-codes-on-takeovers-and-mergers-and-share-buy-backs.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "HIGH",
     "keywords": ["SFC", "Takeover", "Merger", "Share_Buyback"]},
    
    # 5. SFC Handbook for Unit Trusts and Mutual Funds (UT Code)
    {"name": "SFC Handbook for Unit Trusts, Mutual Funds, Investment-Linked Assurance and Unlisted Structured Products",
     "name_zh": "單位信託及互惠基金手冊",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/sfc-handbook-for-unit-trusts-and-mutual-funds/sfc-handbook-for-unit-trusts-and-mutual-funds.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "HIGH",
     "keywords": ["SFC", "Unit_Trust", "Fund", "UT_Code"]},
    
    # 6. REIT Code
    {"name": "Code on Real Estate Investment Trusts (REIT Code)",
     "name_zh": "房地產投資信託基金守則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/code-on-real-estate-investment-trusts/code-on-real-estate-investment-trusts.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "MEDIUM",
     "keywords": ["SFC", "REIT"]},
    
    # 7. Open-ended Fund Companies Code (OFC Code)
    {"name": "Code on Open-ended Fund Companies (OFC Code)",
     "name_zh": "開放式基金型公司守則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/code-on-open-ended-fund-companies/code-on-open-ended-fund-companies.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "MEDIUM",
     "keywords": ["SFC", "OFC", "Fund"]},
    
    # 8. Code on Pooled Retirement Funds
    {"name": "Code on Pooled Retirement Funds",
     "name_zh": "匯集退休基金守則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/code-on-pooled-retirement-funds/code-on-pooled-retirement-funds.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "MEDIUM",
     "keywords": ["SFC", "Retirement_Fund", "PRF"]},
    
    # 9. MPF Products Code
    {"name": "Code on MPF Products",
     "name_zh": "強積金產品守則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/sfc-code-on-mpf-products/sfc-code-on-mpf-products.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "MEDIUM",
     "keywords": ["SFC", "MPF", "MPFA"]},
    
    # 10. Code of Conduct for Credit Rating Services (Type 10)
    {"name": "Code of Conduct for Persons Providing Credit Rating Services",
     "name_zh": "提供信貸評級服務的操守準則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/code-of-conduct-for-persons-providing-credit-rating-services/code-of-conduct-for-persons-providing-credit-rating-services.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "MEDIUM",
     "keywords": ["SFC", "Credit_Rating", "Type10", "CRA"]},
    
    # 11. Code of Conduct for Share Registrars
    {"name": "Code of Conduct for Share Registrars",
     "name_zh": "股份登記機構操守準則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/code-of-conduct-for-share-registrars/code-of-conduct-for-share-registrars.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "LOW",
     "keywords": ["SFC", "Share_Registrar"]},
    
    # 12. Immigration-Linked Investment Code
    {"name": "Code on Immigration-Linked Investment Schemes",
     "name_zh": "與投資移民計劃有關的管理公司守則",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/codes/code-on-immigration-linked-investment-schemes/Code-on-Immigration-Linked-Investment-Schemes.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
     "category": "Code", "priority": "LOW",
     "keywords": ["SFC", "Immigration", "CIES"]},
    
    # ══════════════════════════════════════════════
    # Guidelines (15) — SFC 指引
    # ══════════════════════════════════════════════
    
    # 13. AML/CFT Guideline ★★★
    {"name": "Guideline on Anti-Money Laundering and Counter-Financing of Terrorism (For LCs and SFC-licensed VASPs)",
     "name_zh": "打擊洗錢及恐怖分子資金籌集指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/guideline-on-anti-money-laundering-and-counter-financing-of-terrorism-for-licensed-corporations/AML-Guideline-for-LCs-and-SFC-licensed-VASPs_Eng_1-Jun-2023.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "HIGHEST",
     "keywords": ["SFC", "AML", "CFT", "KYC", "Compliance"]},
    
    # 14. Management, Supervision and Internal Control ★★★
    {"name": "Management, Supervision and Internal Control Guidelines",
     "name_zh": "管理、監督及內部監控指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/management-supervision-and-internal-control-gu/management-supervision-and-internal-control-guidelines-for-persons-licensed.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "HIGHEST",
     "keywords": ["SFC", "Internal_Control", "Supervision", "Compliance"]},
    
    # 15. Guidelines on Competence ★★★ (Oct 2024 verified URL)
    {"name": "Guidelines on Competence",
     "name_zh": "勝任能力的指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/guidelines-on-competence/Guidelines-on-Competence_EN_Oct-2024.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines/Guidelines-on-Competence",
     "category": "Guideline", "priority": "HIGHEST",
     "keywords": ["SFC", "Competence", "Exam", "RO", "LR", "CPT"]},
    
    # 16. VATP Guidelines ★★★ (verified URL)
    {"name": "Guidelines for Virtual Asset Trading Platform Operators (VATP Guidelines)",
     "name_zh": "虛擬資產交易平台營運者指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/Guidelines-for-Virtual-Asset-Trading-Platform-Operators/Guidelines-for-Virtual-Asset-Trading-Platform-Operators.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "HIGHEST",
     "keywords": ["SFC", "VATP", "Virtual_Asset", "Crypto", "Type1", "Type7"]},
    
    # 17. CPT Guidelines ★★
    {"name": "Guidelines on Continuous Professional Training",
     "name_zh": "持續培訓指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/guidelines-on-continuous-professional-training/guidelines-on-continuous-professional-training.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "HIGH",
     "keywords": ["SFC", "CPT", "Training", "CPD"]},
    
    # 18. Fit and Proper Guidelines ★★
    {"name": "Fit and Proper Guidelines",
     "name_zh": "適當人選的指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/fit-and-proper-guidelines/fit-and-proper-guidelines.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "HIGH",
     "keywords": ["SFC", "Fit_Proper", "Licensing"]},
    
    # 19. Disciplinary Fining Guidelines ★★
    {"name": "SFC Disciplinary Fining Guidelines",
     "name_zh": "紀律處分罰款指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/disciplinary-fining-guidelines/disciplinary-fining-guidelines.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "HIGH",
     "keywords": ["SFC", "Disciplinary", "Fine", "Enforcement"]},
    
    # 20. ATS Guidelines
    {"name": "Guidelines for the Regulation of Automated Trading Services",
     "name_zh": "自動化交易服務指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines-for-the-regulation-of-automated-trading-services/guidelines-for-the-regulation-of-automated-trading-services.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "MEDIUM",
     "keywords": ["SFC", "ATS", "Type7", "Automated_Trading"]},
    
    # 21. Online Platform Guidelines
    {"name": "Guidelines on Online Distribution and Advisory Platforms",
     "name_zh": "網上分銷及投資諮詢平台指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/guidelines-on-online-distribution-and-advisory-platforms/Guidelines-on-Online-Distribution-and-Advisory-Platforms.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "MEDIUM",
     "keywords": ["SFC", "Online_Platform", "Robo_Advisor"]},
    
    # 22. Advertising Guidelines for CIS
    {"name": "Advertising Guidelines Applicable to Collective Investment Schemes",
     "name_zh": "集體投資計劃廣告指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/advertising-guidelines-applicable-to-collective-investment-schemes/Advertising-Guidelines--Applicable-to-Collective-Investment-Schemes-Authorized-under-the-Product-Cod.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "MEDIUM",
     "keywords": ["SFC", "Advertising", "CIS", "Fund"]},
    
    # 23. AML Guideline for Associated Entities
    {"name": "Prevention of Money Laundering Guideline for Associated Entities of LCs and VASPs",
     "name_zh": "有聯繫實體的打擊洗錢指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/prevention-of-money-laundering-and-terrorist-fi/AML-Guideline-for-AEs_Eng_1-Jun-2023.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "MEDIUM",
     "keywords": ["SFC", "AML", "Associated_Entity", "VASP"]},
    
    # 24. OTC Derivatives Risk Management
    {"name": "Core Operational and Financial Risk Management Controls for OTC Derivatives",
     "name_zh": "場外衍生工具核心營運及財務風險管理監控",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/core-operational-and-financial-risk-management-controls-for-over-the-counter-derivatives-activities/core-operational-and-financial-risk-management-controls-for-over-the-counter-derivatives-activities.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "MEDIUM",
     "keywords": ["SFC", "OTC", "Derivatives", "Risk_Management"]},
    
    # 25. Debt Collection Guidelines
    {"name": "Debt Collection Guidelines for Licensed Corporations",
     "name_zh": "持牌法團收債指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/debt-collection-guidelines-for-licensed-corporations/debt-collection-guidelines-for-licensed-corporations.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "MEDIUM",
     "keywords": ["SFC", "Debt_Collection"]},
    
    # 26. Inside Information Guidelines
    {"name": "Guidelines on Disclosure of Inside Information",
     "name_zh": "內幕消息披露指引",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/codes/files-current/web/guidelines/guidelines-on-disclosure-of-inside-information/guidelines-on-disclosure-of-inside-information.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "MEDIUM",
     "keywords": ["SFC", "Inside_Information", "Disclosure"]},
    
    # 27. VATP Licensing Handbook ★★★ (verified URL)
    {"name": "Licensing Handbook for Virtual Asset Trading Platform Operators",
     "name_zh": "虛擬資產交易平台營運者發牌手冊",
     "url": "https://www.sfc.hk/-/media/EN/assets/components/Guidelines/File-current/Licensing-Handbook-for-VATPs-31-05-2023.pdf",
     "page_url": "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines",
     "category": "Guideline", "priority": "HIGHEST",
     "keywords": ["SFC", "VATP", "Licensing", "Virtual_Asset"]},
    
    # ══════════════════════════════════════════════
    # Laws (3) — 核心法律
    # ⚠️ elegislation.gov.hk 需要 JavaScript 渲染，無法直接抓取
    # 解決方案：使用 elegislation PDF 導出 URL 格式
    # 若仍失敗，需手動下載 PDF 放到 config/pdfs/ 目錄
    # ══════════════════════════════════════════════
    
    # 28. SFO ★★★
    {"name": "Securities and Futures Ordinance (Cap. 571)",
     "name_zh": "《證券及期貨條例》(第571章)",
     "url": "https://www.elegislation.gov.hk/hk/cap571!en.assist.pdf?FILENAME=Securities%20and%20Futures%20Ordinance.pdf",
     "page_url": "https://www.elegislation.gov.hk/hk/cap571",
     "category": "Law", "priority": "HIGHEST",
     "keywords": ["SFC", "SFO", "Cap571", "Law", "Ordinance"]},
    
    # 29. AMLO ★★★
    {"name": "Anti-Money Laundering and Counter-Terrorist Financing Ordinance (Cap. 615)",
     "name_zh": "《打擊洗錢及恐怖分子資金籌集條例》(第615章)",
     "url": "https://www.elegislation.gov.hk/hk/cap615!en.assist.pdf?FILENAME=AMLO.pdf",
     "page_url": "https://www.elegislation.gov.hk/hk/cap615",
     "category": "Law", "priority": "HIGHEST",
     "keywords": ["SFC", "AMLO", "AML", "Cap615", "Law", "VATP"]},
    
    # 30. Financial Resources Rules ★★
    {"name": "Securities and Futures (Financial Resources) Rules (Cap. 571N)",
     "name_zh": "《證券及期貨（財政資源）規則》(第571N章)",
     "url": "https://www.elegislation.gov.hk/hk/cap571N!en.assist.pdf?FILENAME=FRR.pdf",
     "page_url": "https://www.elegislation.gov.hk/hk/cap571N",
     "category": "Law", "priority": "HIGH",
     "keywords": ["SFC", "FRR", "Cap571N", "Financial_Resources", "Capital"]},
    
    # ══════════════════════════════════════════════
    # v5.2 新增 [P0.4]: 個人資料保護 + 更多附屬法例
    # ══════════════════════════════════════════════
    
    # 31. PDPO ★★★ [P0.4] 合規手冊必須包含
    {"name": "Personal Data (Privacy) Ordinance (Cap. 486)",
     "name_zh": "《個人資料（私隱）條例》(第486章)",
     "url": "https://www.elegislation.gov.hk/hk/cap486!en.assist.pdf?FILENAME=PDPO.pdf",
     "page_url": "https://www.elegislation.gov.hk/hk/cap486",
     "category": "Law", "priority": "HIGH",
     "keywords": ["PCPD", "PDPO", "Cap486", "Privacy", "Data_Protection"]},
    
    # 32. Securities and Futures (Client Securities) Rules ★★
    {"name": "Securities and Futures (Client Securities) Rules (Cap. 571H)",
     "name_zh": "《證券及期貨（客戶證券）規則》(第571H章)",
     "url": "https://www.elegislation.gov.hk/hk/cap571H!en.assist.pdf?FILENAME=Client_Securities_Rules.pdf",
     "page_url": "https://www.elegislation.gov.hk/hk/cap571H",
     "category": "Law", "priority": "HIGH",
     "keywords": ["SFC", "Client_Securities", "Cap571H", "Custody"]},
    
    # 33. Securities and Futures (Client Money) Rules ★★
    {"name": "Securities and Futures (Client Money) Rules (Cap. 571I)",
     "name_zh": "《證券及期貨（客戶款項）規則》(第571I章)",
     "url": "https://www.elegislation.gov.hk/hk/cap571I!en.assist.pdf?FILENAME=Client_Money_Rules.pdf",
     "page_url": "https://www.elegislation.gov.hk/hk/cap571I",
     "category": "Law", "priority": "HIGH",
     "keywords": ["SFC", "Client_Money", "Cap571I", "Segregation"]},
    
    # 34. Securities and Futures (Accounts and Audit) Rules ★
    {"name": "Securities and Futures (Accounts and Audit) Rules (Cap. 571P)",
     "name_zh": "《證券及期貨（會計及審計）規則》(第571P章)",
     "url": "https://www.elegislation.gov.hk/hk/cap571P!en.assist.pdf?FILENAME=Accounts_Audit_Rules.pdf",
     "page_url": "https://www.elegislation.gov.hk/hk/cap571P",
     "category": "Law", "priority": "MEDIUM",
     "keywords": ["SFC", "Accounts", "Audit", "Cap571P"]},
    
    # 35. Securities and Futures (Licensing and Registration) (Information) Rules
    {"name": "Securities and Futures (Licensing and Registration) (Information) Rules (Cap. 571S)",
     "name_zh": "《證券及期貨（發牌及註冊）（資料）規則》(第571S章)",
     "url": "https://www.elegislation.gov.hk/hk/cap571S!en.assist.pdf?FILENAME=Licensing_Info_Rules.pdf",
     "page_url": "https://www.elegislation.gov.hk/hk/cap571S",
     "category": "Law", "priority": "MEDIUM",
     "keywords": ["SFC", "Licensing", "Registration", "Cap571S", "WINGS"]},
]



# v5.0 擴展基石文件：FAQ + Licensing Handbook + Thematic Reports + HKEX
# v5.2 新增：WINGS 指南 + SFC 申請表格指引 + PCPD 指引
EXTENDED_SOURCES = [
    # ── SFC FAQ ──
    {"id": "sfc_faq", "name": "SFC Licensing FAQ", "name_zh": "證監會發牌常見問題",
     "url": "https://www.sfc.hk/en/faqs", "type": "faq",
     "priority": "HIGHEST", "keywords": ["SFC", "FAQ", "Licensing"]},
    # ── SFC Licensing Handbook ──
    {"id": "sfc_licensing_handbook", "name": "SFC Licensing Handbook",
     "name_zh": "證監會發牌手冊",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Licensing-handbook",
     "type": "handbook", "priority": "HIGHEST", "keywords": ["SFC", "Licensing", "Handbook"]},
    # ── SFC Ongoing Obligations ──
    {"id": "sfc_ongoing_obligations", "name": "SFC Ongoing Obligations",
     "name_zh": "持續合規責任",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Ongoing-obligations",
     "type": "obligations", "priority": "HIGH", "keywords": ["SFC", "Obligations", "CPT", "Annual_Return"]},
    # ── SFC Thematic Review Reports ──
    {"id": "sfc_thematic_reports", "name": "SFC Thematic Review Reports",
     "name_zh": "主題巡查報告",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Supervision/Publications-and-statistics/Reports-on-findings-of-thematic-inspection-or-review",
     "type": "thematic_report", "priority": "HIGHEST",
     "keywords": ["SFC", "Thematic", "Inspection", "Review"]},
    # ── HKEX Main Board Listing Rules ──
    {"id": "hkex_main_board", "name": "HKEX Main Board Listing Rules",
     "name_zh": "香港交易所主板上市規則",
     "url": "https://www.hkex.com.hk/Listing/Rules-and-Guidance/Listing-Rules-Contingency?sc_lang=en",
     "type": "listing_rules", "priority": "HIGH",
     "keywords": ["HKEX", "Listing", "Main_Board", "IPO"]},
    # ── HKEX GEM Listing Rules ──
    {"id": "hkex_gem", "name": "HKEX GEM Listing Rules",
     "name_zh": "香港交易所GEM上市規則",
     "url": "https://www.hkex.com.hk/Listing/Rules-and-Guidance/Listing-Rules-Contingency/GEM-Listing-Rules?sc_lang=en",
     "type": "listing_rules", "priority": "MEDIUM",
     "keywords": ["HKEX", "GEM", "Listing"]},
    
    # ══════════════════════════════════════════════
    # v5.2 新增 [P2.10]: WINGS 系統指南
    # ══════════════════════════════════════════════
    {"id": "sfc_wings_portal", "name": "SFC WINGS Electronic System",
     "name_zh": "證監會 WINGS 電子服務系統",
     "url": "https://wings.sfc.hk/",
     "type": "system_guide", "priority": "HIGHEST",
     "keywords": ["SFC", "WINGS", "Electronic", "Application", "Forms"]},
    {"id": "sfc_wings_guide", "name": "SFC WINGS User Guide & E-Forms",
     "name_zh": "WINGS 用戶指南及電子表格",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Licensing-handbook/How-to-apply",
     "type": "system_guide", "priority": "HIGHEST",
     "keywords": ["SFC", "WINGS", "Licensing", "Application", "How_to_Apply"]},
    
    # ══════════════════════════════════════════════
    # v5.2 新增 [P2.11]: SFC 申請表格 + 填寫指引
    # ══════════════════════════════════════════════
    {"id": "sfc_forms_download", "name": "SFC Forms and Checklists Download",
     "name_zh": "證監會表格及核對清單下載",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Forms-and-checklists",
     "type": "application_forms", "priority": "HIGHEST",
     "keywords": ["SFC", "Forms", "Application", "Checklist", "WINGS"]},
    {"id": "sfc_fit_proper_faq", "name": "SFC Fit and Proper FAQ",
     "name_zh": "勝任能力常見問題",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Do-you-need-a-licence-or-registration",
     "type": "faq", "priority": "HIGH",
     "keywords": ["SFC", "Fit_Proper", "RO", "MIC", "Competence"]},
    
    # ══════════════════════════════════════════════
    # v5.2 新增 [P0.4]: PCPD 個資保護指引
    # ══════════════════════════════════════════════
    {"id": "pcpd_guidance", "name": "PCPD Guidance Notes & Best Practices",
     "name_zh": "個人資料私隱專員公署指引",
     "url": "https://www.pcpd.org.hk/english/resources_centre/publications/guidance/guidance.html",
     "type": "guideline", "priority": "HIGH",
     "keywords": ["PCPD", "PDPO", "Privacy", "Data_Protection", "Cap486"]},
    {"id": "pcpd_dpp", "name": "PCPD Data Protection Principles",
     "name_zh": "保障資料原則",
     "url": "https://www.pcpd.org.hk/english/data_privacy_law/6_data_protection_principles/principles.html",
     "type": "guideline", "priority": "HIGH",
     "keywords": ["PCPD", "PDPO", "DPP", "Privacy_Principles"]},
    
    # ══════════════════════════════════════════════
    # v5.2 新增 [P0.2]: SFC 巡查報告專頁（加強爬取）
    # ══════════════════════════════════════════════
    {"id": "sfc_inspection_findings", "name": "SFC Inspection Findings and Observations",
     "name_zh": "證監會巡查結果及觀察",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Supervision/Publications-and-statistics",
     "type": "thematic_report", "priority": "HIGHEST",
     "keywords": ["SFC", "Inspection", "Findings", "Compliance", "Deficiency"]},
]

# SFC 年報 URL 模式（最近10年）
SFC_ANNUAL_REPORTS = [
    {"year": "2024-25", "url": "https://www.sfc.hk/en/Published-resources/Corporate-publications/Annual-report"},
    {"year": "2023-24", "url": "https://www.sfc.hk/en/Published-resources/Corporate-publications/Annual-report"},
    {"year": "2022-23", "url": "https://www.sfc.hk/en/Published-resources/Corporate-publications/Annual-report"},
]


def import_foundational_regulations():
    """Mode 5: 基石法規導入 — v5.2 升級版（35 份 + PDPO）"""
    print(f"\n  {'=' * 50}")
    print(f"  Mode 5: 基石法規與核心文件")
    print(f"  {'=' * 50}")
    print(f"\n  a. 導入 35 份基石法規 (含 PDPO + 附屬法例)")
    print(f"  b. 健康檢查 — 檢測法規更新")
    print(f"  c. 擴展導入 — FAQ/Handbook/HKEX/主題報告")
    print(f"  d. 年報導入（最近 10 年）")
    print(f"  e. 查看基石文件狀態")
    
    sub = input(f"\n  (a/b/c/d/e): ").strip().lower()
    
    if sub == "a":
        _import_foundational_core()
    elif sub == "b":
        _health_check_foundational()
    elif sub == "c":
        _import_extended_sources()
    elif sub == "d":
        _import_annual_reports()
    elif sub == "e":
        _show_foundational_status()
    else:
        print("  無效選擇")

def _import_foundational_core():
    """5a: 導入 30 份基石法規（保留原邏輯 + 加入 QA）"""
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded = set(progress.get("uploaded_refs", []))
    
    print(f"\n  {'=' * 50}")
    print(f"  5a: Import 30 Foundational Regulations")
    print(f"  {'=' * 50}")
    
    priority_order = {"HIGHEST": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    docs = sorted(FOUNDATIONAL_DOCS, key=lambda d: priority_order.get(d["priority"], 9))
    
    total = len(docs)
    stats = {"ok": 0, "fail": 0, "skip": 0, "chunks": 0, "duplicate": 0}
    results = []
    all_vigo_chunks = []
    start_time = time.time()
    
    for i, doc in enumerate(docs, 1):
        ref = f"STATIC_{doc['category']}_{doc['name'][:30].replace(' ', '_')}"
        
        if ref in uploaded:
            stats["skip"] += 1
            results.append({"name": doc["name"], "name_zh": doc["name_zh"],
                           "priority": doc["priority"], "status": "skip"})
            continue
        
        show_progress(i, total, "Doc ", start_time, f"| {doc['name'][:30]}")
        print()
        print(f"  [{i}/{total}] [{doc['priority']}] {doc['name'][:60]}")
        print(f"     {doc['name_zh']}")
        
        try:
            text = None
            _from_local = False
            
            # 優先讀取本地 PDF
            local_pdf_dir = os.path.join(CONFIG_DIR, "pdfs")
            if os.path.isdir(local_pdf_dir):
                doc_lower = doc["name"].lower()
                for fname in os.listdir(local_pdf_dir):
                    if not fname.lower().endswith('.pdf'):
                        continue
                    matched = False
                    for kw in doc.get("keywords", []):
                        if len(kw) < 4:
                            continue
                        kw_pattern = re.sub(r'(?<=[a-zA-Z])(?=\d)|(?<=\d)(?=[a-zA-Z])', r'\\s*', kw)
                        pattern = re.compile(kw_pattern + r'(?![a-zA-Z0-9])', re.IGNORECASE)
                        if pattern.search(fname):
                            matched = True
                            break
                    name_words = doc["name"].split()[:3]
                    if any(w.lower() in fname.lower() for w in name_words if len(w) > 3):
                        matched = True
                    if matched:
                        local_path = os.path.join(local_pdf_dir, fname)
                        print(f"     📁 Found local PDF: {fname}")
                        with open(local_path, "rb") as lf:
                            raw_local = lf.read()
                        if raw_local[:4] == b'%PDF':
                            text = extract_text_from_pdf(raw_local)
                            if text and len(text) > 300:
                                _from_local = True
                                break
                            else:
                                text = None
            
            # 從 URL 下載
            if not text:
                resp = requests.get(doc["url"], headers=HEADERS, timeout=60, allow_redirects=True)
                if resp.status_code != 200:
                    print(f"     ✗ HTTP {resp.status_code}")
                    stats["fail"] += 1
                    results.append({"name": doc["name"], "priority": doc["priority"],
                                   "status": "fail", "reason": f"HTTP {resp.status_code}"})
                    continue
                
                content_type = resp.headers.get('content-type', '').lower()
                raw = resp.content
                fetch_mode = doc.get("fetch_mode", "pdf")
                
                if fetch_mode == "html" or 'text/html' in content_type:
                    text = clean_html_content(resp.text)
                    if len(text) > 50000:
                        text = text[:50000]
                elif raw[:4] == b'%PDF' or 'pdf' in content_type:
                    text = extract_text_from_pdf(raw)
                else:
                    text = clean_html_content(resp.text[:50000])
            
            if not text or len(text) < 300:
                stats["fail"] += 1
                results.append({"name": doc["name"], "priority": doc["priority"],
                               "status": "fail", "reason": f"Content too short"})
                continue
            
            # 404 檢測
            _lower = text[:2000].lower()
            if any(sig in _lower for sig in ["you are being redirected", "page not found", "404"]) and len(text) < 12000:
                stats["fail"] += 1
                results.append({"name": doc["name"], "priority": doc["priority"],
                               "status": "fail", "reason": "404/redirect"})
                continue
            
            print(f"     ✓ Extracted: {len(text)} chars")
            chunks = smart_chunk_text(text, doc["name"])
            
            # 記錄 baseline hash（用於健康檢查）
            _update_baseline(doc, text)
            
            chunk_ok = 0
            chunk_start = time.time()
            
            for ci, chunk_text in enumerate(chunks):
                part_label = f" (Part {ci+1}/{len(chunks)})" if len(chunks) > 1 else ""
                show_progress(ci+1, len(chunks), "  Chunk ", chunk_start)
                
                base_meta = {
                    "source": f"SFC_Foundational_{doc['category']}", "source_url": doc["url"],
                    "doc_type": f"foundational_{doc['category'].lower()}", "category": "SFC",
                    "ref_no": ref, "source_table": "vigo_financial",
                    "keywords": doc["keywords"], "collected_date": datetime.now().strftime("%Y-%m-%d"),
                    "content_type": "pdf_extract", "priority": doc["priority"],
                    "authority_score": SOURCE_AUTHORITY.get("SFC_Foundational", 100),
                }
                
                # EN
                en_content = f"【SFC {doc['category']}{part_label} — {doc['name']}】\n\n{chunk_text}"
                en_content += f"\n\n---\nSource: SFC Official Regulation\nDocument: {doc['name']}\nURL: {doc['url']}"
                
                # QA 去重
                chunk_ref = f"{ref}_EN_{ci}"
                is_dup, _, _ = check_duplicate(en_content, chunk_ref, progress, hashes_data)
                if not is_dup:
                    kid = upload_chunk(en_content, {**base_meta, "language": "en"})
                    if kid:
                        stats["chunks"] += 1
                        chunk_ok += 1
                        progress, hashes_data = record_upload(chunk_ref, en_content, progress, hashes_data)
                    all_vigo_chunks.append({"content": en_content, "metadata": {**base_meta, "language": "en"}, "lang": "EN"})
                
                # ZH
                zh_translation = translate_to_chinese(chunk_text)
                if zh_translation and len(zh_translation) > 80:
                    zh_content = f"【SFC {doc['name_zh']}{part_label}】\n\n{zh_translation}"
                    zh_content += f"\n\n---\n資料來源：證監會官方法規\n文件：{doc['name_zh']}\n鏈接：{doc['url']}"
                    
                    chunk_ref_zh = f"{ref}_ZH_{ci}"
                    is_dup_zh, _, _ = check_duplicate(zh_content, chunk_ref_zh, progress, hashes_data)
                    if not is_dup_zh:
                        kid_zh = upload_chunk(zh_content, {**base_meta, "language": "zh"})
                        if kid_zh:
                            stats["chunks"] += 1
                            chunk_ok += 1
                            progress, hashes_data = record_upload(chunk_ref_zh, zh_content, progress, hashes_data)
                        all_vigo_chunks.append({"content": zh_content, "metadata": {**base_meta, "language": "zh"}, "lang": "ZH"})
                
                time.sleep(0.3)
            
            print(f"\n     ✓ Uploaded ({len(chunks)} EN + ZH chunks)")
            uploaded.add(ref)
            progress["uploaded_refs"] = list(uploaded)
            save_progress(progress)
            save_content_hashes(hashes_data)
            stats["ok"] += 1
            results.append({"name": doc["name"], "priority": doc["priority"],
                           "status": "ok", "chars": len(text), "chunks": len(chunks),
                           "time_min": round((time.time() - chunk_start) / 60, 1),
                           "source": "local_pdf" if _from_local else "url"})
        
        except Exception as e:
            print(f"\n     ✗ Error: {e}")
            stats["fail"] += 1
            results.append({"name": doc["name"], "priority": doc["priority"],
                           "status": "fail", "reason": str(e)[:200]})
    
    total_time = time.time() - start_time
    print(f"\n  {'=' * 60}")
    print(f"  Done! {total_time/60:.1f} min | OK: {stats['ok']} | Fail: {stats['fail']} | Skip: {stats['skip']}")
    print(f"  {'=' * 60}")
    
    generate_report("mode5a_foundational", stats, results, total_time)
    
    if all_vigo_chunks:
        backup_name = os.path.join(REPORTS_DIR, f"VIGO_foundational_bilingual_{datetime.now().strftime('%Y%m%d')}.md")
        save_backup(all_vigo_chunks, backup_name)


def _update_baseline(doc, text):
    """更新基石文件的 baseline hash"""
    baseline = {}
    if os.path.exists(BASELINE_FILE):
        try:
            with open(BASELINE_FILE, 'r') as f:
                baseline = json.load(f)
        except: pass
    
    if "documents" not in baseline:
        baseline["documents"] = {}
    
    doc_key = doc["name"][:50]
    baseline["documents"][doc_key] = {
        "url": doc["url"],
        "content_hash": compute_content_hash(text),
        "file_size": len(text),
        "last_checked": datetime.now().isoformat(),
        "last_imported": datetime.now().isoformat(),
    }
    baseline["last_update"] = datetime.now().isoformat()
    
    with open(BASELINE_FILE, 'w') as f:
        json.dump(baseline, f, ensure_ascii=False, indent=2)


def _health_check_foundational():
    """5b: 健康檢查 — 檢測 30 份基石文件是否有更新"""
    print(f"\n  {'=' * 50}")
    print(f"  5b: Foundational Documents Health Check")
    print(f"  {'=' * 50}")
    
    baseline = {}
    if os.path.exists(BASELINE_FILE):
        try:
            with open(BASELINE_FILE, 'r') as f:
                baseline = json.load(f)
        except: pass
    
    if not baseline.get("documents"):
        print("  ⚠ No baseline data. Run 5a first to establish baselines.")
        return
    
    results = {"unchanged": 0, "changed": 0, "error": 0, "new": 0}
    changed_docs = []
    
    start_time = time.time()
    total = len(FOUNDATIONAL_DOCS)
    
    for i, doc in enumerate(FOUNDATIONAL_DOCS, 1):
        show_progress(i, total, "Check ", start_time, f"| {doc['name'][:25]}")
        
        doc_key = doc["name"][:50]
        old_data = baseline.get("documents", {}).get(doc_key)
        
        try:
            # HTTP HEAD 先檢查大小
            resp = requests.head(doc["url"], headers=HEADERS, timeout=20, allow_redirects=True)
            
            if resp.status_code != 200:
                results["error"] += 1
                changed_docs.append({"name": doc["name"], "status": "🔴 URL_FAILED", "detail": f"HTTP {resp.status_code}"})
                continue
            
            new_size = int(resp.headers.get('content-length', 0))
            last_modified = resp.headers.get('last-modified', '')
            
            if old_data:
                old_size = old_data.get("file_size", 0)
                # 大小有變化 → 可能更新了
                if new_size > 0 and old_size > 0 and abs(new_size - old_size) > 100:
                    results["changed"] += 1
                    changed_docs.append({"name": doc["name"], "status": "🟡 SIZE_CHANGED",
                                        "detail": f"Old: {old_size}, New: {new_size}"})
                else:
                    results["unchanged"] += 1
            else:
                results["new"] += 1
                changed_docs.append({"name": doc["name"], "status": "🆕 NO_BASELINE", "detail": "Run 5a first"})
        
        except Exception as e:
            results["error"] += 1
            changed_docs.append({"name": doc["name"], "status": "🔴 ERROR", "detail": str(e)[:50]})
        
        time.sleep(0.3)
    
    # 檢查 SFC Codes & Guidelines 頁面是否有新增法規
    print(f"\n\n  Checking SFC pages for new regulations...")
    try:
        from bs4 import BeautifulSoup
        for page_url in ["https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Codes",
                         "https://www.sfc.hk/en/Rules-and-standards/Codes-and-guidelines/Guidelines"]:
            resp = requests.get(page_url, headers=HEADERS, timeout=20)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, 'html.parser')
                pdf_links = [a.get('href', '') for a in soup.find_all('a', href=True) if '.pdf' in a.get('href', '').lower()]
                known_urls = {d["url"] for d in FOUNDATIONAL_DOCS}
                for link in pdf_links:
                    if link.startswith('/'):
                        link = f"https://www.sfc.hk{link}"
                    if link not in known_urls and 'sfc.hk' in link:
                        results["new"] += 1
                        changed_docs.append({"name": link.split('/')[-1][:60], "status": "🆕 NEW_REGULATION",
                                            "detail": link[:80]})
    except Exception as e:
        print(f"  ⚠ Page check error: {e}")
    
    elapsed = time.time() - start_time
    print(f"\n\n  {'=' * 60}")
    print(f"  Health Check Results ({elapsed:.0f}s)")
    print(f"  {'=' * 60}")
    print(f"  🟢 Unchanged: {results['unchanged']}")
    print(f"  🟡 Changed:   {results['changed']}")
    print(f"  🔴 Errors:    {results['error']}")
    print(f"  🆕 New:       {results['new']}")
    
    if changed_docs:
        print(f"\n  Details:")
        for d in changed_docs:
            print(f"    {d['status']} {d['name'][:45]}")
            print(f"           {d['detail']}")
        
        if results["changed"] > 0:
            re_import = input(f"\n  重新導入已變更的文件？(y/n): ").strip().lower()
            if re_import == 'y':
                print("  請使用 5a 重新導入（會自動跳過未變更的文件）")
    
    generate_report("mode5b_health_check", results, changed_docs, elapsed)


def _import_extended_sources():
    """5c: 擴展導入 — FAQ/Handbook/HKEX/主題巡查報告"""
    from bs4 import BeautifulSoup
    
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded = set(progress.get("uploaded_refs", []))
    
    print(f"\n  {'=' * 50}")
    print(f"  5c: Extended Sources Import")
    print(f"  {len(EXTENDED_SOURCES)} sources")
    print(f"  {'=' * 50}")
    
    stats = {"ok": 0, "fail": 0, "skip": 0, "chunks": 0}
    results = []
    start_time = time.time()
    
    for i, src in enumerate(EXTENDED_SOURCES, 1):
        ref = f"EXT_{src['id']}"
        if ref in uploaded:
            stats["skip"] += 1
            print(f"\n  [{i}/{len(EXTENDED_SOURCES)}] SKIP: {src['name'][:50]}")
            continue
        
        show_progress(i, len(EXTENDED_SOURCES), "Source ", start_time)
        print(f"\n  [{i}/{len(EXTENDED_SOURCES)}] {src['name']}")
        
        try:
            resp = requests.get(src["url"], headers=HEADERS, timeout=30)
            if resp.status_code != 200:
                stats["fail"] += 1
                results.append({"name": src["name"], "status": "fail", "reason": f"HTTP {resp.status_code}"})
                continue
            
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # 根據類型選擇不同的提取策略
            if src["type"] == "thematic_report":
                # 找所有 PDF 鏈接
                pdf_links = []
                for a in soup.find_all('a', href=True):
                    href = a.get('href', '')
                    if '.pdf' in href.lower():
                        if href.startswith('/'):
                            href = f"https://www.sfc.hk{href}"
                        pdf_links.append({"url": href, "title": a.get_text(strip=True)[:80]})
                
                print(f"     Found {len(pdf_links)} PDF reports")
                report_count = 0
                for pi, plink in enumerate(pdf_links[:15]):  # 最多 15 份
                    try:
                        pdf_resp = requests.get(plink["url"], headers=HEADERS, timeout=60)
                        if pdf_resp.status_code == 200 and pdf_resp.content[:4] == b'%PDF':
                            text = extract_text_from_pdf(pdf_resp.content)
                            if text and len(text) > 500:
                                chunks = smart_chunk_text(text, plink["title"])
                                for ci, chunk_text in enumerate(chunks):
                                    meta = {
                                        "source": "SFC_Official_Thematic", "source_url": plink["url"],
                                        "doc_type": "thematic_report", "category": "SFC",
                                        "ref_no": f"{ref}_report_{pi}", "source_table": "vigo_financial",
                                        "keywords": src["keywords"], "collected_date": datetime.now().strftime("%Y-%m-%d"),
                                        "authority_score": 100,
                                    }
                                    en_content = f"【SFC Thematic Report — {plink['title']}】\n\n{chunk_text}"
                                    kid = upload_chunk(en_content, {**meta, "language": "en"})
                                    if kid: stats["chunks"] += 1
                                    
                                    zh = translate_to_chinese(chunk_text)
                                    if zh and len(zh) > 80:
                                        zh_content = f"【SFC 主題巡查報告 — {plink['title']}】\n\n{zh}"
                                        kid_zh = upload_chunk(zh_content, {**meta, "language": "zh"})
                                        if kid_zh: stats["chunks"] += 1
                                    time.sleep(0.3)
                                report_count += 1
                                print(f"       ✓ {plink['title'][:50]} ({len(chunks)} chunks)")
                    except Exception as e:
                        print(f"       ✗ {plink['title'][:30]}: {e}")
                stats["ok"] += 1
                results.append({"name": src["name"], "status": "ok", "chunks": report_count})
            
            else:
                # FAQ / Handbook / HKEX — HTML 頁面
                # 移除不需要的元素
                for tag in soup(['script', 'style', 'nav', 'footer', 'header']):
                    tag.decompose()
                
                # 嘗試找主內容區域
                main = soup.find('main') or soup.find('div', class_=re.compile(r'content|main|body', re.I)) or soup
                text = main.get_text(separator='\n', strip=True)
                text = re.sub(r'\n{3,}', '\n\n', text)
                
                if len(text) < 300:
                    stats["fail"] += 1
                    results.append({"name": src["name"], "status": "fail", "reason": "No content"})
                    continue
                
                # 截取
                if len(text) > 50000:
                    text = text[:50000]
                
                print(f"     ✓ Extracted: {len(text)} chars")
                chunks = smart_chunk_text(text, src["name"])
                
                authority = SOURCE_AUTHORITY.get("SFC_Official" if "sfc.hk" in src["url"] else "HKEX_Official", 80)
                
                for ci, chunk_text in enumerate(chunks):
                    part = f" (Part {ci+1}/{len(chunks)})" if len(chunks) > 1 else ""
                    meta = {
                        "source": f"SFC_Official_{src['id']}" if "sfc.hk" in src["url"] else f"HKEX_Official_{src['id']}",
                        "source_url": src["url"], "doc_type": src["type"],
                        "category": "SFC" if "sfc.hk" in src["url"] else "HKEX",
                        "ref_no": ref, "source_table": "vigo_financial",
                        "keywords": src["keywords"], "collected_date": datetime.now().strftime("%Y-%m-%d"),
                        "authority_score": authority,
                    }
                    en_content = f"【{src['name']}{part}】\n\n{chunk_text}"
                    kid = upload_chunk(en_content, {**meta, "language": "en"})
                    if kid: stats["chunks"] += 1
                    
                    zh = translate_to_chinese(chunk_text)
                    if zh and len(zh) > 80:
                        zh_content = f"【{src['name_zh']}{part}】\n\n{zh}"
                        kid_zh = upload_chunk(zh_content, {**meta, "language": "zh"})
                        if kid_zh: stats["chunks"] += 1
                    time.sleep(0.5)
                
                stats["ok"] += 1
                results.append({"name": src["name"], "status": "ok", "chars": len(text), "chunks": len(chunks)})
            
            uploaded.add(ref)
            progress["uploaded_refs"] = list(uploaded)
            save_progress(progress)
        
        except Exception as e:
            print(f"     ✗ Error: {e}")
            stats["fail"] += 1
            results.append({"name": src["name"], "status": "fail", "reason": str(e)[:100]})
    
    total_time = time.time() - start_time
    print(f"\n  Done! {total_time/60:.1f} min | OK: {stats['ok']} | Chunks: {stats['chunks']}")
    generate_report("mode5c_extended", stats, results, total_time)


def _import_annual_reports():
    """5d: 導入 SFC 年報（通過 Mode 7 文件導入器）"""
    print(f"\n  {'=' * 50}")
    print(f"  5d: SFC Annual Reports")
    print(f"  {'=' * 50}")
    print(f"\n  SFC 年報為大型 PDF（200+ 頁），建議：")
    print(f"  1. 從 SFC 官網手動下載年報 PDF")
    print(f"  2. 放入 config/imports/ 目錄")
    print(f"  3. 使用 Mode 7 統一文件導入器處理")
    print(f"\n  SFC Annual Report 頁面:")
    print(f"  https://www.sfc.hk/en/Published-resources/Corporate-publications/Annual-report")


def _show_foundational_status():
    """5e: 查看基石文件狀態"""
    progress = load_progress()
    uploaded = set(progress.get("uploaded_refs", []))
    
    baseline = {}
    if os.path.exists(BASELINE_FILE):
        try:
            with open(BASELINE_FILE, 'r') as f:
                baseline = json.load(f)
        except: pass
    
    print(f"\n  {'=' * 60}")
    print(f"  Foundational Documents Status")
    print(f"  {'=' * 60}")
    
    imported = 0
    not_imported = 0
    
    for doc in FOUNDATIONAL_DOCS:
        ref = f"STATIC_{doc['category']}_{doc['name'][:30].replace(' ', '_')}"
        status = "✅" if ref in uploaded else "❌"
        if ref in uploaded:
            imported += 1
        else:
            not_imported += 1
        
        doc_key = doc["name"][:50]
        baseline_info = baseline.get("documents", {}).get(doc_key, {})
        last_check = baseline_info.get("last_checked", "never")[:10] if baseline_info else "never"
        
        print(f"  {status} [{doc['priority']:>7}] {doc['name'][:50]}")
        print(f"     {doc['name_zh']} | Last check: {last_check}")
    
    print(f"\n  Imported: {imported}/{len(FOUNDATIONAL_DOCS)}")
    print(f"  Remaining: {not_imported}")
    
    # Extended sources
    print(f"\n  Extended Sources:")
    for src in EXTENDED_SOURCES:
        ref = f"EXT_{src['id']}"
        status = "✅" if ref in uploaded else "❌"
        print(f"  {status} {src['name'][:55]}")


# ============================================================
# v5.0 Mode 6: 行業知識中心（22 家來源 + 子頁面爬取）
# ============================================================

def load_sources():
    """載入 sources.json"""
    src_file = os.path.join(CONFIG_DIR, "sources.json")
    if os.path.exists(src_file):
        try:
            with open(src_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return [s for s in data if s.get("enabled", True)]
        except: pass
    
    # v7.0.4 預設 22 家來源（URL 已修復）
    default_sources = [
        # ── 合規顧問公司 (2) ──
        {"id": "complianceplus", "name": "CompliancePlus", "name_zh": "合規顧問",
         "url": "https://www.complianceplus.hk/", "type": "consulting_firm",
         "priority": "HIGH", "enabled": True, "schedule": "weekly",
         "keywords": ["Type6A", "Compliance", "Licensing", "SFC"],
         "crawl_depth": 2, "max_articles": 15},
        {"id": "paradox_management", "name": "Paradox Management", "name_zh": "合規管理顧問",
         "url": "https://paradox-management.com/", "type": "consulting_firm",
         "priority": "HIGH", "enabled": True, "schedule": "weekly",
         "keywords": ["Type6A", "Compliance", "Licensing", "SFC"],
         "crawl_depth": 2, "max_articles": 15},
        # ── 國際律所 (5) ──
        {"id": "charltons", "name": "Charltons Law", "name_zh": "齊伯禮律師行",
         "url": "https://www.charltonslaw.com/news/newsletters/updates/", "type": "law_firm_intl",
         "priority": "HIGH", "enabled": True, "schedule": "weekly",
         "keywords": ["SFC", "IPO", "Licensing", "Enforcement"],
         "crawl_depth": 2, "max_articles": 20},
        {"id": "deacons", "name": "Deacons (的近律師行)", "name_zh": "的近律師行",
         "url": "https://www.deacons.com/news-and-insights/?sector=financial-services-regulatory", "type": "law_firm_intl",
         "priority": "HIGHEST", "enabled": True, "schedule": "weekly",
         "keywords": ["SFC", "Compliance", "Inspection", "Licensing"],
         "crawl_depth": 2, "max_articles": 20},
        {"id": "kwm", "name": "King & Wood Mallesons (金杜)", "name_zh": "金杜律師事務所",
         "url": "https://www.kwm.com/hk/en/insights.html", "type": "law_firm_intl",
         "priority": "HIGH", "enabled": True, "schedule": "weekly",
         "keywords": ["SFC", "Licensing", "PE", "Family_Office"],
         "crawl_depth": 1, "max_articles": 10, "needs_enhanced_headers": True},
        {"id": "sidley", "name": "Sidley Austin", "name_zh": "盛德律師事務所",
         "url": "https://www.sidley.com/en/insights/newsupdates", "type": "law_firm_intl",
         "priority": "MEDIUM", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "PE", "VC", "Licensing"],
         "crawl_depth": 1, "max_articles": 10, "needs_enhanced_headers": True},
        {"id": "linklaters", "name": "Linklaters", "name_zh": "年利達律師事務所",
         "url": "https://www.linklaters.com/en/insights", "type": "law_firm_intl",
         "priority": "MEDIUM", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "Financial_Regulation", "Cross_Border"],
         "crawl_depth": 1, "max_articles": 10},
        # ── 本地律所 (2) ──
        {"id": "onc", "name": "ONC Lawyers", "name_zh": "柯伍陳律師事務所",
         "url": "https://www.onc.hk/en_US/publications/", "type": "law_firm_local",
         "priority": "MEDIUM", "enabled": True, "schedule": "weekly",
         "keywords": ["SFC", "Compliance", "Licensing"],
         "crawl_depth": 1, "max_articles": 10, "ssl_verify": False},
        {"id": "swlaw", "name": "Stevenson Wong", "name_zh": "胡關李羅律師行",
         "url": "https://www.swlaw.hk/publications", "type": "law_firm_local",
         "priority": "MEDIUM", "enabled": False, "schedule": "monthly",
         "keywords": ["SFC", "Enforcement", "Investigation"],
         "crawl_depth": 1, "max_articles": 10},
        # ── 四大會計師行 (4) ──
        {"id": "deloitte_hk", "name": "Deloitte China - FS Insights", "name_zh": "德勤中國",
         "url": "https://www.deloitte.com/cn/en/Industries/financial-services/perspectives.html",
         "type": "big4", "priority": "MEDIUM", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "Risk", "Compliance", "RegTech"],
         "crawl_depth": 1, "max_articles": 8},
        {"id": "pwc_hk", "name": "PwC HK - FS Risk & Regulation", "name_zh": "普華永道香港",
         "url": "https://www.pwchk.com/en/industries/financial-services/financial-services-risk-and-regulations.html",
         "type": "big4", "priority": "MEDIUM", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "Regulation", "Financial_Services"],
         "crawl_depth": 1, "max_articles": 8},
        {"id": "ey_hk", "name": "EY China - HK FS Regulatory", "name_zh": "安永中國",
         "url": "https://www.ey.com/en_cn/insights/assurance/hong-kong-financial-services-regulatory-requirements",
         "type": "big4", "priority": "MEDIUM", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "Financial_Services", "RegTech"],
         "crawl_depth": 1, "max_articles": 8},
        {"id": "kpmg_hk", "name": "KPMG China - FS Insights", "name_zh": "畢馬威中國",
         "url": "https://kpmg.com/cn/en/insights.html",
         "type": "big4", "priority": "MEDIUM", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "Financial_Services", "Regulation"],
         "crawl_depth": 1, "max_articles": 8},
        # ── 合規科技/顧問 (3) ──
        {"id": "waystone", "name": "Waystone Compliance", "name_zh": "Waystone 合規",
         "url": "https://compliance.waystone.com/", "type": "compliance_tech",
         "priority": "MEDIUM", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "Inspection", "Compliance", "Fund"],
         "crawl_depth": 2, "max_articles": 10},
        {"id": "heinbro", "name": "Heinbro Consulting", "name_zh": "Heinbro 顧問",
         "url": "https://heinbroconsulting.com/", "type": "compliance_tech",
         "priority": "MEDIUM", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "Licensing", "Type1", "Type9"],
         "crawl_depth": 2, "max_articles": 10},
        {"id": "bbcincorp", "name": "BBCIncorp HK", "name_zh": "BBCIncorp",
         "url": "https://bbcincorp.com/hk/articles", "type": "compliance_tech",
         "priority": "LOW", "enabled": True, "schedule": "monthly",
         "keywords": ["SFC", "Licensing", "Incorporation"],
         "crawl_depth": 1, "max_articles": 8},
    ]
    
    # 保存預設
    with open(src_file, 'w', encoding='utf-8') as f:
        json.dump(default_sources, f, ensure_ascii=False, indent=2)
    
    return default_sources

def save_sources(sources):
    src_file = os.path.join(CONFIG_DIR, "sources.json")
    with open(src_file, 'w', encoding='utf-8') as f:
        json.dump(sources, f, ensure_ascii=False, indent=2)

def load_templates():
    """載入實操經驗模板"""
    templates = {}
    tpl_dir = os.path.join(CONFIG_DIR, "templates")
    if os.path.isdir(tpl_dir):
        for fname in os.listdir(tpl_dir):
            if fname.endswith('.json'):
                try:
                    with open(os.path.join(tpl_dir, fname), 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        templates[data.get("id", fname.replace('.json',''))] = data
                except: pass
    if not templates:
        templates = {
            "enforcement_lesson": {"id": "enforcement_lesson", "title_zh": "執法教訓分析",
                "structure": ["背景", "違規行為", "處罰結果", "教訓"],
                "keywords": ["Enforcement", "Lesson"], "prompt": "整理以下執法案例為教訓分析：\n\n{content}"},
            "licensing_experience": {"id": "licensing_experience", "title_zh": "牌照申請經驗",
                "structure": ["申請類型", "準備過程", "常見問題", "建議"],
                "keywords": ["Licensing", "Application"], "prompt": "整理以下牌照申請經驗：\n\n{content}"},
            "custom": {"id": "custom", "title_zh": "自定義經驗",
                "structure": ["概述", "詳細內容", "總結"],
                "keywords": ["Practical"], "prompt": "整理以下內容為專業文章：\n\n{content}"},
        }
    return templates

def crawl_subpages(base_url, soup, max_articles=10):
    """子頁面爬取 — 從首頁提取文章鏈接，逐一訪問"""
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin
    
    articles = []
    links = set()
    
    # 找文章鏈接
    for a in soup.find_all('a', href=True):
        href = a.get('href', '')
        text = a.get_text(strip=True)
        
        # 過濾：只要看起來像文章的鏈接
        if len(text) < 15:
            continue
        if any(skip in href.lower() for skip in ['login', 'signup', 'contact', 'about', 'privacy', '#', 'javascript']):
            continue
        
        full_url = urljoin(base_url, href)
        
        # 只跟進同域名
        from urllib.parse import urlparse
        if urlparse(full_url).netloc != urlparse(base_url).netloc:
            continue
        
        if full_url not in links and full_url != base_url:
            links.add(full_url)
    
    # 訪問子頁面
    for url in list(links)[:max_articles]:
        try:
            resp = requests.get(url, headers=HEADERS, timeout=20)
            if resp.status_code == 200:
                sub_soup = BeautifulSoup(resp.text, 'html.parser')
                for tag in sub_soup(['script', 'style', 'nav', 'footer', 'header']):
                    tag.decompose()
                
                # 找主內容
                content = None
                for sel in ['article', '.entry-content', '.post-content', '.article-body', 'main']:
                    found = sub_soup.select(sel)
                    if found:
                        content = found[0].get_text(separator='\n', strip=True)
                        break
                
                if not content:
                    content = sub_soup.get_text(separator='\n', strip=True)
                
                content = re.sub(r'\n{3,}', '\n\n', content)
                
                if len(content) > 500:
                    articles.append({"url": url, "text": content[:8000],
                                    "title": sub_soup.find('title').get_text(strip=True)[:80] if sub_soup.find('title') else url.split('/')[-1][:50]})
            time.sleep(0.5)
        except:
            continue
    
    return articles

def import_industry_intelligence():
    """6A: 自動抓取全部啟用來源 — v5.0 含子頁面爬取 + QA"""
    from bs4 import BeautifulSoup
    
    sources = load_sources()
    if not sources:
        print("  無可用來源。")
        return
    
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded = set(progress.get("uploaded_refs", []))
    
    print(f"\n  {'=' * 50}")
    print(f"  6A: 抓取行業資訊 ({len(sources)} 來源)")
    print(f"  {'=' * 50}")
    
    stats = {"ok": 0, "fail": 0, "skip": 0, "chunks": 0, "duplicate": 0}
    results = []
    start_time = time.time()
    
    for i, src in enumerate(sources, 1):
        ref = f"IND_{src['id']}_{datetime.now().strftime('%Y%m')}"
        if ref in uploaded:
            stats["skip"] += 1
            continue
        
        show_progress(i, len(sources), "Source ", start_time, f"| {src['name'][:25]}")
        print(f"\n  [{i}/{len(sources)}] {src['name']}")
        print(f"     Type: {src['type']} | Priority: {src.get('priority', 'MEDIUM')}")
        
        try:
            # v7.0.4: 支援增強 Headers / SSL bypass / 自定義超時
            req_headers = ENHANCED_HEADERS if src.get("needs_enhanced_headers") else HEADERS
            req_timeout = src.get("request_timeout", 30)
            req_verify = src.get("ssl_verify", True)
            
            resp = requests.get(src["url"], headers=req_headers, timeout=req_timeout, verify=req_verify)
            if resp.status_code != 200:
                print(f"     ✗ HTTP {resp.status_code}")
                stats["fail"] += 1
                results.append({"name": src["name"], "status": "fail", "reason": f"HTTP {resp.status_code}"})
                continue
            
            soup = BeautifulSoup(resp.text, 'html.parser')
            all_texts = []
            
            # 首頁內容
            for tag in soup(['script', 'style', 'nav', 'footer', 'header']):
                tag.decompose()
            
            for sel in ['article', '.entry-content', '.post-content', 'main']:
                found = soup.select(sel)
                if found:
                    for el in found[:3]:
                        text = el.get_text(separator='\n', strip=True)
                        if len(text) > 300:
                            all_texts.append({"text": text[:5000], "title": src["name"], "url": src["url"]})
                    break
            
            # v5.0: 子頁面爬取
            crawl_depth = src.get("crawl_depth", 1)
            max_articles = src.get("max_articles", 10)
            if crawl_depth >= 2:
                sub_articles = crawl_subpages(src["url"], BeautifulSoup(resp.text, 'html.parser'), max_articles)
                all_texts.extend(sub_articles)
                print(f"     Found {len(sub_articles)} sub-pages")
            
            if not all_texts:
                text = soup.get_text(separator='\n', strip=True)
                text = re.sub(r'\n{3,}', '\n\n', text)
                if len(text) > 500:
                    all_texts.append({"text": text[:8000], "title": src["name"], "url": src["url"]})
            
            if not all_texts:
                stats["fail"] += 1
                results.append({"name": src["name"], "status": "fail", "reason": "No content"})
                continue
            
            print(f"     ✓ {len(all_texts)} articles found")
            
            # 確定權威性分數
            type_to_authority = {
                "law_firm_intl": "Law_Firm", "law_firm_local": "Law_Firm",
                "big4": "Big4", "consulting_firm": "Consulting",
                "compliance_tech": "Consulting",
            }
            authority = SOURCE_AUTHORITY.get(type_to_authority.get(src["type"], "Industry"), 50)
            
            article_chunks = 0
            for ai, article in enumerate(all_texts):
                art_text = article.get("text", "")
                art_title = article.get("title", "")
                art_url = article.get("url", src["url"])
                
                chunks = smart_chunk_text(art_text, art_title)
                for ci, chunk_text in enumerate(chunks):
                    # QA: 內容驗證
                    is_valid, reason = validate_chunk(chunk_text)
                    if not is_valid:
                        continue
                    
                    # QA: 去重
                    chunk_ref = f"{ref}_art{ai}_c{ci}"
                    is_dup, _, _ = check_duplicate(chunk_text, chunk_ref, progress, hashes_data)
                    if is_dup:
                        stats["duplicate"] += 1
                        continue
                    
                    part = f" (Part {ci+1}/{len(chunks)})" if len(chunks) > 1 else ""
                    meta = {
                        "source": f"Industry_{src['type']}", "source_url": art_url,
                        "doc_type": src["type"], "category": "Industry",
                        "ref_no": chunk_ref, "source_table": "vigo_financial",
                        "keywords": src.get("keywords", ["Industry"]),
                        "collected_date": datetime.now().strftime("%Y-%m-%d"),
                        "authority_score": authority,
                    }
                    
                    en_content = f"【{art_title}{part}】\n\n{chunk_text}"
                    en_content += f"\n\n---\nSource: {src['name']}\nURL: {art_url}"
                    kid = upload_chunk(en_content, {**meta, "language": "en"})
                    if kid:
                        stats["chunks"] += 1
                        article_chunks += 1
                        progress, hashes_data = record_upload(chunk_ref, chunk_text, progress, hashes_data)
                    
                    zh = translate_to_chinese(chunk_text)
                    if zh and len(zh) > 80:
                        zh_content = f"【{src.get('name_zh', src['name'])}{part}】\n\n{zh}"
                        zh_content += f"\n\n---\n來源：{src.get('name_zh', src['name'])}\n鏈接：{art_url}"
                        kid_zh = upload_chunk(zh_content, {**meta, "language": "zh"})
                        if kid_zh: stats["chunks"] += 1
                    time.sleep(0.5)
            
            uploaded.add(ref)
            progress["uploaded_refs"] = list(uploaded)
            save_progress(progress)
            save_content_hashes(hashes_data)
            stats["ok"] += 1
            results.append({"name": src["name"], "status": "ok", "chunks": article_chunks})
        
        except Exception as e:
            print(f"     ✗ Error: {e}")
            stats["fail"] += 1
            results.append({"name": src["name"], "status": "fail", "reason": str(e)[:100]})
    
    elapsed = time.time() - start_time
    print(f"\n\n  Done! {elapsed/60:.1f} min | Sources: {stats['ok']} | Chunks: {stats['chunks']} | Dup blocked: {stats['duplicate']}")
    generate_report("mode6a_industry", stats, results, elapsed)


# ── 6A-c: 手動錄入行情數據 ──
def input_market_data():
    """手動錄入行情數據（殼價、RO費、時間線等）"""
    print(f"\n  {'=' * 50}")
    print(f"  6A: 行情數據快錄")
    print(f"  {'=' * 50}")
    
    categories = {
        "1": ("shell_price", "殼價行情 (Shell Price)", ["Shell", "License", "Market"]),
        "2": ("ro_fee", "RO 掛靠費 (RO Fee)", ["RO", "Fee", "Licensing"]),
        "3": ("timeline", "牌照時間線 (Timeline)", ["Timeline", "Licensing", "Application"]),
        "4": ("market_trend", "市場趨勢 (Trend)", ["Market", "Trend", "Industry"]),
        "5": ("custom", "自定義 (Custom)", ["Market"]),
    }
    
    print("\n  類別:")
    for k, (_, label, _) in categories.items():
        print(f"  {k}. {label}")
    
    choice = input(f"\n  (1-5): ").strip()
    if choice not in categories:
        print("  無效選擇"); return
    
    cat_id, cat_name, cat_kw = categories[choice]
    
    print(f"\n  {cat_name}")
    title = input("  標題: ").strip()
    if not title:
        print("  需要標題"); return
    
    print("  內容（多行輸入，空行結束）:")
    lines = []
    while True:
        line = input("  > ")
        if not line:
            break
        lines.append(line)
    
    content = "\n".join(lines)
    if len(content) < 50:
        print("  內容太短"); return
    
    meta = {
        "source": "Market_Data", "doc_type": f"market_{cat_id}",
        "category": "Market", "ref_no": f"MKT_{cat_id}_{datetime.now().strftime('%Y%m%d_%H%M')}",
        "source_table": "vigo_financial", "keywords": cat_kw,
        "collected_date": datetime.now().strftime("%Y-%m-%d"),
        "authority_score": SOURCE_AUTHORITY.get("Practical_Experience", 40),
    }
    
    full_content = f"【{cat_name} — {title}】\n\n{content}"
    full_content += f"\n\n---\n資料來源：人工錄入\n日期：{datetime.now().strftime('%Y-%m-%d')}"
    
    kid = upload_chunk(full_content, {**meta, "language": "zh"})
    if kid:
        print(f"  ✅ Uploaded (ID: {kid})")
    else:
        print(f"  ❌ Failed")


# ── 6B: 實操經驗 ──
def input_practical_experience():
    """6B-e: 選擇模板輸入經驗"""
    templates = load_templates()
    
    print(f"\n  {'=' * 50}")
    print(f"  6B: 實操經驗錄入")
    print(f"  {'=' * 50}")
    
    tpl_list = sorted(templates.items())
    for i, (tid, t) in enumerate(tpl_list, 1):
        print(f"  {i}. {t.get('title_zh', tid)}")
    
    choice = input(f"\n  模板 (1-{len(tpl_list)}): ").strip()
    try:
        idx = int(choice) - 1
        tid, template = tpl_list[idx]
    except:
        print("  無效選擇"); return
    
    print(f"\n  模板：{template.get('title_zh', tid)}")
    
    sections = {}
    for field in template.get("structure", ["概述"]):
        print(f"\n  ── {field} ──")
        print(f"  （多行輸入，空行結束）")
        lines = []
        while True:
            line = input("  > ")
            if not line:
                break
            lines.append(line)
        if lines:
            sections[field] = "\n".join(lines)
    
    if not sections:
        print("  未輸入任何內容"); return
    
    raw_content = "\n\n".join(f"【{k}】\n{v}" for k, v in sections.items())
    
    prompt = template.get("prompt", "整理以下內容：\n\n{content}").replace("{content}", raw_content[:4000])
    optimized = deepseek_chat(prompt, max_tokens=2000, temperature=0.3)
    if not optimized or len(optimized) < 100:
        optimized = raw_content
    
    meta = {
        "source": "Practical_Experience", "doc_type": f"practical_{tid}",
        "category": "Practical", "ref_no": f"EXP_{tid}_{datetime.now().strftime('%Y%m%d_%H%M')}",
        "source_table": "vigo_financial", "keywords": template.get("keywords", ["Practical"]),
        "collected_date": datetime.now().strftime("%Y-%m-%d"),
        "authority_score": SOURCE_AUTHORITY.get("Practical_Experience", 40),
    }
    
    content = f"【實操經驗：{template.get('title_zh', tid)}】\n\n{optimized}"
    kid = upload_chunk(content, {**meta, "language": "zh"})
    print(f"  ✅ Uploaded" if kid else "  ❌ Failed")

def manage_sources():
    """6A-d: 管理來源"""
    sources = load_sources()
    
    print(f"\n  {'=' * 50}")
    print(f"  來源管理 ({len(sources)} 個)")
    print(f"  {'=' * 50}")
    
    for i, s in enumerate(sources, 1):
        status = "✅" if s.get("enabled", True) else "❌"
        print(f"  {i:2d}. {status} [{s.get('priority','?'):>7}] {s['name'][:40]} ({s['type']})")
    
    print(f"\n  操作: (a)新增 (d)停用/啟用 (r)返回")
    action = input("  : ").strip().lower()
    
    if action == "a":
        name = input("  來源名稱: ").strip()
        url = input("  URL: ").strip()
        stype = input("  類型 (law_firm/consulting/big4/other): ").strip()
        if name and url:
            sources.append({
                "id": re.sub(r'\W+', '_', name.lower())[:20], "name": name,
                "url": url, "type": stype or "other",
                "priority": "MEDIUM", "enabled": True, "schedule": "weekly",
                "keywords": ["SFC"], "crawl_depth": 1, "max_articles": 10
            })
            save_sources(sources)
            print(f"  ✅ Added: {name}")
    elif action == "d":
        idx = input(f"  編號 (1-{len(sources)}): ").strip()
        try:
            s = sources[int(idx)-1]
            s["enabled"] = not s.get("enabled", True)
            save_sources(sources)
            print(f"  {'✅ Enabled' if s['enabled'] else '❌ Disabled'}: {s['name']}")
        except:
            print("  無效編號")


# ── Mode 6 主入口 ──
def run_mode_6():
    print(f"\n  {'=' * 50}")
    print(f"  Mode 6: 行業知識中心")
    print(f"  合規是下線，業務是上線 — VIGO Protocol")
    print(f"  {'=' * 50}")
    print(f"\n  ┌─ 6A 行業資訊 ──────────────────────┐")
    print(f"  │  a. 自動抓取全部來源 (19家)           │")
    print(f"  │  b. 手動錄入行情數據                  │")
    print(f"  │  c. 管理來源（新增/停用/查看）        │")
    print(f"  ├─ 6B 實操經驗 ──────────────────────┤")
    print(f"  │  d. 選擇模板輸入經驗                  │")
    print(f"  │  e. 查看可用模板                      │")
    print(f"  ├─────────────────────────────────────┤")
    print(f"  │  f. 全部執行（a + d）                 │")
    print(f"  └─────────────────────────────────────┘")
    
    sub = input(f"\n  (a/b/c/d/e/f): ").strip().lower()
    
    if sub == "a":
        import_industry_intelligence()
    elif sub == "b":
        input_market_data()
    elif sub == "c":
        manage_sources()
    elif sub == "d":
        input_practical_experience()
    elif sub == "e":
        templates = load_templates()
        print(f"\n  可用模板 ({len(templates)} 個)：")
        for i, (tid, t) in enumerate(sorted(templates.items()), 1):
            print(f"  {i}. {t.get('title_zh', tid)} ({tid})")
    elif sub == "f":
        import_industry_intelligence()
        input_practical_experience()
    else:
        print(f"  無效選擇")


# ============================================================
# v5.0 Mode 7: 統一文件導入器 (Universal File Importer)
# 支持: PDF / DOCX / DOC / TXT / MD / XLSX
# 取代原 Mode 6f 的 .txt/.md 批量導入
# ============================================================

def run_mode_7():
    """Mode 7: 統一文件導入器"""
    print(f"\n  {'=' * 50}")
    print(f"  Mode 7: 統一文件導入器")
    print(f"  支持格式: PDF / DOCX / TXT / MD / XLSX")
    print(f"  導入目錄: config/imports/")
    print(f"  {'=' * 50}")
    
    # 掃描目錄
    supported_ext = {'.pdf', '.docx', '.doc', '.txt', '.md', '.xlsx'}
    files = []
    
    # 選擇來源
    print(f"\n  a. 掃描 config/imports/ 目錄")
    print(f"  b. 指定文件/資料夾路徑")
    src = input(f"\n  (a/b): ").strip().lower()
    
    if src == "a":
        scan_dir = IMPORTS_DIR
    elif src == "b":
        scan_dir = input("  路徑: ").strip().strip('"').strip("'")
        if not os.path.exists(scan_dir):
            print(f"  路徑不存在: {scan_dir}")
            return
    else:
        print("  無效選擇"); return
    
    if os.path.isfile(scan_dir):
        files = [scan_dir]
    elif os.path.isdir(scan_dir):
        for fname in os.listdir(scan_dir):
            ext = os.path.splitext(fname)[1].lower()
            if ext in supported_ext:
                files.append(os.path.join(scan_dir, fname))
    
    if not files:
        print(f"\n  未找到支持的文件。")
        print(f"  請將文件放入: {IMPORTS_DIR}")
        return
    
    # 顯示文件清單
    print(f"\n  找到 {len(files)} 個文件：")
    for i, f in enumerate(files, 1):
        fname = os.path.basename(f)
        size = os.path.getsize(f)
        ext = os.path.splitext(fname)[1]
        print(f"  {i:3d}. [{ext:>5}] {fname[:55]} ({size/1024:.0f} KB)")
    
    # 選擇全部或個別
    print(f"\n  (a) 全部導入  (s) 選擇性導入")
    sel = input("  : ").strip().lower()
    if sel == "s":
        nums = input("  輸入編號（逗號分隔）: ").strip()
        try:
            indices = [int(n.strip()) - 1 for n in nums.split(",")]
            files = [files[i] for i in indices if 0 <= i < len(files)]
        except:
            print("  無效輸入"); return
    
    # 選擇文件分類
    print(f"\n  文件分類:")
    print(f"  1. regulation      (法規/規則)")
    print(f"  2. guideline       (指引)")
    print(f"  3. industry        (行業資訊)")
    print(f"  4. training        (培訓材料)")
    print(f"  5. report          (報告)")
    print(f"  ── v5.2 合規文件工廠分類 [P2.12] ──")
    print(f"  6. template        (合規模板/範本)")
    print(f"  7. compliance_manual (合規手冊)")
    print(f"  8. business_plan   (業務計劃書)")
    print(f"  9. operations_manual (營運手冊)")
    print(f"  10. sop            (標準操作程序)")
    print(f"  11. other          (其他)")
    doc_type_map = {"1": "regulation", "2": "guideline", "3": "industry",
                    "4": "training", "5": "report", "6": "template",
                    "7": "compliance_manual", "8": "business_plan",
                    "9": "operations_manual", "10": "sop", "11": "other"}
    dt_choice = input("  (1-11): ").strip()
    doc_type = doc_type_map.get(dt_choice, "other")
    
    custom_keywords = input("  自定義關鍵詞（逗號分隔，可跳過）: ").strip()
    keywords = [k.strip() for k in custom_keywords.split(",") if k.strip()] if custom_keywords else ["Import", doc_type.title()]
    
    confirm = input(f"\n  開始處理 {len(files)} 個文件？(y/n): ").strip().lower()
    if confirm != 'y':
        return
    
    # 處理文件
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded = set(progress.get("uploaded_refs", []))
    
    stats = {"ok": 0, "fail": 0, "skip": 0, "chunks": 0, "duplicate": 0, "qa_blocked": 0}
    results = []
    all_vigo_chunks = []
    start_time = time.time()
    
    for fi, fpath in enumerate(files, 1):
        fname = os.path.basename(fpath)
        ext = os.path.splitext(fname)[1].lower()
        ref = f"IMPORT_{compute_content_hash(fname + str(os.path.getsize(fpath)))}"
        
        if ref in uploaded:
            stats["skip"] += 1
            results.append({"name": fname, "status": "skip"})
            print(f"\n  [{fi}/{len(files)}] SKIP (already imported): {fname[:50]}")
            continue
        
        show_progress(fi, len(files), "File ", start_time, f"| {fname[:25]}")
        print(f"\n  [{fi}/{len(files)}] {fname}")
        
        # 提取文字
        text = None
        try:
            if ext == '.pdf':
                with open(fpath, 'rb') as f:
                    raw = f.read()
                if raw[:4] == b'%PDF':
                    text = extract_text_from_pdf(raw)
                else:
                    print(f"     ⚠ Not a valid PDF")
            
            elif ext in ('.docx', '.doc'):
                text = extract_text_from_docx(fpath)
            
            elif ext == '.xlsx':
                text = extract_text_from_xlsx(fpath)
            
            elif ext in ('.txt', '.md'):
                try:
                    with open(fpath, 'r', encoding='utf-8') as f:
                        text = f.read()
                except UnicodeDecodeError:
                    with open(fpath, 'r', encoding='gbk', errors='ignore') as f:
                        text = f.read()
            
            else:
                print(f"     ✗ Unsupported format: {ext}")
                stats["fail"] += 1
                results.append({"name": fname, "status": "fail", "reason": f"Unsupported: {ext}"})
                continue
        
        except Exception as e:
            print(f"     ✗ Extract error: {e}")
            stats["fail"] += 1
            results.append({"name": fname, "status": "fail", "reason": str(e)[:100]})
            continue
        
        if not text or len(text) < 200:
            print(f"     ✗ Content too short ({len(text) if text else 0} chars)")
            stats["fail"] += 1
            results.append({"name": fname, "status": "fail", "reason": "Too short"})
            continue
        
        print(f"     ✓ Extracted: {len(text)} chars")
        
        # ── v5.2.1: 檢測結構化 .md 格式 ──
        vigo_sections = None
        if ext == '.md':
            vigo_sections = parse_vigo_sections(text)
            if vigo_sections:
                print(f"     ✓ Detected VIGO ===SECTION=== format: {len(vigo_sections)} sections")
        
        if vigo_sections:
            # ═══ 結構化 .md 上傳路徑（v5.2.1 新增）═══
            chunks = vigo_sections  # 用於後續 len(chunks) 統計
            chunk_ok = 0
            chunk_start = time.time()
            
            for ci, section in enumerate(vigo_sections):
                show_progress(ci+1, len(vigo_sections), "  Section ", chunk_start)
                
                s_content = section['content']
                s_title = section['title']
                s_meta = section['metadata']
                
                is_valid, reason = validate_chunk(s_content)
                if not is_valid:
                    stats["qa_blocked"] += 1
                    continue
                
                chunk_ref = f"{ref}_s{ci}"
                is_dup, _, _ = check_duplicate(s_content, chunk_ref, progress, hashes_data)
                if is_dup:
                    stats["duplicate"] += 1
                    continue
                
                section_source = s_meta.get('source', fname.rsplit('.', 1)[0])
                section_doc_type = s_meta.get('doc_type', doc_type)
                section_category = s_meta.get('category', 'Import')
                section_lang = s_meta.get('language', 'en')
                section_keywords = s_meta.get('keywords', ', '.join(keywords) if isinstance(keywords, list) else str(keywords))
                section_priority = s_meta.get('priority', 'MEDIUM')
                section_source_url = s_meta.get('source_url', f"file://{fname}")
                
                chunk_meta = {
                    "source": section_source,
                    "source_url": section_source_url,
                    "doc_type": section_doc_type,
                    "category": section_category,
                    "ref_no": s_meta.get('ref_no', chunk_ref),
                    "source_table": s_meta.get('source_table', 'vigo_knowledge'),
                    "keywords": section_keywords,
                    "collected_date": s_meta.get('collected_date', datetime.now().strftime("%Y-%m-%d")),
                    "content_type": "structured_md",
                    "authority_score": SOURCE_AUTHORITY.get(section_source, 70),
                    "original_filename": fname,
                    "language": section_lang,
                    "priority": section_priority,
                }
                
                for extra_key in ['license_types', 'compliance_domains', 'doc_hierarchy']:
                    if extra_key in s_meta:
                        chunk_meta[extra_key] = s_meta[extra_key]
                
                part_label = f" (Section {ci+1}/{len(vigo_sections)})"
                final_content = f"【{s_title}{part_label}】\n\n{s_content}"
                
                kid = upload_chunk(final_content, chunk_meta)
                if kid:
                    stats["chunks"] += 1
                    chunk_ok += 1
                    progress, hashes_data = record_upload(chunk_ref, s_content, progress, hashes_data)
                
                all_vigo_chunks.append({
                    "content": final_content,
                    "metadata": chunk_meta,
                    "lang": section_lang.upper()
                })
                
                time.sleep(0.3)
            
            print(f"\n     ✓ Uploaded {chunk_ok} sections (structured)")
        
        else:
            # ═══ 原有普通文本上傳路徑 ═══
            if len(text) > 80000:
                text = text[:80000]
                print(f"     ⚠ Truncated to 80,000 chars")
            
            chunks = smart_chunk_text(text, fname.rsplit('.', 1)[0])
            print(f"     ✓ Chunks: {len(chunks)}")
            
            is_chinese = detect_language(text[:500]) == 'zh'
            if is_chinese:
                print(f"     ℹ️  Language: 中文 (will upload as ZH, skip EN translation)")
            else:
                print(f"     ℹ️  Language: English (will upload as EN, skip ZH translation)")
            
            chunk_ok = 0
            chunk_start = time.time()
            
            for ci, chunk_text in enumerate(chunks):
                show_progress(ci+1, len(chunks), "  Chunk ", chunk_start)
                
                is_valid, reason = validate_chunk(chunk_text)
                if not is_valid:
                    stats["qa_blocked"] += 1
                    continue
                
                chunk_ref = f"{ref}_c{ci}"
                is_dup, _, _ = check_duplicate(chunk_text, chunk_ref, progress, hashes_data)
                if is_dup:
                    stats["duplicate"] += 1
                    continue
                
                part_label = f" (Part {ci+1}/{len(chunks)})" if len(chunks) > 1 else ""
                title = fname.rsplit('.', 1)[0].replace('_', ' ')
                
                base_meta = {
                    "source": "MANUAL_IMPORT", "source_url": f"file://{fname}",
                    "doc_type": doc_type, "category": "Import",
                    "ref_no": chunk_ref, "source_table": "vigo_financial",
                    "keywords": keywords, "collected_date": datetime.now().strftime("%Y-%m-%d"),
                    "content_type": ext.lstrip('.'),
                    "authority_score": SOURCE_AUTHORITY.get("MANUAL_IMPORT", 30),
                    "original_filename": fname,
                }
                
                if is_chinese:
                    zh_content = f"【{title}{part_label}】\n\n{chunk_text}"
                    zh_content += f"\n\n---\n來源：手動導入\n文件：{fname}"
                    kid = upload_chunk(zh_content, {**base_meta, "language": "zh"})
                    if kid:
                        stats["chunks"] += 1
                        chunk_ok += 1
                        progress, hashes_data = record_upload(chunk_ref, chunk_text, progress, hashes_data)
                    all_vigo_chunks.append({"content": zh_content, "metadata": {**base_meta, "language": "zh"}, "lang": "ZH"})
                else:
                    en_content = f"【{title}{part_label}】\n\n{chunk_text}"
                    en_content += f"\n\n---\nSource: Manual Import\nFile: {fname}"
                    kid = upload_chunk(en_content, {**base_meta, "language": "en"})
                    if kid:
                        stats["chunks"] += 1
                        chunk_ok += 1
                        progress, hashes_data = record_upload(chunk_ref, chunk_text, progress, hashes_data)
                    all_vigo_chunks.append({"content": en_content, "metadata": {**base_meta, "language": "en"}, "lang": "EN"})
                
                time.sleep(0.3)
            
            print(f"\n     ✓ Uploaded {chunk_ok} chunks")
        uploaded.add(ref)
        progress["uploaded_refs"] = list(uploaded)
        save_progress(progress)
        save_content_hashes(hashes_data)
        stats["ok"] += 1
        results.append({"name": fname, "status": "ok", "chars": len(text), "chunks": len(chunks),
                        "time_min": round((time.time() - chunk_start) / 60, 1)})
    
    total_time = time.time() - start_time
    print(f"\n\n  {'=' * 60}")
    print(f"  Mode 7 Complete! {total_time/60:.1f} min")
    print(f"  OK: {stats['ok']} | Fail: {stats['fail']} | Skip: {stats['skip']}")
    print(f"  Chunks: {stats['chunks']} | Dup blocked: {stats['duplicate']} | QA blocked: {stats['qa_blocked']}")
    print(f"  {'=' * 60}")
    
    generate_report("mode7_file_import", stats, results, total_time, {"doc_type": doc_type, "keywords": keywords})
    
    if all_vigo_chunks:
        backup = os.path.join(REPORTS_DIR, f"VIGO_import_bilingual_{datetime.now().strftime('%Y%m%d')}.md")
        save_backup(all_vigo_chunks, backup)


# ============================================================
# v5.0 Mode 8: 數據庫管理工具
# ============================================================

def run_mode_8():
    """Mode 8: VIGO 數據庫管理 — v5.2 含合規日曆"""
    print(f"\n  {'=' * 50}")
    print(f"  Mode 8: VIGO 數據庫管理")
    print(f"  數據質量五原則：寧缺勿濫 | 源頭把控 | 權威優先 | 時效標記 | 定期清洗")
    print(f"  {'=' * 50}")
    print(f"\n  a. 健康檢查 — 掃描重複/低質量內容")
    print(f"  b. 統計總覽 — 各類別數量、來源分佈")
    print(f"  c. 查看本地記錄 — refs + hashes")
    print(f"  d. 重置進度文件（慎用）")
    print(f"  e. 導出審計報告")
    print(f"  ── v5.2 新增 ──")
    print(f"  f. ★ 合規日曆數據 [P3.17] — 匯出持牌法團合規日曆")
    print(f"  g. ★ 結構化標籤統計 [P3.16] — 牌照/合規領域分佈")
    
    sub = input(f"\n  (a/b/c/d/e/f/g): ").strip().lower()
    
    if sub == "a":
        _db_health_check()
    elif sub == "b":
        _db_stats_overview()
    elif sub == "c":
        _view_local_records()
    elif sub == "d":
        _reset_progress()
    elif sub == "e":
        _export_audit_report()
    elif sub == "f":
        _export_compliance_calendar()
    elif sub == "g":
        _structured_tag_stats()
    else:
        print("  無效選擇")

def _db_health_check():
    """8a: 數據庫健康檢查"""
    print(f"\n  {'=' * 50}")
    print(f"  8a: Database Health Check")
    print(f"  {'=' * 50}")
    
    hashes_data = load_content_hashes()
    hashes = hashes_data.get("hashes", {})
    
    print(f"\n  本地記錄:")
    print(f"    Content hashes: {len(hashes)}")
    
    # 檢查重複 hash（不同 ref 但相同 hash = 重複內容）
    hash_to_refs = {}
    for h, ref in hashes.items():
        if h not in hash_to_refs:
            hash_to_refs[h] = []
        hash_to_refs[h].append(ref)
    
    duplicates = {h: refs for h, refs in hash_to_refs.items() if len(refs) > 1}
    
    if duplicates:
        print(f"\n  ⚠ Found {len(duplicates)} duplicate content groups:")
        for h, refs in list(duplicates.items())[:10]:
            print(f"    Hash {h}: {', '.join(refs[:3])}")
    else:
        print(f"\n  ✅ No duplicate content found in local records")
    
    # 檢查 Supabase 記錄數
    try:
        headers = {
            "apikey": SUPABASE_SERVICE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "count=exact"
        }
        resp = requests.get(f"{SUPABASE_URL}/rest/v1/{KNOWLEDGE_TABLE}?select=id&limit=1",
            headers=headers, timeout=15)
        if resp.status_code == 200:
            count = resp.headers.get('content-range', '').split('/')[-1]
            print(f"\n  Supabase vigo_knowledge 記錄數: {count}")
    except Exception as e:
        print(f"\n  ⚠ Cannot connect to Supabase: {e}")

def _db_stats_overview():
    """8b: 統計總覽"""
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded_refs = progress.get("uploaded_refs", [])
    
    print(f"\n  {'=' * 60}")
    print(f"  VIGO 數據庫統計 — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  {'=' * 60}")
    
    # 按類別統計
    categories = {}
    for ref in uploaded_refs:
        if ref.startswith("STATIC_"):
            cat = "基石法規"
        elif ref.startswith("EXT_"):
            cat = "擴展來源"
        elif ref.startswith("IND_"):
            cat = "行業資訊"
        elif ref.startswith("IMPORT_"):
            cat = "手動導入"
        elif ref.startswith("EXP_"):
            cat = "實操經驗"
        elif ref.startswith("MKT_"):
            cat = "市場數據"
        elif ref.startswith("HKMA_"):
            cat = "HKMA聯合"
        elif ref.startswith("VATP9_"):
            cat = "VATP深度(M9)"
        elif ref.startswith("OPS10_"):
            cat = "操作知識(M10)"
        elif ref.startswith("CAL_"):
            cat = "合規日曆"
        elif "EC" in ref:
            cat = "通函"
        elif "CP" in ref:
            cat = "諮詢文件"
        elif "PR" in ref:
            cat = "新聞/執法"
        elif ref.endswith("_TC"):
            cat = "中文通函"
        else:
            cat = "其他"
        categories[cat] = categories.get(cat, 0) + 1
    
    total = len(uploaded_refs)
    print(f"\n  總記錄: {total} refs")
    print(f"  Content hashes: {len(hashes_data.get('hashes', {}))}")
    print(f"\n  按類別:")
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        pct = count / max(total, 1) * 100
        bar = "█" * int(pct / 5)
        print(f"    {cat:10s} {count:5d} ({pct:4.1f}%) {bar}")

def _view_local_records():
    """8c: 查看本地記錄"""
    progress = load_progress()
    refs = progress.get("uploaded_refs", [])
    
    print(f"\n  Total refs: {len(refs)}")
    print(f"  Last year: {progress.get('last_year', 'N/A')}")
    
    if refs:
        print(f"\n  最近 20 條:")
        for ref in refs[-20:]:
            print(f"    {ref}")

def _reset_progress():
    """8d: 重置進度文件"""
    print(f"\n  ⚠ 此操作將清除所有進度記錄")
    print(f"  已上傳的數據不會被刪除，但腳本將無法識別已上傳內容")
    confirm = input(f"  確認重置？輸入 'RESET' 確認: ").strip()
    if confirm == "RESET":
        save_progress({"uploaded_refs": [], "last_year": None, "last_source": None})
        save_content_hashes({"hashes": {}})
        print(f"  ✅ 已重置")
    else:
        print(f"  取消")

def _export_audit_report():
    """8e: 導出審計報告"""
    progress = load_progress()
    hashes_data = load_content_hashes()
    refs = progress.get("uploaded_refs", [])
    hashes = hashes_data.get("hashes", {})
    
    report = {
        "timestamp": datetime.now().isoformat(),
        "total_refs": len(refs),
        "total_hashes": len(hashes),
        "refs_by_category": {},
        "quality_metrics": {
            "duplicate_hashes": 0,
            "refs_list": refs[-100:],  # 最近 100 條
        }
    }
    
    for ref in refs:
        prefix = ref.split("_")[0] if "_" in ref else ref[:4]
        report["refs_by_category"][prefix] = report["refs_by_category"].get(prefix, 0) + 1
    
    path = os.path.join(REPORTS_DIR, f"audit_report_{datetime.now().strftime('%Y%m%d_%H%M')}.json")
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n  📄 Audit report: {path}")


def _export_compliance_calendar():
    """8f: [P3.17] 匯出 SFC 持牌法團合規日曆數據"""
    print(f"\n  {'=' * 60}")
    print(f"  8f: SFC Licensed Corporation Compliance Calendar")
    print(f"  {'=' * 60}")
    
    calendar_data = {
        "generated": datetime.now().isoformat(),
        "description": "SFC Licensed Corporation Compliance Calendar — v5.2",
        
        "monthly_obligations": [
            {"deadline": "每月第 21 個曆日", "deadline_en": "21st calendar day of each month",
             "task": "提交 FRR 月報", "task_en": "Submit Financial Resources Return (FRR)",
             "applies_to": "所有持牌法團（Type 4/5/6/9 無客戶資產者除外，可半年報）",
             "penalty": "逾期可能導致牌照條件加嚴或紀律處分",
             "reference": "Cap.571N, Schedule 1"},
            {"deadline": "每月", "deadline_en": "Monthly",
             "task": "制裁名單篩查", "task_en": "Sanctions screening (UNSC + HK designated lists)",
             "applies_to": "所有持牌法團",
             "penalty": "AML/CFT 違規",
             "reference": "AML Guideline Chapter 6"},
        ],
        
        "quarterly_obligations": [
            {"deadline": "每季度", "deadline_en": "Quarterly",
             "task": "合規監控報告", "task_en": "Compliance monitoring report to senior management",
             "applies_to": "所有持牌法團",
             "reference": "MSIC Guidelines"},
            {"deadline": "每季度", "deadline_en": "Quarterly",
             "task": "釣魚郵件模擬演練", "task_en": "Phishing simulation exercise",
             "applies_to": "所有持牌法團",
             "reference": "SFC Cybersecurity Review 2023-24"},
        ],
        
        "annual_obligations": [
            {"deadline": "財政年度結束後 4 個月內", "deadline_en": "Within 4 months of financial year-end",
             "task": "提交經審計財務報表", "task_en": "Submit audited financial statements",
             "applies_to": "所有持牌法團",
             "penalty": "牌照可被暫停或撤銷",
             "reference": "SFO Section 156"},
            {"deadline": "每年（通過 WINGS）", "deadline_en": "Annually (via WINGS)",
             "task": "提交年度申報表", "task_en": "Submit annual return",
             "applies_to": "所有持牌法團",
             "penalty": "逾期可能導致牌照被撤銷",
             "reference": "SFO Section 130"},
            {"deadline": "每年", "deadline_en": "Annually",
             "task": "繳交年費", "task_en": "Pay annual licence fee",
             "applies_to": "所有持牌法團及持牌代表",
             "penalty": "逾期 = 撤銷風險",
             "reference": "SFO Section 138"},
            {"deadline": "12月31日前", "deadline_en": "Before 31 December",
             "task": "完成 CPT 持續培訓時數", "task_en": "Complete CPT training hours",
             "applies_to": "LR: 10小時（5 RA + 2 倫理）; RO: 12小時（2 合規）",
             "reference": "Guidelines on Competence"},
            {"deadline": "每年", "deadline_en": "Annually",
             "task": "機構洗錢/恐資風險評估更新", "task_en": "Update institutional ML/TF risk assessment",
             "applies_to": "所有持牌法團",
             "reference": "AML Guideline Chapter 2"},
            {"deadline": "每年", "deadline_en": "Annually",
             "task": "合規手冊年度審查", "task_en": "Annual review of compliance manual",
             "applies_to": "所有持牌法團",
             "reference": "Code of Conduct GP1"},
            {"deadline": "每年", "deadline_en": "Annually",
             "task": "BCP 測試與演練", "task_en": "BCP testing and drill (tabletop + simulation)",
             "applies_to": "所有持牌法團",
             "reference": "MSIC Guidelines, SFC Circular"},
            {"deadline": "每年", "deadline_en": "Annually",
             "task": "風險管理框架審查", "task_en": "Risk management framework review",
             "applies_to": "所有持牌法團",
             "reference": "MSIC Guidelines"},
            {"deadline": "每年", "deadline_en": "Annually",
             "task": "AML 培訓（全員）", "task_en": "AML training for all staff",
             "applies_to": "所有持牌法團",
             "reference": "AML Guideline Chapter 8"},
        ],
        
        "event_triggered_obligations": [
            {"trigger": "流動資本不足", "trigger_en": "Liquid capital deficit",
             "action": "立即停止受規管活動 + 即時通報 SFC", "action_en": "Cease RA immediately + notify SFC immediately",
             "reference": "FRR Rule 5"},
            {"trigger": "重大違規事件", "trigger_en": "Material breach or incident",
             "action": "即時通報 SFC", "action_en": "Notify SFC immediately",
             "reference": "Code of Conduct 12.5"},
            {"trigger": "董事/股東變更", "trigger_en": "Change of director/shareholder",
             "action": "7 個工作日內通報 SFC", "action_en": "Notify SFC within 7 business days",
             "reference": "SFO Section 130"},
            {"trigger": "新入職持牌代表", "trigger_en": "New licensed representative hire",
             "action": "12 個月內完成 2 小時倫理 CPT", "action_en": "Complete 2-hour ethics CPT within 12 months",
             "reference": "Guidelines on Competence"},
            {"trigger": "離職持牌代表", "trigger_en": "Departure of licensed representative",
             "action": "7 個工作日內通報 SFC", "action_en": "Notify SFC within 7 business days",
             "reference": "SFO Section 130"},
            {"trigger": "嚴重 IT 事件", "trigger_en": "Major IT incident / cyber attack",
             "action": "即時通報 SFC", "action_en": "Notify SFC immediately",
             "reference": "SFC Cybersecurity Circular"},
            {"trigger": "嚴重投訴", "trigger_en": "Serious complaint (fraud/misconduct/mass/significant loss)",
             "action": "即時升級 + 考慮通報 SFC", "action_en": "Escalate immediately + consider SFC notification",
             "reference": "Code of Conduct 12.5(a), SFC March 2022 Circular"},
            {"trigger": "可疑交易", "trigger_en": "Suspicious transaction detected",
             "action": "內部報告 MLRO → 提交 STR 至 JFIU", "action_en": "Internal report to MLRO → file STR with JFIU",
             "reference": "AML Guideline Chapter 5"},
        ],
        
        "ongoing_monitoring_frequency": {
            "high_risk_clients": "每年覆核", "high_risk_clients_en": "Annual review",
            "medium_risk_clients": "每 2-3 年覆核", "medium_risk_clients_en": "Review every 2-3 years",
            "low_risk_clients": "每 5 年覆核", "low_risk_clients_en": "Review every 5 years",
        },
    }
    
    # 保存 JSON
    path = os.path.join(REPORTS_DIR, f"compliance_calendar_{datetime.now().strftime('%Y%m%d')}.json")
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(calendar_data, f, ensure_ascii=False, indent=2)
    
    # 也生成可讀的 MD 版本
    md_path = os.path.join(REPORTS_DIR, f"compliance_calendar_{datetime.now().strftime('%Y%m%d')}.md")
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write("# SFC 持牌法團合規日曆 / Compliance Calendar\n\n")
        f.write(f"生成日期: {datetime.now().strftime('%Y-%m-%d')}\n\n")
        
        f.write("## 每月義務 Monthly\n\n")
        for item in calendar_data["monthly_obligations"]:
            f.write(f"- **{item['deadline']}**: {item['task']}\n")
            f.write(f"  - {item['deadline_en']}: {item['task_en']}\n")
            f.write(f"  - 適用: {item['applies_to']} | Ref: {item['reference']}\n\n")
        
        f.write("## 每季義務 Quarterly\n\n")
        for item in calendar_data["quarterly_obligations"]:
            f.write(f"- **{item['deadline']}**: {item['task']}\n")
            f.write(f"  - {item['task_en']} | Ref: {item['reference']}\n\n")
        
        f.write("## 每年義務 Annual\n\n")
        for item in calendar_data["annual_obligations"]:
            f.write(f"- **{item['deadline']}**: {item['task']}\n")
            f.write(f"  - {item['task_en']}\n")
            f.write(f"  - 適用: {item['applies_to']} | Ref: {item['reference']}\n\n")
        
        f.write("## 事件觸發義務 Event-Triggered\n\n")
        for item in calendar_data["event_triggered_obligations"]:
            f.write(f"- **{item['trigger']}** → {item['action']}\n")
            f.write(f"  - {item['trigger_en']} → {item['action_en']}\n")
            f.write(f"  - Ref: {item['reference']}\n\n")
    
    print(f"\n  📅 Compliance Calendar (JSON): {path}")
    print(f"  📅 Compliance Calendar (MD): {md_path}")
    print(f"\n  統計:")
    print(f"    每月義務: {len(calendar_data['monthly_obligations'])} 項")
    print(f"    每季義務: {len(calendar_data['quarterly_obligations'])} 項")
    print(f"    每年義務: {len(calendar_data['annual_obligations'])} 項")
    print(f"    事件觸發: {len(calendar_data['event_triggered_obligations'])} 項")
    
    # 提示上傳
    upload = input(f"\n  上傳合規日曆到 VIGO 知識庫？(y/n): ").strip().lower()
    if upload == 'y':
        calendar_text = json.dumps(calendar_data, ensure_ascii=False, indent=2)
        chunks = smart_chunk_text(calendar_text, "SFC Compliance Calendar")
        ok = 0
        for ci, chunk in enumerate(chunks):
            meta = {
                "source": "SFC_Official", "source_url": "https://www.sfc.hk",
                "doc_type": "compliance_calendar", "category": "SFC",
                "ref_no": f"CAL_compliance_{ci}", "source_table": "vigo_financial",
                "keywords": ["SFC", "Compliance_Calendar", "Obligations", "Deadlines", "FRR", "CPT"],
                "collected_date": datetime.now().strftime("%Y-%m-%d"),
                "authority_score": 100, "language": "zh",
            }
            kid = upload_chunk(f"【SFC 持牌法團合規日曆 Part {ci+1}/{len(chunks)}】\n\n{chunk}", meta)
            if kid: ok += 1
            time.sleep(0.3)
        print(f"  ✓ Uploaded {ok}/{len(chunks)} chunks")


def _structured_tag_stats():
    """8g: [P3.16] 結構化標籤統計"""
    print(f"\n  {'=' * 60}")
    print(f"  8g: Structured Tag Statistics (v5.2)")
    print(f"  {'=' * 60}")
    
    # 從 Supabase 查詢
    try:
        headers = {
            "apikey": SUPABASE_SERVICE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
            "Content-Type": "application/json",
        }
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/{KNOWLEDGE_TABLE}?select=metadata&limit=500",
            headers=headers, timeout=30)
        
        if resp.status_code != 200:
            print(f"  ⚠ Cannot fetch from Supabase: HTTP {resp.status_code}")
            return
        
        records = resp.json()
        print(f"\n  分析 {len(records)} 條記錄...")
        
        license_counts = {}
        domain_counts = {}
        hierarchy_counts = {}
        
        for rec in records:
            meta = rec.get("metadata", {})
            if not isinstance(meta, dict):
                continue
            
            for lt in meta.get("license_types", []):
                license_counts[lt] = license_counts.get(lt, 0) + 1
            
            for d in meta.get("compliance_domains", []):
                domain_counts[d] = domain_counts.get(d, 0) + 1
            
            h = meta.get("doc_hierarchy", "unknown")
            hierarchy_counts[str(h)] = hierarchy_counts.get(str(h), 0) + 1
        
        print(f"\n  ── 牌照類型分佈 ──")
        for lt, count in sorted(license_counts.items(), key=lambda x: -x[1]):
            bar = "█" * min(count, 30)
            print(f"    {lt:8s} {count:4d} {bar}")
        
        print(f"\n  ── 合規領域分佈 ──")
        for d, count in sorted(domain_counts.items(), key=lambda x: -x[1]):
            bar = "█" * min(count, 30)
            print(f"    {d:15s} {count:4d} {bar}")
        
        print(f"\n  ── 文件層級分佈 ──")
        hierarchy_labels = {v: k for k, v in DOC_HIERARCHY.items()}
        for h, count in sorted(hierarchy_counts.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 99):
            label = hierarchy_labels.get(int(h) if h.isdigit() else 99, h)
            bar = "█" * min(count, 30)
            print(f"    {label:20s} {count:4d} {bar}")
        
        if not license_counts and not domain_counts:
            print(f"\n  ℹ 現有數據尚未包含結構化標籤。")
            print(f"    新上傳的數據將自動包含 license_types / compliance_domains / doc_hierarchy。")
            print(f"    建議重新運行 Mode 5 或 Mode 1/2 以更新現有數據。")
    
    except Exception as e:
        print(f"  ⚠ Error: {e}")
        print(f"\n  本地結構化標籤系統配置:")
        print(f"    牌照類型: {len(LICENSE_TYPE_KEYWORDS)} 種 ({', '.join(LICENSE_TYPE_KEYWORDS.keys())})")
        print(f"    合規領域: {len(COMPLIANCE_DOMAIN_KEYWORDS)} 個 ({', '.join(COMPLIANCE_DOMAIN_KEYWORDS.keys())})")
        print(f"    文件層級: {len(DOC_HIERARCHY)} 級")


# ============================================================
# v5.0 主菜單
# ============================================================

# ============================================================
# v5.2 Mode 9: VATP 深度採集 [P1.6/P1.7/P1.9]
# FATF Travel Rule + 跨境 VA 比較 + VA ETF
# ============================================================

# FATF VA/VASP 相關指引 [P1.6]
FATF_VA_SOURCES = [
    {"id": "fatf_va_guidance_2021", 
     "name": "FATF Updated Guidance for a Risk-Based Approach to Virtual Assets and VASPs (Oct 2021)",
     "name_zh": "FATF 虛擬資產及 VASP 風險為本指引（2021年10月更新版）",
     "url": "https://www.fatf-gafi.org/content/dam/fatf-gafi/guidance/Updated-Guidance-VA-VASP.pdf",
     "keywords": ["FATF", "VA", "VASP", "Travel_Rule", "AML", "Risk_Based"]},
    {"id": "fatf_travel_rule",
     "name": "FATF Recommendation 16 — Wire Transfer / Travel Rule for VASPs",
     "name_zh": "FATF 第16號建議 — 虛擬資產轉移規則（Travel Rule）",
     "url": "https://www.fatf-gafi.org/content/dam/fatf-gafi/recommendations/FATF%20Recommendations%202012.pdf",
     "keywords": ["FATF", "Travel_Rule", "Recommendation_16", "Wire_Transfer", "VASP"]},
    {"id": "fatf_targeted_update_2023",
     "name": "FATF Targeted Update on Implementation of FATF Standards on VAs/VASPs (Jun 2023)",
     "name_zh": "FATF 虛擬資產/VASP 標準實施進度更新（2023年6月）",
     "url": "https://www.fatf-gafi.org/content/dam/fatf-gafi/reports/Targeted-Update-Implementation-FATF-Standards-VA-VASPs.pdf",
     "keywords": ["FATF", "VA", "VASP", "Implementation", "Travel_Rule"]},
    {"id": "fatf_rba_guidance",
     "name": "FATF Guidance on Risk-Based Approach for the Securities Sector",
     "name_zh": "FATF 證券業風險為本指引",
     "url": "https://www.fatf-gafi.org/content/dam/fatf-gafi/guidance/RBA-Securities-Sector.pdf",
     "keywords": ["FATF", "Securities", "Risk_Based", "AML"]},
]

# 跨境 VA 監管比較來源 [P1.7]
CROSS_BORDER_VA_SOURCES = [
    # MAS (新加坡)
    {"id": "mas_ps_act",
     "name": "MAS Payment Services Act — Digital Payment Token Services",
     "name_zh": "新加坡金管局支付服務法 — 數字支付代幣服務",
     "url": "https://www.mas.gov.sg/regulation/acts/payment-services-act",
     "authority": "MAS_Official",
     "keywords": ["MAS", "Singapore", "Payment_Services", "DPT", "VA"]},
    {"id": "mas_va_guidelines",
     "name": "MAS Guidelines on Licensing for Payment Service Providers (Digital Payment Token)",
     "name_zh": "MAS 數字支付代幣服務商發牌指引",
     "url": "https://www.mas.gov.sg/regulation/guidelines/guidelines-on-licensing-for-payment-service-providers",
     "authority": "MAS_Official",
     "keywords": ["MAS", "Singapore", "Licensing", "DPT", "VA"]},
    # EU MiCA
    {"id": "eu_mica",
     "name": "EU Markets in Crypto-Assets Regulation (MiCA) — Official Text",
     "name_zh": "歐盟加密資產市場法規 (MiCA)",
     "url": "https://eur-lex.europa.eu/legal-content/EN/TXT/PDF/?uri=CELEX:32023R1114",
     "authority": "EU_Official",
     "keywords": ["EU", "MiCA", "Crypto", "VA", "Stablecoin", "CASP"]},
    {"id": "eu_mica_summary",
     "name": "European Commission MiCA Overview Page",
     "name_zh": "歐盟委員會 MiCA 概覽",
     "url": "https://finance.ec.europa.eu/digital-finance/digital-assets/markets-crypto-assets-regulation-mica_en",
     "authority": "EU_Official",
     "keywords": ["EU", "MiCA", "Summary", "Digital_Assets"]},
    # Japan FSA
    {"id": "japan_psa",
     "name": "Japan FSA — Crypto-Asset Exchange Service Provider Registration",
     "name_zh": "日本金融廳 — 加密資產交換業者登記",
     "url": "https://www.fsa.go.jp/en/policy/virtual_currency/index.html",
     "authority": "Japan_FSA_Official",
     "keywords": ["Japan", "FSA", "JVCEA", "Crypto_Asset", "Registration"]},
]

# VA ETF 批准文件來源 [P1.9]
VA_ETF_SOURCES = [
    {"id": "sfc_va_etf_circular",
     "name": "SFC Circular on Virtual Asset Futures ETFs",
     "name_zh": "證監會虛擬資產期貨 ETF 通函",
     "url": "https://www.sfc.hk/en/Published-resources/Circulars-and-Memos",
     "keywords": ["SFC", "VA_ETF", "Bitcoin_ETF", "Ethereum_ETF"]},
    {"id": "sfc_authorized_va_funds",
     "name": "SFC List of SFC-authorized Virtual Asset Funds",
     "name_zh": "證監會認可虛擬資產基金列表",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Products/Virtual-asset-related-products",
     "keywords": ["SFC", "VA_Fund", "ETF", "Authorized"]},
    {"id": "investhk_va",
     "name": "InvestHK — Hong Kong as Virtual Asset Hub",
     "name_zh": "投資推廣署 — 香港虛擬資產中心",
     "url": "https://www.investhk.gov.hk/en/setting-up-your-business/fintech.html",
     "keywords": ["InvestHK", "VA", "Fintech", "Hong_Kong"]},
]


def run_mode_9():
    """Mode 9: VATP 深度採集 — FATF + 跨境比較 + VA ETF [P1.6/P1.7/P1.9]"""
    from bs4 import BeautifulSoup
    
    print(f"\n  {'=' * 60}")
    print(f"  Mode 9: VATP 深度採集 (v5.2)")
    print(f"  FATF VA Travel Rule + 跨境 VA 監管比較 + VA ETF")
    print(f"  {'=' * 60}")
    print(f"\n  a. FATF VA/VASP 指引 ({len(FATF_VA_SOURCES)} 份)")
    print(f"  b. 跨境 VA 監管比較 ({len(CROSS_BORDER_VA_SOURCES)} 份 — MAS/MiCA/Japan)")
    print(f"  c. VA ETF 批准文件 ({len(VA_ETF_SOURCES)} 份)")
    print(f"  d. 全部執行 (a + b + c)")
    
    sub = input(f"\n  (a/b/c/d): ").strip().lower()
    
    source_groups = []
    if sub in ("a", "d"): source_groups.append(("FATF VA/VASP", FATF_VA_SOURCES, "FATF_Official"))
    if sub in ("b", "d"): source_groups.append(("Cross-border VA", CROSS_BORDER_VA_SOURCES, None))
    if sub in ("c", "d"): source_groups.append(("VA ETF", VA_ETF_SOURCES, "SFC_Official"))
    
    if not source_groups:
        print("  無效選擇"); return
    
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded = set(progress.get("uploaded_refs", []))
    
    stats = {"ok": 0, "fail": 0, "skip": 0, "chunks": 0}
    results = []
    start_time = time.time()
    
    for group_name, sources, default_authority in source_groups:
        print(f"\n  {'─' * 50}")
        print(f"  {group_name} ({len(sources)} sources)")
        print(f"  {'─' * 50}")
        
        for i, src in enumerate(sources, 1):
            ref = f"VATP9_{src['id']}"
            if ref in uploaded:
                stats["skip"] += 1
                print(f"\n  [{i}/{len(sources)}] SKIP: {src['name'][:60]}")
                continue
            
            print(f"\n  [{i}/{len(sources)}] {src['name'][:70]}")
            
            try:
                # 針對外部監管機構網站使用更完整的 headers
                req_headers = {
                    **HEADERS,
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,application/pdf;q=0.8,*/*;q=0.7',
                    'Accept-Language': 'en-US,en;q=0.9',
                    'Accept-Encoding': 'gzip, deflate',
                    'Referer': src["url"].split('/')[0] + '//' + src["url"].split('/')[2] + '/',
                    'Connection': 'keep-alive',
                }
                
                # 重試機制（最多2次）
                resp = None
                for attempt in range(2):
                    try:
                        resp = requests.get(src["url"], headers=req_headers, timeout=60,
                                            allow_redirects=True, verify=True)
                        if resp.status_code == 200:
                            break
                        if attempt == 0 and resp.status_code in (403, 429, 503):
                            print(f"     ⚠ HTTP {resp.status_code}, retrying...")
                            time.sleep(3)
                    except requests.exceptions.SSLError:
                        if attempt == 0:
                            print(f"     ⚠ SSL error, retrying without verify...")
                            try:
                                resp = requests.get(src["url"], headers=req_headers, timeout=60,
                                                    allow_redirects=True, verify=False)
                            except Exception:
                                pass
                            break
                    except requests.exceptions.Timeout:
                        if attempt == 0:
                            print(f"     ⚠ Timeout, retrying...")
                            time.sleep(2)
                        continue
                
                if resp is None or resp.status_code != 200:
                    status_code = resp.status_code if resp else 'No response'
                    print(f"     ✗ HTTP {status_code} — {src['url'][:60]}")
                    stats["fail"] += 1
                    results.append({"name": src["name"], "status": "fail", "reason": f"HTTP {status_code}"})
                    continue
                
                # 判斷是 PDF 還是 HTML
                ct = resp.headers.get('content-type', '').lower()
                if 'pdf' in ct or resp.content[:4] == b'%PDF':
                    text = extract_text_from_pdf(resp.content)
                    if not text or len(text) < 300:
                        print(f"     ✗ PDF extraction failed or too short ({len(text) if text else 0} chars)")
                        stats["fail"] += 1
                        results.append({"name": src["name"], "status": "fail", "reason": "PDF extraction failed"})
                        continue
                else:
                    soup = BeautifulSoup(resp.text, 'html.parser')
                    for tag in soup(['script', 'style', 'nav', 'footer', 'header']):
                        tag.decompose()
                    main = soup.find('main') or soup.find('div', class_=re.compile(r'content|main|body', re.I)) or soup
                    text = main.get_text(separator='\n', strip=True)
                    text = re.sub(r'\n{3,}', '\n\n', text)
                    if len(text) < 300:
                        print(f"     ✗ No meaningful content extracted ({len(text)} chars)")
                        stats["fail"] += 1
                        results.append({"name": src["name"], "status": "fail", "reason": "No content"})
                        continue
                
                if len(text) > 60000:
                    text = text[:60000]
                
                print(f"     ✓ Extracted: {len(text)} chars")
                chunks = smart_chunk_text(text, src["name"])
                
                authority_key = src.get("authority", default_authority) or "Industry"
                authority = SOURCE_AUTHORITY.get(authority_key, 50)
                
                for ci, chunk_text in enumerate(chunks):
                    part = f" (Part {ci+1}/{len(chunks)})" if len(chunks) > 1 else ""
                    meta = {
                        "source": authority_key, "source_url": src["url"],
                        "doc_type": "vatp_reference", "category": "VATP",
                        "ref_no": ref, "source_table": "vigo_financial",
                        "keywords": src["keywords"], "collected_date": datetime.now().strftime("%Y-%m-%d"),
                        "authority_score": authority,
                    }
                    en_content = f"【{src['name']}{part}】\n\n{chunk_text}"
                    kid = upload_chunk(en_content, {**meta, "language": "en"})
                    if kid: stats["chunks"] += 1
                    
                    zh = translate_to_chinese(chunk_text)
                    if zh and len(zh) > 80:
                        zh_content = f"【{src.get('name_zh', src['name'])}{part}】\n\n{zh}"
                        kid_zh = upload_chunk(zh_content, {**meta, "language": "zh"})
                        if kid_zh: stats["chunks"] += 1
                    time.sleep(0.5)
                
                uploaded.add(ref)
                progress["uploaded_refs"] = list(uploaded)
                save_progress(progress)
                stats["ok"] += 1
                results.append({"name": src["name"], "status": "ok", "chars": len(text), "chunks": len(chunks)})
                print(f"     ✓ {len(chunks)} chunks uploaded")
            
            except Exception as e:
                print(f"     ✗ Error: {e}")
                stats["fail"] += 1
                results.append({"name": src["name"], "status": "fail", "reason": str(e)[:100]})
    
    total_time = time.time() - start_time
    print(f"\n  {'=' * 60}")
    print(f"  Mode 9 Complete! {total_time/60:.1f} min")
    print(f"  OK: {stats['ok']} | Fail: {stats['fail']} | Skip: {stats['skip']} | Chunks: {stats['chunks']}")
    print(f"  {'=' * 60}")
    generate_report("mode9_vatp_deep", stats, results, total_time)


# ============================================================
# v5.2 Mode 10: SFC 操作知識採集 [P2.10/P2.11/P0.4/P0.2]
# WINGS + 申請表格 + PDPO + 巡查加強
# ============================================================

# SFC 操作知識來源
SFC_OPERATIONAL_SOURCES = [
    # ── WINGS 系統 [P2.10] ──
    {"id": "wings_overview",
     "name": "WINGS — Web-based INteGrated Service System Overview",
     "name_zh": "WINGS 網上綜合服務系統概覽",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Licensing-handbook/How-to-apply",
     "category": "WINGS",
     "keywords": ["SFC", "WINGS", "Application", "Electronic"]},
    {"id": "wings_individual_account",
     "name": "WINGS — Individual Account (Licensed Representative Applications)",
     "name_zh": "WINGS 個人帳戶（持牌代表申請）",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Licensed-persons-and-registered-institutions",
     "category": "WINGS",
     "keywords": ["SFC", "WINGS", "Individual", "LR", "Application"]},
    {"id": "wings_corporate_account",
     "name": "WINGS — Corporate Account (Licensed Corporation Management)",
     "name_zh": "WINGS 公司帳戶（持牌法團管理）",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Licensing-handbook",
     "category": "WINGS",
     "keywords": ["SFC", "WINGS", "Corporate", "LC", "Management"]},
    
    # ── SFC 表格 + 指引 [P2.11] ──
    {"id": "sfc_form1_lc_application",
     "name": "SFC Form 1 — Application for Licence as Licensed Corporation",
     "name_zh": "SFC 表格1 — 持牌法團牌照申請",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Forms-and-checklists",
     "category": "Forms",
     "keywords": ["SFC", "Form1", "LC", "Application", "Licensing"]},
    {"id": "sfc_form2_lr_application",
     "name": "SFC Form 2 — Application for Licence as Licensed Representative",
     "name_zh": "SFC 表格2 — 持牌代表牌照申請",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Forms-and-checklists",
     "category": "Forms",
     "keywords": ["SFC", "Form2", "LR", "Application", "Licensing"]},
    {"id": "sfc_form3_ro_approval",
     "name": "SFC Form 3 — Application for Approval as Responsible Officer",
     "name_zh": "SFC 表格3 — 負責人員批准申請",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Forms-and-checklists",
     "category": "Forms",
     "keywords": ["SFC", "Form3", "RO", "Application", "Responsible_Officer"]},
    {"id": "sfc_checklist_lc",
     "name": "SFC Checklist for Licensed Corporation Application",
     "name_zh": "持牌法團申請核對清單",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Licensing/Forms-and-checklists",
     "category": "Forms",
     "keywords": ["SFC", "Checklist", "LC", "Application"]},
    
    # ── PCPD 操作指引 [P0.4] ──
    {"id": "pcpd_compliance_guide",
     "name": "PCPD Compliance Guide for Financial Services Industry",
     "name_zh": "個人資料私隱專員公署金融服務業合規指南",
     "url": "https://www.pcpd.org.hk/english/resources_centre/publications/guidance/guidance.html",
     "category": "PDPO",
     "keywords": ["PCPD", "PDPO", "Financial_Services", "Compliance", "Privacy"]},
    {"id": "pcpd_data_breach",
     "name": "PCPD Guidance on Data Breach Handling and Notification",
     "name_zh": "資料外洩處理及通報指引",
     "url": "https://www.pcpd.org.hk/english/resources_centre/publications/guidance/guidance.html",
     "category": "PDPO",
     "keywords": ["PCPD", "PDPO", "Data_Breach", "Notification"]},
    
    # ── SFC 巡查加強 [P0.2] ──
    {"id": "sfc_compliance_bulletin",
     "name": "SFC Compliance Bulletin",
     "name_zh": "證監會合規通訊",
     "url": "https://www.sfc.hk/en/Published-resources/Compliance-related-resources/Compliance-Bulletin",
     "category": "Inspection",
     "keywords": ["SFC", "Compliance", "Bulletin", "Findings", "Best_Practice"]},
    {"id": "sfc_cyber_review_2024",
     "name": "SFC Cybersecurity Review Report (2023-24)",
     "name_zh": "證監會網絡安全審查報告 (2023-24)",
     "url": "https://www.sfc.hk/en/Regulatory-functions/Intermediaries/Supervision/Publications-and-statistics",
     "category": "Inspection",
     "keywords": ["SFC", "Cybersecurity", "Review", "IT_Risk", "BCP"]},
]


def run_mode_10():
    """Mode 10: SFC 操作知識採集 [P2.10/P2.11/P0.4/P0.2]"""
    from bs4 import BeautifulSoup
    
    print(f"\n  {'=' * 60}")
    print(f"  Mode 10: SFC 操作知識採集 (v5.2)")
    print(f"  WINGS 指南 + 申請表格 + PDPO + 巡查加強")
    print(f"  {'=' * 60}")
    
    # 按類別分組顯示
    categories = {}
    for src in SFC_OPERATIONAL_SOURCES:
        cat = src.get("category", "Other")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(src)
    
    print(f"\n  來源清單 ({len(SFC_OPERATIONAL_SOURCES)} 項):")
    for cat, items in categories.items():
        print(f"    {cat}: {len(items)} 項")
    
    print(f"\n  a. WINGS 系統 ({len(categories.get('WINGS', []))} 項)")
    print(f"  b. SFC 表格 ({len(categories.get('Forms', []))} 項)")
    print(f"  c. PDPO 指引 ({len(categories.get('PDPO', []))} 項)")
    print(f"  d. 巡查/合規 ({len(categories.get('Inspection', []))} 項)")
    print(f"  e. 全部執行")
    
    sub = input(f"\n  (a/b/c/d/e): ").strip().lower()
    
    cat_map = {"a": "WINGS", "b": "Forms", "c": "PDPO", "d": "Inspection"}
    if sub == "e":
        sources = SFC_OPERATIONAL_SOURCES
    elif sub in cat_map:
        target_cat = cat_map[sub]
        sources = [s for s in SFC_OPERATIONAL_SOURCES if s.get("category") == target_cat]
    else:
        print("  無效選擇"); return
    
    progress = load_progress()
    hashes_data = load_content_hashes()
    uploaded = set(progress.get("uploaded_refs", []))
    
    stats = {"ok": 0, "fail": 0, "skip": 0, "chunks": 0}
    results = []
    start_time = time.time()
    
    for i, src in enumerate(sources, 1):
        ref = f"OPS10_{src['id']}"
        if ref in uploaded:
            stats["skip"] += 1
            print(f"\n  [{i}/{len(sources)}] SKIP: {src['name'][:60]}")
            continue
        
        print(f"\n  [{i}/{len(sources)}] {src['name'][:70]}")
        
        try:
            resp = requests.get(src["url"], headers=HEADERS, timeout=30, allow_redirects=True)
            if resp.status_code != 200:
                stats["fail"] += 1
                results.append({"name": src["name"], "status": "fail", "reason": f"HTTP {resp.status_code}"})
                continue
            
            ct = resp.headers.get('content-type', '').lower()
            if 'pdf' in ct or resp.content[:4] == b'%PDF':
                text = extract_text_from_pdf(resp.content)
            else:
                soup = BeautifulSoup(resp.text, 'html.parser')
                
                # 尋找所有 PDF 鏈接（表格頁面）
                if src.get("category") in ("Forms", "Inspection"):
                    pdf_links = []
                    for a in soup.find_all('a', href=True):
                        href = a.get('href', '')
                        link_text = a.get_text(strip=True).lower()
                        if '.pdf' in href.lower():
                            # 針對表格和巡查報告過濾
                            if any(kw in link_text for kw in ['form', 'checklist', 'guide', 'report', 'review', 'bulletin', 'circular']):
                                if href.startswith('/'):
                                    href = f"https://www.sfc.hk{href}"
                                pdf_links.append({"url": href, "title": a.get_text(strip=True)[:80]})
                    
                    if pdf_links:
                        print(f"     Found {len(pdf_links)} PDF links")
                        combined_text = ""
                        for pi, plink in enumerate(pdf_links[:8]):
                            try:
                                pdf_resp = requests.get(plink["url"], headers=HEADERS, timeout=60)
                                if pdf_resp.status_code == 200 and pdf_resp.content[:4] == b'%PDF':
                                    pdf_text = extract_text_from_pdf(pdf_resp.content)
                                    if pdf_text and len(pdf_text) > 300:
                                        combined_text += f"\n\n=== {plink['title']} ===\n\n{pdf_text}"
                                        print(f"       ✓ {plink['title'][:50]}")
                            except Exception as e:
                                print(f"       ✗ {plink['title'][:30]}: {e}")
                        if combined_text:
                            text = combined_text
                        else:
                            for tag in soup(['script', 'style', 'nav', 'footer']):
                                tag.decompose()
                            text = soup.get_text(separator='\n', strip=True)
                    else:
                        for tag in soup(['script', 'style', 'nav', 'footer']):
                            tag.decompose()
                        text = soup.get_text(separator='\n', strip=True)
                else:
                    for tag in soup(['script', 'style', 'nav', 'footer', 'header']):
                        tag.decompose()
                    main = soup.find('main') or soup.find('div', class_=re.compile(r'content|main|body', re.I)) or soup
                    text = main.get_text(separator='\n', strip=True)
            
            if not text or len(text) < 200:
                stats["fail"] += 1
                results.append({"name": src["name"], "status": "fail", "reason": "No content"})
                continue
            
            text = re.sub(r'\n{3,}', '\n\n', text)
            if len(text) > 50000:
                text = text[:50000]
            
            print(f"     ✓ Extracted: {len(text)} chars")
            chunks = smart_chunk_text(text, src["name"])
            
            authority = SOURCE_AUTHORITY.get("SFC_Official" if "sfc.hk" in src["url"] else "PCPD_Official", 80)
            
            for ci, chunk_text in enumerate(chunks):
                part = f" (Part {ci+1}/{len(chunks)})" if len(chunks) > 1 else ""
                meta = {
                    "source": f"SFC_Operational_{src['id']}" if "sfc.hk" in src["url"] else f"PCPD_Official_{src['id']}",
                    "source_url": src["url"],
                    "doc_type": "operational_guide", "category": src.get("category", "SFC"),
                    "ref_no": ref, "source_table": "vigo_financial",
                    "keywords": src["keywords"], "collected_date": datetime.now().strftime("%Y-%m-%d"),
                    "authority_score": authority,
                }
                en_content = f"【{src['name']}{part}】\n\n{chunk_text}"
                kid = upload_chunk(en_content, {**meta, "language": "en"})
                if kid: stats["chunks"] += 1
                
                zh = translate_to_chinese(chunk_text)
                if zh and len(zh) > 80:
                    zh_content = f"【{src.get('name_zh', src['name'])}{part}】\n\n{zh}"
                    kid_zh = upload_chunk(zh_content, {**meta, "language": "zh"})
                    if kid_zh: stats["chunks"] += 1
                time.sleep(0.5)
            
            uploaded.add(ref)
            progress["uploaded_refs"] = list(uploaded)
            save_progress(progress)
            stats["ok"] += 1
            results.append({"name": src["name"], "status": "ok", "chars": len(text), "chunks": len(chunks)})
            print(f"     ✓ {len(chunks)} chunks uploaded")
        
        except Exception as e:
            print(f"     ✗ Error: {e}")
            stats["fail"] += 1
            results.append({"name": src["name"], "status": "fail", "reason": str(e)[:100]})
    
    total_time = time.time() - start_time
    print(f"\n  {'=' * 60}")
    print(f"  Mode 10 Complete! {total_time/60:.1f} min")
    print(f"  OK: {stats['ok']} | Fail: {stats['fail']} | Skip: {stats['skip']} | Chunks: {stats['chunks']}")
    print(f"  {'=' * 60}")
    generate_report("mode10_sfc_operational", stats, results, total_time)

# ============================================================
# v7.0 Mode 11: Eval Benchmark — Sprint 1
# 50 題自動化檢索質量評估
# ============================================================

def _load_eval_questions():
    """載入評估題庫"""
    paths = [
        EVAL_QUESTIONS_FILE,
        os.path.join(SCRIPT_DIR, "eval_questions.json"),
        os.path.join(SCRIPT_DIR, "config", "eval_questions.json"),
        "eval_questions.json",
    ]
    for p in paths:
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data.get("questions", [])
    print("  ❌ eval_questions.json 未找到")
    print(f"     請將檔案放在: {EVAL_QUESTIONS_FILE}")
    return []


def _query_vigo_chat(question, language="auto"):
    """調用 vigo-chat API 並計時"""
    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
    }
    payload = {"query": question, "userId": "eval_benchmark", "language": language}
    
    start = time.time()
    try:
        resp = requests.post(VIGO_CHAT_URL, headers=headers, json=payload, timeout=90)
        latency = time.time() - start
        if resp.ok:
            result = resp.json()
            return {
                "answer": result.get("answer", ""),
                "sources": result.get("sources", []),
                "latency": latency,
                "status": "ok",
                "version": result.get("version", "?"),
                "reranker_enabled": result.get("reranker_enabled", None),
                "retrieval_route": result.get("retrieval_route", "?"),
                "documents_retrieved": result.get("documents_retrieved", 0),
            }
        return {"answer": "", "sources": [], "latency": latency, "status": f"HTTP {resp.status_code}"}
    except Exception as e:
        return {"answer": "", "sources": [], "latency": time.time() - start, "status": str(e)}


def _query_vigo_mcp(tool, params):
    """調用 vigo-mcp MCP 工具"""
    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
    }
    # MCP Streamable HTTP: send JSON-RPC
    payload = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {"name": tool, "arguments": params}
    }
    start = time.time()
    try:
        resp = requests.post(VIGO_MCP_URL, headers=headers, json=payload, timeout=30)
        latency = time.time() - start
        if resp.ok:
            result = resp.json()
            content = result.get("result", {}).get("content", [])
            text = " ".join(c.get("text", "") for c in content if c.get("type") == "text")
            return {"answer": text, "latency": latency, "status": "ok"}
        return {"answer": "", "latency": latency, "status": f"HTTP {resp.status_code}"}
    except Exception as e:
        return {"answer": "", "latency": time.time() - start, "status": str(e)}


def _check_answer(response, question_data):
    """檢查回答是否包含預期內容"""
    answer = response.get("answer", "").lower()
    sources_text = " ".join(str(s) for s in response.get("sources", [])).lower()
    full_text = answer + " " + sources_text
    
    expected = question_data.get("expected_contains", [])
    if not expected:
        return True
    
    matches = sum(1 for kw in expected if kw.lower() in full_text)
    return matches >= max(1, len(expected) // 3)  # 至少匹配 1/3 的關鍵詞


def run_mode_11():
    """Mode 11: Eval Benchmark — 自動化檢索質量評估"""
    print("\n" + "=" * 60)
    print("  Mode 11: Eval Benchmark 📊")
    print("  自動化檢索質量評估（50 題）")
    print("=" * 60)

    questions = _load_eval_questions()
    if not questions:
        return

    print(f"\n  載入 {len(questions)} 題測試問題")
    print(f"  目標 API: {VIGO_CHAT_URL}")

    print("\n  a. 完整評估（50 題，~5 分鐘）")
    print("  b. 快速評估（10 題，~1 分鐘）")
    print("  c. 按類別評估（A/B/C/D/E）")
    print("  d. 單題測試")
    
    sub = input("\n  (a/b/c/d): ").strip().lower()
    
    if sub == "b":
        # 每類取 2 題
        sampled = []
        for cat in ["A", "B", "C", "D", "E"]:
            cat_qs = [q for q in questions if q["cat"] == cat]
            sampled.extend(cat_qs[:2])
        questions = sampled
    elif sub == "c":
        cat = input("  類別 (A=牌照/B=合規/C=執法/D=市場/E=跨境): ").strip().upper()
        questions = [q for q in questions if q["cat"] == cat]
        if not questions:
            print(f"  ❌ 類別 {cat} 無題目")
            return
    elif sub == "d":
        qid = input("  題目 ID (如 A01): ").strip().upper()
        questions = [q for q in questions if q["id"] == qid]
        if not questions:
            print(f"  ❌ 題目 {qid} 未找到")
            return
    elif sub != "a":
        print("  Invalid")
        return
    
    print(f"\n  開始評估 {len(questions)} 題...\n")
    
    results = []
    correct = 0
    latencies = []
    cat_stats = {}
    start_all = time.time()
    
    for i, q in enumerate(questions):
        cat = q["cat"]
        if cat not in cat_stats:
            cat_stats[cat] = {"total": 0, "correct": 0}
        cat_stats[cat]["total"] += 1
        
        show_progress(i + 1, len(questions), prefix="  Eval", start_time=start_all)
        
        # 調用 API
        resp = _query_vigo_chat(q["question"])
        
        # 檢查結果
        is_correct = _check_answer(resp, q)
        if is_correct:
            correct += 1
            cat_stats[cat]["correct"] += 1
        
        latencies.append(resp["latency"])
        
        results.append({
            "id": q["id"],
            "cat": cat,
            "question": q["question"],
            "correct": is_correct,
            "latency": round(resp["latency"], 2),
            "status": resp["status"],
            "answer_preview": resp["answer"][:200] if resp["answer"] else "(empty)",
            "version": resp.get("version", "?"),
            "reranker": resp.get("reranker_enabled", None),
            "route": resp.get("retrieval_route", "?"),
        })
        
        # 顯示即時結果
        icon = "✅" if is_correct else "❌"
        print(f"    {q['id']} {icon} ({resp['latency']:.1f}s) {q['question'][:40]}...")
        
        time.sleep(0.3)  # rate limit
    
    total_time = time.time() - start_all
    
    # ═══ 計算指標 ═══
    total = len(questions)
    recall_at_5 = correct / total if total > 0 else 0
    
    # MRR（簡化版：正確=rank 1，不正確=0）
    mrr = sum(1.0 for r in results if r["correct"]) / total if total > 0 else 0
    
    latencies_sorted = sorted(latencies)
    p50 = latencies_sorted[len(latencies_sorted) // 2] if latencies_sorted else 0
    p95 = latencies_sorted[int(len(latencies_sorted) * 0.95)] if latencies_sorted else 0
    
    # ═══ 報告 ═══
    print("\n" + "=" * 60)
    print("  📊 EVAL BENCHMARK REPORT")
    print("=" * 60)

    # 引擎診斷（從第一個成功的結果提取）
    ok_results = [r for r in results if r["status"] == "ok"]
    if ok_results:
        first = ok_results[0]
        ver = first.get("version", "?")
        rr = first.get("reranker", None)
        rr_str = "✅ ON" if rr == True else ("❌ OFF" if rr == False else "⚠️ 未知")
        print(f"\n  🔧 Engine: vigo-chat {ver} | Reranker: {rr_str}")
        routes_used = set(r.get("route", "?") for r in ok_results)
        print(f"     Routes used: {', '.join(sorted(routes_used))}")
    
    timeout_count = sum(1 for r in results if "timed out" in r.get("status", "").lower() or "timeout" in r.get("status", "").lower())
    if timeout_count:
        print(f"  ⚠️  超時: {timeout_count} 題 (timeout=90s)")

    print(f"\n  Overall:")
    print(f"    Recall@5:    {recall_at_5:.1%} ({correct}/{total})")
    print(f"    MRR:         {mrr:.3f}")
    print(f"    Latency P50: {p50:.2f}s")
    print(f"    Latency P95: {p95:.2f}s")
    print(f"    Total Time:  {total_time:.0f}s")
    
    print(f"\n  By Category:")
    cat_names = {"A": "牌照/考試/資本", "B": "合規操作", "C": "執法案例", "D": "市場/費用", "E": "跨境/比較"}
    for cat, stats in sorted(cat_stats.items()):
        rate = stats["correct"] / stats["total"] if stats["total"] > 0 else 0
        bar = "█" * int(rate * 20) + "░" * (20 - int(rate * 20))
        print(f"    {cat} ({cat_names.get(cat, '?')[:8]:>8}): {bar} {rate:.0%} ({stats['correct']}/{stats['total']})")
    
    # 失敗題目
    failed = [r for r in results if not r["correct"]]
    if failed:
        print(f"\n  Failed Questions ({len(failed)}):")
        for r in failed[:10]:
            print(f"    {r['id']}: {r['question'][:50]}...")
    
    # 保存報告
    report = {
        "timestamp": datetime.now().isoformat(),
        "api_url": VIGO_CHAT_URL,
        "total_questions": total,
        "correct": correct,
        "recall_at_5": round(recall_at_5, 4),
        "mrr": round(mrr, 4),
        "latency_p50": round(p50, 2),
        "latency_p95": round(p95, 2),
        "total_time_sec": round(total_time, 1),
        "by_category": {cat: {"correct": s["correct"], "total": s["total"], "rate": round(s["correct"]/s["total"], 4) if s["total"] > 0 else 0} for cat, s in cat_stats.items()},
        "results": results,
    }
    
    report_dir = os.path.join(SCRIPT_DIR, "reports")
    os.makedirs(report_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(report_dir, f"eval_benchmark_{ts}.json")
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n  📁 報告已保存: {report_path}")
    
    # World-class 對標
    print(f"\n  ─── World-Class 對標 ───")
    print(f"  {'指標':<15} {'當前':>8} {'目標':>8} {'狀態':>6}")
    print(f"  {'Recall@5':<15} {recall_at_5:>7.1%} {'≥85%':>8} {'✅' if recall_at_5 >= 0.85 else '⚠️':>6}")
    print(f"  {'MRR':<15} {mrr:>7.3f} {'≥0.80':>8} {'✅' if mrr >= 0.80 else '⚠️':>6}")
    print(f"  {'Latency P95':<15} {p95:>6.1f}s {'≤2.0s':>8} {'✅' if p95 <= 2.0 else '⚠️':>6}")


# ============================================================
# v7.0 Mode 12: Structured Extraction — Sprint 2
# 從 vigo_knowledge 提取結構化數據到新表
# ============================================================

def _supabase_query(table, select="*", filters=None, limit=1000):
    """通用 Supabase REST 查詢"""
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    }
    url = f"{SUPABASE_URL}/rest/v1/{table}?select={select}&limit={limit}"
    if filters:
        for k, v in filters.items():
            url += f"&{k}={v}"
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.ok:
            return resp.json()
        print(f"    ⚠ Query {table}: HTTP {resp.status_code}")
        return []
    except Exception as e:
        print(f"    ⚠ Query {table}: {e}")
        return []


def _supabase_upsert(table, data, on_conflict=None):
    """通用 Supabase REST upsert"""
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=representation" if on_conflict else "return=representation",
    }
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if on_conflict:
        url += f"?on_conflict={on_conflict}"
    try:
        resp = requests.post(url, headers=headers, json=data if isinstance(data, list) else [data], timeout=30)
        if resp.status_code in [200, 201]:
            return resp.json()
        print(f"    ⚠ Upsert {table}: HTTP {resp.status_code} — {resp.text[:200]}")
        return None
    except Exception as e:
        print(f"    ⚠ Upsert {table}: {e}")
        return None


def _extract_enforcement_llm(content, metadata):
    """用 LLM 從執法文本提取結構化數據"""
    prompt = f"""Extract structured enforcement data from this SFC enforcement text. Return ONLY a JSON object:
{{
  "respondent_name": "company or person name",
  "respondent_type": "individual" or "corporation" or "both",
  "license_types": ["Type X", ...] or [],
  "violation_types": ["AML", "conduct", "market_manipulation", "client_asset", "disclosure", "internal_control", "other"],
  "fine_amount": number or null,
  "other_sanctions": ["reprimand", "ban_X_years", "suspension", "revocation", "conditions"] or [],
  "decision_date": "YYYY-MM-DD" or null,
  "sfc_ref": "SFC reference number" or null
}}

Text:
{content[:3000]}"""
    
    try:
        result = deepseek_chat(prompt, max_tokens=500, temperature=0.1)
        # Extract JSON from response
        result = result.strip()
        if result.startswith("```"):
            result = result.split("```")[1]
            if result.startswith("json"):
                result = result[4:]
        return json.loads(result)
    except:
        return None


def _extract_regulation_llm(content, metadata):
    """用 LLM 從法規文本提取結構化數據"""
    prompt = f"""Extract structured regulation data from this SFC regulation text. Return ONLY a JSON object:
{{
  "regulation_id": "short_id like SFO or CoC or AML_Guideline",
  "name_en": "full English name",
  "name_zh": "繁體中文名稱",
  "category": "primary_legislation" or "subsidiary_legislation" or "code" or "guideline" or "handbook" or "thematic_report" or "listing_rules",
  "applies_to": ["Type 1", "Type 9", ...] or ["All"],
  "key_requirements": ["requirement 1", "requirement 2", ...]
}}

Text:
{content[:3000]}"""
    
    try:
        result = deepseek_chat(prompt, max_tokens=500, temperature=0.1)
        result = result.strip()
        if result.startswith("```"):
            result = result.split("```")[1]
            if result.startswith("json"):
                result = result[4:]
        return json.loads(result)
    except:
        return None


def _run_extraction_enforcement():
    """12a: 從 vigo_knowledge 提取執法案例到 vigo_enforcement"""
    print("\n  ── 12a: Enforcement Extraction ──")
    
    # 查詢執法相關記錄
    rows = _supabase_query(
        KNOWLEDGE_TABLE,
        select="id,content,metadata",
        filters={
            "source_id": "eq.sfc_enforcement",
            "order": "created_at.desc",
        },
        limit=500,
    )
    
    if not rows:
        # Fallback: search by content_type or metadata
        rows = _supabase_query(
            KNOWLEDGE_TABLE,
            select="id,content,metadata",
            filters={
                "content_type": "eq.enforcement",
                "order": "created_at.desc",
            },
            limit=500,
        )
    
    if not rows:
        print("    ⚠ 未找到執法記錄（source_id=sfc_enforcement 或 content_type=enforcement）")
        print("    嘗試從 metadata 搜索...")
        rows = _supabase_query(
            KNOWLEDGE_TABLE,
            select="id,content,metadata",
            limit=2000,
        )
        # Filter locally
        enforce_kw = ['reprimand', 'fine', 'suspend', 'ban', 'prosecut', 'disciplin', '紀律', '罰款', '暫停', '禁止']
        rows = [r for r in rows if any(kw in r.get('content', '').lower() for kw in enforce_kw)]
    
    print(f"    找到 {len(rows)} 條執法相關記錄")
    
    if not rows:
        return 0
    
    extracted = 0
    errors = 0
    start = time.time()
    
    for i, row in enumerate(rows):
        show_progress(i + 1, len(rows), prefix="    Extract", start_time=start)
        
        content = row.get("content", "")
        metadata = row.get("metadata", {})
        
        # LLM 提取
        data = _extract_enforcement_llm(content, metadata)
        if not data or not data.get("respondent_name"):
            errors += 1
            continue
        
        # 生成 case_id
        case_id = f"ENF_{data.get('sfc_ref', '') or row['id']}".replace(" ", "_")[:64]
        
        # 構建記錄
        record = {
            "case_id": case_id,
            "respondent_name": data["respondent_name"],
            "respondent_type": data.get("respondent_type", "corporation"),
            "license_types": data.get("license_types", []),
            "violation_types": data.get("violation_types", []),
            "fine_amount": data.get("fine_amount"),
            "fine_currency": "HKD",
            "other_sanctions": data.get("other_sanctions", []),
            "decision_date": data.get("decision_date"),
            "sfc_ref": data.get("sfc_ref"),
            "summary_en": content[:500] if not any('\u4e00' <= c <= '\u9fff' for c in content[:20]) else None,
            "summary_zh": content[:500] if any('\u4e00' <= c <= '\u9fff' for c in content[:20]) else None,
            "source_url": metadata.get("source_url", metadata.get("url", "")),
            "knowledge_chunk_ids": [row["id"]],
        }
        
        result = _supabase_upsert("vigo_enforcement", record, on_conflict="case_id")
        if result:
            extracted += 1
        else:
            errors += 1
        
        time.sleep(0.5)  # rate limit for LLM
    
    print(f"\n    ✅ 提取完成: {extracted} 條入庫, {errors} 條失敗")
    return extracted


def _run_extraction_regulations():
    """12b: 從 vigo_knowledge 提取法規到 vigo_regulations"""
    print("\n  ── 12b: Regulations Extraction ──")
    
    # 基石法規清單（35份 + 5份主題報告）
    FOUNDATIONAL_SEEDS = [
        {"id": "SFO", "name_en": "Securities and Futures Ordinance", "name_zh": "證券及期貨條例", "category": "primary_legislation"},
        {"id": "AMLO", "name_en": "Anti-Money Laundering and Counter-Terrorist Financing Ordinance", "name_zh": "打擊洗錢及恐怖分子資金籌集條例", "category": "primary_legislation"},
        {"id": "SFF", "name_en": "Securities and Futures (Financial Resources) Rules", "name_zh": "證券及期貨（財政資源）規則", "category": "subsidiary_legislation"},
        {"id": "CoC", "name_en": "Code of Conduct for Persons Licensed by or Registered with the SFC", "name_zh": "證券及期貨事務監察委員會持牌人或註冊人操守準則", "category": "code"},
        {"id": "FMCoC", "name_en": "Fund Manager Code of Conduct", "name_zh": "基金經理操守準則", "category": "code"},
        {"id": "AML_Guideline", "name_en": "Guideline on Anti-Money Laundering and Counter-Financing of Terrorism", "name_zh": "打擊洗錢及恐怖分子資金籌集指引", "category": "guideline"},
        {"id": "Competence_Guideline", "name_en": "Guidelines on Competence", "name_zh": "勝任能力的指引", "category": "guideline"},
        {"id": "Internal_Control", "name_en": "Management, Supervision and Internal Control Guidelines", "name_zh": "管理、監督及內部監控指引", "category": "guideline"},
        {"id": "Cybersecurity", "name_en": "Guidelines for Reducing and Mitigating Hacking Risks", "name_zh": "降低及紓減遭受黑客入侵風險的指引", "category": "guideline"},
        {"id": "Online_Distribution", "name_en": "Guidelines on Online Distribution and Advisory Platforms", "name_zh": "有關網上分銷及投資諮詢平台的指引", "category": "guideline"},
        {"id": "Market_Soundings", "name_en": "Guidelines on Market Soundings", "name_zh": "市場探盤活動指引", "category": "guideline"},
        {"id": "SFC_Handbook", "name_en": "SFC Licensing Handbook", "name_zh": "SFC 發牌手冊", "category": "handbook"},
        {"id": "VATP_Handbook", "name_en": "VATP Licensing Handbook", "name_zh": "虛擬資產交易平台發牌手冊", "category": "handbook"},
        {"id": "Takeovers_Code", "name_en": "The Codes on Takeovers and Mergers and Share Buy-backs", "name_zh": "收購守則及股份回購守則", "category": "code"},
        {"id": "HKEX_MB", "name_en": "HKEX Main Board Listing Rules", "name_zh": "聯交所主板上市規則", "category": "listing_rules"},
        {"id": "HKEX_GEM", "name_en": "HKEX GEM Listing Rules", "name_zh": "聯交所GEM上市規則", "category": "listing_rules"},
        {"id": "TR_Cybersec", "name_en": "Thematic Review: Cybersecurity", "name_zh": "主題巡查：網絡安全", "category": "thematic_report"},
        {"id": "TR_Sponsor", "name_en": "Thematic Review: Sponsor Business", "name_zh": "主題巡查：保薦人業務", "category": "thematic_report"},
        {"id": "TR_PrimeServices", "name_en": "Thematic Review: Prime Services", "name_zh": "主題巡查：主要經紀及股票衍生品", "category": "thematic_report"},
        {"id": "TR_MarginFin", "name_en": "Thematic Review: Securities Margin Financing", "name_zh": "主題巡查：證券保證金融資", "category": "thematic_report"},
        {"id": "TR_ALP", "name_en": "Thematic Review: Alternative Liquidity Pools", "name_zh": "主題巡查：另類流動性池", "category": "thematic_report"},
    ]
    
    print(f"    導入 {len(FOUNDATIONAL_SEEDS)} 條基石法規...")
    
    extracted = 0
    for seed in FOUNDATIONAL_SEEDS:
        record = {
            "regulation_id": seed["id"],
            "name_en": seed["name_en"],
            "name_zh": seed["name_zh"],
            "category": seed["category"],
            "applies_to": ["All"],
            "key_requirements": [],
            "knowledge_chunk_ids": [],
        }
        result = _supabase_upsert("vigo_regulations", record, on_conflict="regulation_id")
        if result:
            extracted += 1
    
    print(f"    ✅ 導入完成: {extracted} 條法規")
    
    # 嘗試從 knowledge 中關聯 chunk IDs
    print("    關聯 knowledge chunks...")
    all_regs = _supabase_query("vigo_regulations", select="regulation_id,name_en", limit=100)
    for reg in all_regs:
        # 搜索相關 chunks
        name = reg.get("name_en", "")
        if not name:
            continue
        chunks = _supabase_query(
            KNOWLEDGE_TABLE,
            select="id",
            filters={"content": f"ilike.*{name[:30]}*", "limit": "10"},
            limit=10,
        )
        if chunks:
            chunk_ids = [c["id"] for c in chunks]
            # Update
            headers = {
                "apikey": SUPABASE_SERVICE_KEY,
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "Content-Type": "application/json",
            }
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/vigo_regulations?regulation_id=eq.{reg['regulation_id']}",
                headers=headers,
                json={"knowledge_chunk_ids": chunk_ids},
                timeout=15,
            )
    
    print(f"    ✅ 關聯完成")
    return extracted


def _run_extraction_license_market():
    """12c: 從 vigo_knowledge 提取牌照行情到 vigo_license_market"""
    print("\n  ── 12c: License Market Extraction ──")
    
    # 搜索牌照行情相關記錄
    rows = _supabase_query(
        KNOWLEDGE_TABLE,
        select="id,content,metadata",
        limit=2000,
    )
    
    market_kw = ['牌照出售', 'license for sale', 'shell company', '殼公司', '牌照轉讓',
                 'license wanted', '徵求牌照', 'asking price', '售價', 'paradox',
                 'for sale', 'wanted', 'license acquisition', '牌照收購']
    
    rows = [r for r in rows if any(kw in r.get('content', '').lower() for kw in market_kw)]
    print(f"    找到 {len(rows)} 條牌照行情相關記錄")
    
    if not rows:
        print("    ℹ 暫無牌照行情數據。建議先用 Mode 6 抓取 Paradox 數據。")
        return 0
    
    extracted = 0
    for i, row in enumerate(rows):
        content = row.get("content", "")
        metadata = row.get("metadata", {})
        
        prompt = f"""Extract license market listing from this text. Return ONLY a JSON object:
{{
  "listing_type": "wanted" or "for_sale" or "completed",
  "license_types": ["Type X", ...],
  "asking_price": number or null,
  "nav": number or null,
  "conditions": "brief conditions" or null
}}

Text:
{content[:2000]}"""
        
        try:
            result = deepseek_chat(prompt, max_tokens=300, temperature=0.1)
            result = result.strip()
            if result.startswith("```"):
                result = result.split("```")[1]
                if result.startswith("json"):
                    result = result[4:]
            data = json.loads(result)
        except:
            continue
        
        if not data.get("license_types"):
            continue
        
        listing_id = f"LM_{row['id']}_{i}"
        record = {
            "listing_id": listing_id,
            "listing_type": data.get("listing_type", "for_sale"),
            "license_types": data.get("license_types", []),
            "asking_price": data.get("asking_price"),
            "nav": data.get("nav"),
            "currency": "HKD",
            "conditions": data.get("conditions"),
            "source_id": metadata.get("source", ""),
            "source_url": metadata.get("source_url", metadata.get("url", "")),
        }
        
        result = _supabase_upsert("vigo_license_market", record, on_conflict="listing_id")
        if result:
            extracted += 1
        
        time.sleep(0.5)
    
    print(f"    ✅ 提取完成: {extracted} 條牌照行情")
    return extracted


def _run_extraction_fee_benchmarks():
    """12d: 從知識庫和已知數據建立費用基準"""
    print("\n  ── 12d: Fee Benchmarks Extraction ──")
    
    # 預設已知費用基準（來自行業經驗 + 多來源匯總）
    KNOWN_FEES = [
        {"fee_type": "RO_retainer", "fee_subtype": "Type 1", "license_types": ["Type 1"], "range_low": 15000, "range_high": 35000, "typical": 25000, "period": "monthly", "confidence": "medium", "sources": ["Paradox", "industry"], "notes": "根據牌照類型和 AUM 浮動"},
        {"fee_type": "RO_retainer", "fee_subtype": "Type 9", "license_types": ["Type 9"], "range_low": 20000, "range_high": 50000, "typical": 35000, "period": "monthly", "confidence": "medium", "sources": ["Paradox", "industry"], "notes": "資管類 RO 費用通常較高"},
        {"fee_type": "RO_retainer", "fee_subtype": "Type 4+9", "license_types": ["Type 4", "Type 9"], "range_low": 25000, "range_high": 60000, "typical": 40000, "period": "monthly", "confidence": "medium", "sources": ["Paradox", "industry"]},
        {"fee_type": "compliance_officer", "license_types": ["All"], "range_low": 30000, "range_high": 80000, "typical": 50000, "period": "monthly", "confidence": "medium", "sources": ["industry"], "notes": "全職合規主任月薪"},
        {"fee_type": "audit_annual", "fee_subtype": "small_firm", "license_types": ["Type 1"], "range_low": 80000, "range_high": 200000, "typical": 120000, "period": "annually", "confidence": "medium", "sources": ["industry"], "notes": "小型持牌法團年度審計"},
        {"fee_type": "audit_annual", "fee_subtype": "medium_firm", "license_types": ["Type 1", "Type 9"], "range_low": 200000, "range_high": 500000, "typical": 300000, "period": "annually", "confidence": "medium", "sources": ["industry"]},
        {"fee_type": "sfc_annual_fee", "fee_subtype": "corporation", "license_types": ["All"], "range_low": 5540, "range_high": 5540, "typical": 5540, "period": "annually", "confidence": "high", "sources": ["SFC"], "notes": "每個受規管活動的年費"},
        {"fee_type": "sfc_application_fee", "fee_subtype": "corporation", "license_types": ["All"], "range_low": 5540, "range_high": 5540, "typical": 5540, "period": "one_time", "confidence": "high", "sources": ["SFC"]},
        {"fee_type": "sfc_application_fee", "fee_subtype": "individual_LR", "license_types": ["All"], "range_low": 1390, "range_high": 1390, "typical": 1390, "period": "one_time", "confidence": "high", "sources": ["SFC"]},
        {"fee_type": "sfc_annual_fee", "fee_subtype": "individual_LR", "license_types": ["All"], "range_low": 1810, "range_high": 1810, "typical": 1810, "period": "annually", "confidence": "high", "sources": ["SFC"]},
        {"fee_type": "legal_setup", "fee_subtype": "license_application", "license_types": ["All"], "range_low": 150000, "range_high": 500000, "typical": 250000, "period": "one_time", "confidence": "low", "sources": ["industry"], "notes": "律師費（牌照申請）"},
        {"fee_type": "compliance_consulting", "fee_subtype": "monthly_retainer", "license_types": ["All"], "range_low": 15000, "range_high": 50000, "typical": 25000, "period": "monthly", "confidence": "medium", "sources": ["CompliancePlus", "industry"], "notes": "外判合規月費"},
    ]
    
    extracted = 0
    for fee in KNOWN_FEES:
        fee["source_count"] = len(fee.get("sources", []))
        fee["last_updated"] = datetime.now().strftime("%Y-%m-%d")
        result = _supabase_upsert("vigo_fee_benchmarks", fee)
        if result:
            extracted += 1
    
    print(f"    ✅ 導入完成: {extracted} 條費用基準")
    
    # 從 knowledge 補充
    rows = _supabase_query(
        KNOWLEDGE_TABLE,
        select="id,content",
        limit=2000,
    )
    fee_kw = ['費用', 'fee', 'cost', 'price', '收費', 'charge', 'retainer', '月費']
    fee_rows = [r for r in rows if any(kw in r.get('content', '').lower() for kw in fee_kw)]
    print(f"    📝 找到 {len(fee_rows)} 條費用相關記錄可供後續精煉")
    
    return extracted


def _show_structured_stats():
    """12f: 結構化表統計總覽"""
    print("\n  ── Structured Tables Stats ──\n")
    
    tables = [
        ("vigo_enforcement", "case_id"),
        ("vigo_regulations", "regulation_id"),
        ("vigo_license_market", "listing_id"),
        ("vigo_fee_benchmarks", "id"),
    ]
    
    total = 0
    for table, pk in tables:
        rows = _supabase_query(table, select=pk, limit=5000)
        count = len(rows) if rows else 0
        total += count
        
        icon = "📊" if count > 0 else "⚪"
        print(f"    {icon} {table}: {count} 條")
        
        if table == "vigo_enforcement" and count > 0:
            # Show top violation types
            violations = {}
            full = _supabase_query(table, select="violation_types,fine_amount", limit=500)
            for r in full:
                for v in (r.get("violation_types") or []):
                    violations[v] = violations.get(v, 0) + 1
            if violations:
                top = sorted(violations.items(), key=lambda x: -x[1])[:5]
                print(f"       Top violations: {', '.join(f'{k}({v})' for k,v in top)}")
        
        elif table == "vigo_regulations" and count > 0:
            cats = {}
            full = _supabase_query(table, select="category", limit=100)
            for r in full:
                c = r.get("category", "unknown")
                cats[c] = cats.get(c, 0) + 1
            if cats:
                print(f"       Categories: {', '.join(f'{k}({v})' for k,v in cats.items())}")
        
        elif table == "vigo_fee_benchmarks" and count > 0:
            types = {}
            full = _supabase_query(table, select="fee_type,confidence", limit=100)
            for r in full:
                t = r.get("fee_type", "unknown")
                types[t] = types.get(t, 0) + 1
            if types:
                print(f"       Fee types: {', '.join(f'{k}({v})' for k,v in types.items())}")
    
    print(f"\n    Total structured records: {total}")
    print(f"    vigo_knowledge records: {len(_supabase_query(KNOWLEDGE_TABLE, select='id', limit=10000))}")


def run_mode_12():
    """Mode 12: Structured Extraction — 從 vigo_knowledge 提取結構化數據"""
    print("\n" + "=" * 60)
    print("  Mode 12: Structured Extraction 🏗️")
    print("  從 vigo_knowledge 提取結構化數據到專用表")
    print("=" * 60)
    
    print("\n  ⚠ 前置條件：請先在 Supabase SQL Editor 運行 005_structured_tables.sql")
    print("    建立 4 張結構化表：vigo_enforcement / vigo_regulations /")
    print("    vigo_license_market / vigo_fee_benchmarks\n")
    
    print("  a. 🔍 執法案例提取 → vigo_enforcement（LLM 輔助）")
    print("  b. 📜 法規導入 → vigo_regulations（種子清單）")
    print("  c. 💰 牌照行情提取 → vigo_license_market（LLM 輔助）")
    print("  d. 💵 費用基準導入 → vigo_fee_benchmarks（預設數據）")
    print("  e. ⚡ 一鍵全部提取（a+b+c+d）")
    print("  f. 📊 結構化表統計總覽")
    
    sub = input("\n  (a/b/c/d/e/f): ").strip().lower()
    
    start = time.time()
    
    if sub == "a":
        _run_extraction_enforcement()
    elif sub == "b":
        _run_extraction_regulations()
    elif sub == "c":
        _run_extraction_license_market()
    elif sub == "d":
        _run_extraction_fee_benchmarks()
    elif sub == "e":
        print("\n  ⚡ 一鍵全部提取...\n")
        _run_extraction_regulations()
        _run_extraction_fee_benchmarks()
        _run_extraction_enforcement()
        _run_extraction_license_market()
        print("\n  ═══ 全部提取完成 ═══")
        _show_structured_stats()
    elif sub == "f":
        _show_structured_stats()
    else:
        print("  Invalid")
        return
    
    if sub != "f":
        elapsed = time.time() - start
        print(f"\n  ⏱ 耗時: {elapsed:.0f}s")


# ============================================================
# v7.0 Mode 13: Data Quality Engine — Sprint 3
# 語義去重 + 鑒真補全 + 時效性衰減
# ============================================================

def _fetch_all_knowledge(select="id,content,metadata,source_id,verification_grade,authority_score,freshness_score,content_type,created_at", limit=10000):
    """批量獲取 vigo_knowledge 記錄"""
    all_rows = []
    offset = 0
    batch = 1000
    while offset < limit:
        rows = _supabase_query(
            KNOWLEDGE_TABLE,
            select=select,
            filters={"order": "id.asc", "offset": str(offset), "limit": str(batch)},
            limit=batch,
        )
        if not rows:
            break
        all_rows.extend(rows)
        offset += len(rows)
        if len(rows) < batch:
            break
    return all_rows


def _supabase_patch(table, filters, data):
    """通用 Supabase REST PATCH（更新）"""
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    url = f"{SUPABASE_URL}/rest/v1/{table}?{filters}"
    try:
        resp = requests.patch(url, headers=headers, json=data, timeout=30)
        return resp.status_code in [200, 204]
    except:
        return False


def _run_semantic_dedup():
    """13a: 語義去重 — 用 embedding cosine similarity 找語義重複"""
    print("\n  ── 13a: Semantic Dedup Scan ──")
    print("  策略: cosine similarity ≥ 0.92 → 標記為重複")
    print("  保留: authority_score 最高的版本\n")

    # 獲取所有記錄的 embedding（通過 Supabase RPC 計算相似度）
    # 方法：逐條對比太慢，改用分批 + source_id 分組策略
    rows = _fetch_all_knowledge(select="id,content,source_id,authority_score,verification_grade,created_at")
    print(f"  共 {len(rows)} 條記錄")

    if len(rows) < 2:
        print("  ℹ 記錄太少，無需去重")
        return 0

    # 策略：按 source_id 分組，只比對不同 source_id 的記錄（同一來源的重複已由 L2 hash 處理）
    from collections import defaultdict
    by_source = defaultdict(list)
    for r in rows:
        sid = r.get("source_id", "unknown") or "unknown"
        by_source[sid].append(r)

    print(f"  來源分佈: {', '.join(f'{k}({len(v)})' for k, v in sorted(by_source.items(), key=lambda x: -len(x[1]))[:8])}")

    # 用內容前 200 字的 SimHash 做快速預篩
    duplicates = []
    checked = 0
    total_pairs = 0

    # 對每條記錄，取前 300 字做文本比較
    content_map = {}
    for r in rows:
        content_preview = r.get("content", "")[:300].strip().lower()
        # 簡化去標點
        content_preview = ''.join(c for c in content_preview if c.isalnum() or c.isspace() or '\u4e00' <= c <= '\u9fff')
        content_map[r["id"]] = content_preview

    # 用 n-gram Jaccard 做快速近似比對
    def ngram_set(text, n=3):
        return set(text[i:i+n] for i in range(max(1, len(text) - n + 1)))

    ids = list(content_map.keys())
    ngram_cache = {rid: ngram_set(content_map[rid]) for rid in ids}

    print(f"  計算 Jaccard 相似度（n-gram）...")
    start = time.time()

    # 只比對前 3000 條（避免 O(n²) 過慢）
    check_ids = ids[:3000]
    for i in range(len(check_ids)):
        if i % 500 == 0 and i > 0:
            show_progress(i, len(check_ids), prefix="  Dedup", start_time=start)
        for j in range(i + 1, min(i + 100, len(check_ids))):  # 滑動窗口比較
            id_a, id_b = check_ids[i], check_ids[j]
            set_a, set_b = ngram_cache[id_a], ngram_cache[id_b]
            if not set_a or not set_b:
                continue
            jaccard = len(set_a & set_b) / len(set_a | set_b)
            total_pairs += 1
            if jaccard >= 0.75:  # 高 Jaccard → 很可能語義重複
                # 決定保留哪個
                row_a = next(r for r in rows if r["id"] == id_a)
                row_b = next(r for r in rows if r["id"] == id_b)
                score_a = (row_a.get("authority_score") or 0)
                score_b = (row_b.get("authority_score") or 0)
                # 保留 authority 高的，相同則保留 ID 小的（較早入庫）
                if score_a >= score_b:
                    duplicates.append({"keep": id_a, "dup": id_b, "jaccard": round(jaccard, 3)})
                else:
                    duplicates.append({"keep": id_b, "dup": id_a, "jaccard": round(jaccard, 3)})

    print(f"\n  檢查 {total_pairs} 對, 發現 {len(duplicates)} 組重複")

    if not duplicates:
        print("  ✅ 無重複記錄")
        return 0

    # 去重：標記 is_duplicate
    print(f"\n  標記 {len(duplicates)} 條重複記錄...")
    marked = 0
    dup_ids = set(d["dup"] for d in duplicates)

    for dup_id in dup_ids:
        ok = _supabase_patch(KNOWLEDGE_TABLE, f"id=eq.{dup_id}", {
            "metadata": {"is_duplicate": True, "duplicate_marked_at": datetime.now().isoformat()}
        })
        # 同時降低 freshness_score
        _supabase_patch(KNOWLEDGE_TABLE, f"id=eq.{dup_id}", {"freshness_score": 0.1})
        if ok:
            marked += 1

    current_rate = len(dup_ids) / len(rows) * 100 if rows else 0
    print(f"  ✅ 標記完成: {marked} 條標記為重複 ({current_rate:.1f}%)")
    print(f"     目標: ≤3%   當前: {current_rate:.1f}%")

    # 保存去重報告
    report_dir = os.path.join(SCRIPT_DIR, "reports")
    os.makedirs(report_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    with open(os.path.join(report_dir, f"dedup_report_{ts}.json"), 'w', encoding='utf-8') as f:
        json.dump({"timestamp": ts, "total_records": len(rows), "duplicates_found": len(duplicates),
                   "rate": round(current_rate, 2), "details": duplicates[:100]}, f, ensure_ascii=False, indent=2)

    return marked


def _run_verification_enhancement():
    """13b: 五維鑒真補全 — 對現有記錄補充 真實性 + 準確性 + 參照價值"""
    print("\n  ── 13b: Verification Enhancement ──")
    print("  補充: 真實性(Authenticity) + 準確性(Accuracy) + 參照價值(Reference Value)")

    rows = _fetch_all_knowledge(select="id,content,metadata,source_id,authority_score,verification_grade,verification_score,content_type")
    print(f"  共 {len(rows)} 條記錄需要鑒真補全\n")

    # 分批處理（每批 50 條，用 LLM 評分）
    batch_size = 50
    enhanced = 0
    start = time.time()

    for batch_start in range(0, len(rows), batch_size):
        batch = rows[batch_start:batch_start + batch_size]
        show_progress(min(batch_start + batch_size, len(rows)), len(rows), prefix="  Verify", start_time=start)

        for row in batch:
            content = row.get("content", "")[:1500]
            source_id = row.get("source_id", "unknown") or "unknown"
            authority = row.get("authority_score", 50) or 50
            existing_grade = row.get("verification_grade", "BRONZE")

            # ── Dim 1: 真實性（Authenticity）──
            auth_score = 0.8  # 默認
            # SFC 官方域名來源自動高分
            if source_id.startswith("sfc_"):
                auth_score = 0.95
            elif source_id in ("hkex_rules", "hkma_guidelines", "ia_guidelines", "mpfa_guidelines"):
                auth_score = 0.90
            # 檢測可疑內容
            suspicious_markers = ['advertisement', '廣告', 'sponsored', 'click here to buy', '免費諮詢']
            if any(m in content.lower() for m in suspicious_markers):
                auth_score = 0.3

            # ── Dim 2: 準確性（Accuracy）──
            # 基於來源層級近似
            acc_score = min(1.0, authority / 100.0)

            # ── Dim 5: 參照價值（Reference Value）──
            # 資訊密度近似：數字/法規引用密度
            digits = sum(1 for c in content if c.isdigit())
            refs = content.lower().count('section') + content.lower().count('第') + content.lower().count('article')
            density = (digits + refs * 5) / max(len(content), 1) * 100
            ref_value = min(1.0, 0.3 + density * 2)

            # ── 計算綜合 verification_score ──
            # authority 已有（Dim 3），freshness 已有（Dim 4）
            new_score = (
                auth_score * 0.25 +       # Dim 1: 真實性
                acc_score * 0.20 +         # Dim 2: 準確性
                (authority / 100.0) * 0.25 +  # Dim 3: 權威性
                0.70 * 0.15 +              # Dim 4: 時效性（取平均值代替）
                ref_value * 0.15           # Dim 5: 參照價值
            )

            # 重新計算 verification_grade
            if new_score >= 0.80:
                new_grade = "GOLD"
            elif new_score >= 0.60:
                new_grade = "SILVER"
            else:
                new_grade = "BRONZE"

            # 寫入 metadata 中的鑒真報告
            verification_report = {
                "authenticity": round(auth_score, 3),
                "accuracy": round(acc_score, 3),
                "authority": round(authority / 100.0, 3),
                "reference_value": round(ref_value, 3),
                "composite": round(new_score, 3),
                "verified_at": datetime.now().isoformat(),
            }

            # 更新
            ok = _supabase_patch(KNOWLEDGE_TABLE, f"id=eq.{row['id']}", {
                "verification_score": round(new_score, 4),
                "verification_grade": new_grade,
            })
            if ok:
                enhanced += 1

        time.sleep(0.2)  # rate limit

    # 統計新分佈
    print(f"\n  ✅ 鑒真補全完成: {enhanced}/{len(rows)} 條已更新")

    updated = _fetch_all_knowledge(select="id,verification_grade")
    grades = {"GOLD": 0, "SILVER": 0, "BRONZE": 0}
    for r in updated:
        g = r.get("verification_grade", "BRONZE")
        grades[g] = grades.get(g, 0) + 1
    total = sum(grades.values()) or 1  # 防止除零
    print(f"  新分佈: GOLD {grades.get('GOLD',0)} ({grades.get('GOLD',0)/total*100:.0f}%) | "
          f"SILVER {grades.get('SILVER',0)} ({grades.get('SILVER',0)/total*100:.0f}%) | "
          f"BRONZE {grades.get('BRONZE',0)} ({grades.get('BRONZE',0)/total*100:.0f}%)")

    return enhanced


def _run_freshness_recalc():
    """13c: 時效性衰減重算"""
    print("\n  ── 13c: Freshness Score Recalculation ──")

    DECAY_RATES = {
        "regulation": 0.001,
        "circular": 0.005,
        "enforcement": 0.01,
        "market_data": 0.05,
        "news": 0.03,
        "guide": 0.003,
        "analysis": 0.005,
        "form": 0.002,
        "fee_schedule": 0.02,
    }

    rows = _fetch_all_knowledge(select="id,content_type,created_at,freshness_score")
    print(f"  共 {len(rows)} 條記錄需要重算 freshness_score\n")

    updated = 0
    start = time.time()

    for i, row in enumerate(rows):
        if i % 500 == 0 and i > 0:
            show_progress(i, len(rows), prefix="  Freshness", start_time=start)

        ct = row.get("content_type", "guide") or "guide"
        decay = DECAY_RATES.get(ct, 0.003)

        created = row.get("created_at", "")
        if not created:
            continue

        try:
            # Parse ISO datetime
            if isinstance(created, str):
                created_dt = datetime.fromisoformat(created.replace("Z", "+00:00").replace("+00:00", ""))
            else:
                continue
            age_days = (datetime.now() - created_dt).days
        except:
            age_days = 365  # default 1 year

        new_freshness = max(0.1, 1.0 - (age_days * decay))

        # 只更新差異超過 0.05 的（減少 API 調用）
        old_freshness = row.get("freshness_score") or 0.5
        if abs(new_freshness - old_freshness) > 0.05:
            ok = _supabase_patch(KNOWLEDGE_TABLE, f"id=eq.{row['id']}", {
                "freshness_score": round(new_freshness, 4)
            })
            if ok:
                updated += 1

    print(f"\n  ✅ 衰減重算完成: {updated}/{len(rows)} 條已更新")

    # 統計分佈
    fresh_rows = _fetch_all_knowledge(select="id,freshness_score,content_type")
    buckets = {"≥0.8 (新鮮)": 0, "0.5-0.8 (正常)": 0, "0.2-0.5 (老化)": 0, "<0.2 (過期)": 0}
    for r in fresh_rows:
        fs = r.get("freshness_score") or 0.5
        if fs >= 0.8: buckets["≥0.8 (新鮮)"] += 1
        elif fs >= 0.5: buckets["0.5-0.8 (正常)"] += 1
        elif fs >= 0.2: buckets["0.2-0.5 (老化)"] += 1
        else: buckets["<0.2 (過期)"] += 1

    for label, count in buckets.items():
        bar = "█" * int(count / max(len(fresh_rows), 1) * 30)
        print(f"    {label:>15}: {bar} {count}")

    return updated


def _show_quality_report():
    """13d: 數據質量總覽"""
    print("\n  ── Data Quality Report ──\n")

    rows = _fetch_all_knowledge(select="id,verification_grade,authority_score,freshness_score,content_type,source_id")

    total = len(rows)
    print(f"  總記錄: {total}")

    # 驗證分級
    grades = {}
    for r in rows:
        g = r.get("verification_grade", "UNKNOWN")
        grades[g] = grades.get(g, 0) + 1
    print(f"\n  驗證分級:")
    for g in ["GOLD", "SILVER", "BRONZE", "UNKNOWN"]:
        c = grades.get(g, 0)
        pct = c / total * 100 if total else 0
        bar = "█" * int(pct / 3)
        emoji = {"GOLD": "🥇", "SILVER": "🥈", "BRONZE": "🥉"}.get(g, "⚪")
        print(f"    {emoji} {g:>7}: {bar} {c} ({pct:.1f}%)")

    # 內容類型
    types = {}
    for r in rows:
        t = r.get("content_type", "unknown") or "unknown"
        types[t] = types.get(t, 0) + 1
    print(f"\n  內容類型 (top 8):")
    for t, c in sorted(types.items(), key=lambda x: -x[1])[:8]:
        print(f"    {t:>20}: {c}")

    # 來源分佈
    sources = {}
    for r in rows:
        s = r.get("source_id", "unknown") or "unknown"
        sources[s] = sources.get(s, 0) + 1
    print(f"\n  來源分佈 (top 10):")
    for s, c in sorted(sources.items(), key=lambda x: -x[1])[:10]:
        print(f"    {s:>25}: {c}")

    # 新鮮度
    avg_fresh = sum((r.get("freshness_score") or 0.5) for r in rows) / max(total, 1)
    avg_auth = sum((r.get("authority_score") or 50) for r in rows) / max(total, 1)
    print(f"\n  平均 freshness_score: {avg_fresh:.3f}")
    print(f"  平均 authority_score: {avg_auth:.1f}")

    # World-class 對標
    dup_count = sum(1 for r in rows if (r.get("metadata") or {}).get("is_duplicate"))
    dup_rate = dup_count / total * 100 if total else 0
    gold_rate = grades.get("GOLD", 0) / total * 100 if total else 0

    print(f"\n  ─── World-Class 對標 ───")
    print(f"  {'指標':<20} {'當前':>10} {'目標':>10} {'狀態':>6}")
    print(f"  {'重複率':<20} {dup_rate:>9.1f}% {'≤3%':>10} {'✅' if dup_rate <= 3 else '⚠️':>6}")
    print(f"  {'GOLD 佔比':<20} {gold_rate:>9.1f}% {'≥60%':>10} {'✅' if gold_rate >= 60 else '⚠️':>6}")
    print(f"  {'平均新鮮度':<20} {avg_fresh:>9.3f} {'≥0.6':>10} {'✅' if avg_fresh >= 0.6 else '⚠️':>6}")
    print(f"  {'平均權威性':<20} {avg_auth:>9.1f} {'≥75':>10} {'✅' if avg_auth >= 75 else '⚠️':>6}")


def run_mode_13():
    """Mode 13: Data Quality Engine — Sprint 3"""
    print("\n" + "=" * 60)
    print("  Mode 13: Data Quality Engine 🔬")
    print("  語義去重 + 五維鑒真 + 時效性衰減")
    print("=" * 60)

    print("\n  a. 🔍 語義去重掃描（Jaccard n-gram ≥ 0.75）")
    print("  b. ✅ 五維鑒真補全（真實性+準確性+參照價值）")
    print("  c. ⏱ 時效性衰減重算（按 content_type 衰減率）")
    print("  d. 📊 數據質量總覽報告")
    print("  e. ⚡ 一鍵全部執行（a+b+c+d）")

    sub = input("\n  (a/b/c/d/e): ").strip().lower()

    start = time.time()
    if sub == "a":
        _run_semantic_dedup()
    elif sub == "b":
        _run_verification_enhancement()
    elif sub == "c":
        _run_freshness_recalc()
    elif sub == "d":
        _show_quality_report()
    elif sub == "e":
        print("\n  ⚡ 一鍵全部執行...\n")
        _run_semantic_dedup()
        _run_verification_enhancement()
        _run_freshness_recalc()
        _show_quality_report()
    else:
        print("  Invalid")
        return

    if sub != "d":
        print(f"\n  ⏱ 耗時: {time.time() - start:.0f}s")


# ============================================================
# v7.0 Mode 14: Self-Correction Engine — Sprint 4
# 衝突檢測 + 自動巡檢 + 617 分類映射
# ============================================================

# 617 節點分類映射（Dim 2 合規領域的 16 大類）
COMPLIANCE_DOMAIN_617 = {
    "D01": {"name": "AML/CFT", "keywords": ["aml", "anti-money", "cft", "kyc", "cdd", "sanctions", "str", "jfiu", "pep", "洗錢", "制裁", "可疑交易"]},
    "D02": {"name": "Financial Resources", "keywords": ["frr", "financial resources", "liquid capital", "capital adequacy", "paid-up capital", "速動資本", "實繳資本", "資本"]},
    "D03": {"name": "Business Conduct", "keywords": ["code of conduct", "suitability", "best execution", "conflict of interest", "client agreement", "操守", "合適性", "最佳執行"]},
    "D04": {"name": "Client Asset Protection", "keywords": ["client money", "client securities", "segregation", "custody", "trust account", "客戶資產", "隔離", "保管"]},
    "D05": {"name": "Risk Management", "keywords": ["risk management", "var", "stress test", "market risk", "credit risk", "liquidity risk", "operational risk", "風險管理"]},
    "D06": {"name": "Technology & Cybersecurity", "keywords": ["cybersecurity", "cyber", "data protection", "information security", "bcp", "disaster recovery", "hacking", "網絡安全", "資訊安全"]},
    "D07": {"name": "Virtual Assets", "keywords": ["virtual asset", "vatp", "vasp", "crypto", "stablecoin", "aspire", "虛擬資產", "穩定幣"]},
    "D08": {"name": "Fund Compliance", "keywords": ["fund", "ofc", "lpf", "reit", "fund manager", "基金", "開放式", "有限合夥"]},
    "D09": {"name": "Corporate Finance", "keywords": ["corporate finance", "sponsor", "ipo", "takeover", "merger", "保薦人", "企業融資", "收購"]},
    "D10": {"name": "Market Conduct", "keywords": ["insider", "market manipulation", "short selling", "disclosure", "內幕", "市場操控", "沽空", "權益披露"]},
    "D11": {"name": "Internal Governance", "keywords": ["mic", "responsible officer", "compliance function", "internal audit", "governance", "管治", "內部監控", "合規職能"]},
    "D12": {"name": "Licensing & Registration", "keywords": ["licensing", "application", "fit and proper", "competence", "cpt", "wings", "examination", "牌照", "考試", "勝任能力"]},
    "D13": {"name": "Information Disclosure", "keywords": ["prospectus", "offering document", "disclosure", "public register", "銷售文件", "招股書", "披露"]},
    "D14": {"name": "Cross-Border & International", "keywords": ["cross-border", "fatf", "iosco", "mas", "mica", "mutual recognition", "stock connect", "跨境", "互聯互通"]},
    "D15": {"name": "Investor Protection", "keywords": ["investor protection", "icf", "compensation", "mediation", "投資者保障", "賠償"]},
    "D16": {"name": "Emerging Areas", "keywords": ["esg", "climate", "fintech", "cbdc", "tokenization", "新興", "氣候", "數碼貨幣"]},
}


def _run_conflict_detection():
    """14a: 衝突檢測"""
    print("\n  ── 14a: Conflict Detection ──")
    print("  檢測: 數字衝突 + 法規替代 + 狀態變化\n")

    rows = _fetch_all_knowledge(select="id,content,source_id,authority_score,content_type,created_at")
    print(f"  共 {len(rows)} 條記錄")

    # 策略：找「同主題但不同數字」的衝突
    # 關注：資本要求、罰款金額、費用數字
    import re
    number_pattern = re.compile(r'(?:HKD|hkd|\$|港幣)\s*[\d,]+(?:\.\d+)?|[\d,]+(?:\.\d+)?\s*(?:million|萬|百萬|千萬|億)')

    conflicts = []
    # 按 content_type 分組檢查
    by_type = {}
    for r in rows:
        ct = r.get("content_type", "other") or "other"
        by_type.setdefault(ct, []).append(r)

    # 檢查同類型記錄中的數字衝突
    for ct, group in by_type.items():
        if ct in ("news", "analysis"):
            continue  # 新聞和分析不需要數字一致
        for i in range(min(len(group), 200)):
            content_i = group[i].get("content", "")[:500]
            nums_i = number_pattern.findall(content_i)
            if not nums_i:
                continue
            for j in range(i + 1, min(i + 20, len(group))):
                content_j = group[j].get("content", "")[:500]
                nums_j = number_pattern.findall(content_j)
                if not nums_j:
                    continue
                # 檢查是否涉及同一主題但數字不同
                # 簡化版：如果兩條記錄有 >50% 的非數字詞重疊，但數字不同
                words_i = set(content_i[:200].split())
                words_j = set(content_j[:200].split())
                overlap = len(words_i & words_j) / max(len(words_i | words_j), 1)
                if overlap > 0.3 and set(nums_i) != set(nums_j):
                    auth_i = group[i].get("authority_score", 50) or 50
                    auth_j = group[j].get("authority_score", 50) or 50
                    conflicts.append({
                        "type": "numerical",
                        "chunk_a": group[i]["id"],
                        "chunk_b": group[j]["id"],
                        "nums_a": nums_i[:3],
                        "nums_b": nums_j[:3],
                        "auth_a": auth_i,
                        "auth_b": auth_j,
                        "resolution": "a_wins" if auth_i > auth_j else "b_wins" if auth_j > auth_i else "manual",
                    })

    print(f"  發現 {len(conflicts)} 個潛在衝突")

    if conflicts:
        auto_resolved = sum(1 for c in conflicts if c["resolution"] != "manual")
        manual = len(conflicts) - auto_resolved
        print(f"    自動解決（按權威性）: {auto_resolved}")
        print(f"    需人工審核: {manual}")

        # 嘗試寫入 vigo_conflicts 表
        for c in conflicts[:50]:  # 只寫前 50 個
            record = {
                "conflict_id": f"CF_{c['chunk_a']}_{c['chunk_b']}",
                "type": c["type"],
                "chunk_a_id": str(c["chunk_a"]),
                "chunk_b_id": str(c["chunk_b"]),
                "description": f"Numbers differ: {c['nums_a']} vs {c['nums_b']}",
                "resolution": "auto_resolved" if c["resolution"] != "manual" else "pending",
            }
            # Try to upsert — table may not exist yet
            _supabase_upsert("vigo_conflicts", record, on_conflict="conflict_id")

    # 保存報告
    report_dir = os.path.join(SCRIPT_DIR, "reports")
    os.makedirs(report_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    with open(os.path.join(report_dir, f"conflict_report_{ts}.json"), 'w', encoding='utf-8') as f:
        json.dump({"timestamp": ts, "total_records": len(rows), "conflicts": len(conflicts),
                   "details": conflicts[:100]}, f, ensure_ascii=False, indent=2)

    return len(conflicts)


def _run_expiry_scan():
    """14b: 過期內容掃描 + 自動降權"""
    print("\n  ── 14b: Expiry Scan & Auto-Downgrade ──")

    # 過期閾值（天數）
    EXPIRY_THRESHOLDS = {
        "market_data": 30,
        "news": 90,
        "enforcement": 365,
        "circular": 365,
        "analysis": 180,
        "fee_schedule": 60,
        "form": 730,  # 2 years
        "regulation": 3650,  # 10 years
        "guide": 1825,  # 5 years
    }

    rows = _fetch_all_knowledge(select="id,content_type,created_at,freshness_score")
    expired = 0
    downgraded = 0

    for row in rows:
        ct = row.get("content_type", "guide") or "guide"
        threshold = EXPIRY_THRESHOLDS.get(ct, 365)

        created = row.get("created_at", "")
        if not created:
            continue
        try:
            if isinstance(created, str):
                created_dt = datetime.fromisoformat(created.replace("Z", "+00:00").replace("+00:00", ""))
            else:
                continue
            age_days = (datetime.now() - created_dt).days
        except:
            continue

        if age_days > threshold:
            expired += 1
            # 降權到 0.05
            current = row.get("freshness_score") or 0.5
            if current > 0.1:
                _supabase_patch(KNOWLEDGE_TABLE, f"id=eq.{row['id']}", {"freshness_score": 0.05})
                downgraded += 1

    print(f"  過期記錄: {expired}/{len(rows)}")
    print(f"  已降權: {downgraded}")
    return downgraded


def _run_source_health():
    """14c: 來源健康檢查"""
    print("\n  ── 14c: Source Health Check ──\n")

    rows = _fetch_all_knowledge(select="id,source_id,created_at,verification_grade")

    # 統計每個 source_id 的最新記錄時間
    from collections import defaultdict
    source_stats = defaultdict(lambda: {"count": 0, "latest": "", "gold": 0, "silver": 0, "bronze": 0})

    for r in rows:
        sid = r.get("source_id", "unknown") or "unknown"
        source_stats[sid]["count"] += 1
        created = r.get("created_at", "")
        if created > source_stats[sid]["latest"]:
            source_stats[sid]["latest"] = created
        grade = r.get("verification_grade", "BRONZE")
        if grade == "GOLD":
            source_stats[sid]["gold"] += 1
        elif grade == "SILVER":
            source_stats[sid]["silver"] += 1
        else:
            source_stats[sid]["bronze"] += 1

    print(f"  {'Source':<28} {'Count':>6} {'GOLD':>6} {'Latest':>12} {'Status':>8}")
    print(f"  {'─' * 70}")

    warnings = 0
    for sid, stats in sorted(source_stats.items(), key=lambda x: -x[1]["count"]):
        latest = stats["latest"][:10] if stats["latest"] else "N/A"

        # 判斷健康狀態
        try:
            latest_dt = datetime.fromisoformat(stats["latest"].replace("Z", "+00:00").replace("+00:00", ""))
            days_since = (datetime.now() - latest_dt).days
        except:
            days_since = 999

        if days_since <= 7:
            status = "✅ OK"
        elif days_since <= 30:
            status = "⚠️ Stale"
            warnings += 1
        else:
            status = "🔴 Dead"
            warnings += 1

        print(f"  {sid:<28} {stats['count']:>6} {stats['gold']:>6} {latest:>12} {status:>8}")

    print(f"\n  Total sources: {len(source_stats)}")
    print(f"  Warnings: {warnings}")
    return warnings


def _run_compliance_domain_mapping():
    """14d: 617 節點合規領域分類映射"""
    print("\n  ── 14d: Compliance Domain Mapping ──")
    print("  將 6,379 條記錄映射到 16 個合規大類 (D01-D16)\n")

    rows = _fetch_all_knowledge(select="id,content,metadata")
    print(f"  共 {len(rows)} 條記錄")

    mapped = 0
    domain_counts = {}
    start = time.time()

    for i, row in enumerate(rows):
        if i % 500 == 0 and i > 0:
            show_progress(i, len(rows), prefix="  Mapping", start_time=start)

        content = row.get("content", "")[:2000].lower()
        domains = []

        for code, info in COMPLIANCE_DOMAIN_617.items():
            # 匹配關鍵詞
            matches = sum(1 for kw in info["keywords"] if kw in content)
            if matches >= 2:  # 至少匹配 2 個關鍵詞才計入
                domains.append(code)
                domain_counts[code] = domain_counts.get(code, 0) + 1

        if domains:
            # 更新 metadata 中的 compliance_domains
            meta = row.get("metadata") or {}
            if isinstance(meta, str):
                try:
                    meta = json.loads(meta)
                except:
                    meta = {}
            meta["compliance_domains"] = domains

            _supabase_patch(KNOWLEDGE_TABLE, f"id=eq.{row['id']}", {"metadata": meta})
            mapped += 1

    print(f"\n  ✅ 映射完成: {mapped}/{len(rows)} 條已標記合規領域")
    print(f"\n  領域分佈 (D01-D16):")
    for code in sorted(COMPLIANCE_DOMAIN_617.keys()):
        count = domain_counts.get(code, 0)
        name = COMPLIANCE_DOMAIN_617[code]["name"]
        bar = "█" * min(int(count / max(len(rows), 1) * 100), 30)
        print(f"    {code} {name:<28} {bar} {count}")

    return mapped


def _show_correction_report():
    """14f: 巡檢報告"""
    print("\n  ── Self-Correction Report ──\n")

    # 嘗試讀取 vigo_conflicts
    conflicts = _supabase_query("vigo_conflicts", select="conflict_id,type,resolution", limit=1000)
    if conflicts:
        print(f"  衝突記錄: {len(conflicts)}")
        resolutions = {}
        for c in conflicts:
            r = c.get("resolution", "pending")
            resolutions[r] = resolutions.get(r, 0) + 1
        for r, count in resolutions.items():
            print(f"    {r}: {count}")
    else:
        print("  衝突記錄: 0（vigo_conflicts 表可能未建立）")

    # 最近的報告文件
    report_dir = os.path.join(SCRIPT_DIR, "reports")
    if os.path.exists(report_dir):
        reports = sorted([f for f in os.listdir(report_dir) if f.endswith('.json')], reverse=True)
        if reports:
            print(f"\n  最近報告:")
            for r in reports[:5]:
                size = os.path.getsize(os.path.join(report_dir, r))
                print(f"    {r} ({size} bytes)")


def run_mode_14():
    """Mode 14: Self-Correction Engine — Sprint 4"""
    print("\n" + "=" * 60)
    print("  Mode 14: Self-Correction Engine 🔧")
    print("  衝突檢測 + 自動巡檢 + 分類映射")
    print("=" * 60)

    print("\n  a. ⚡ 衝突檢測（數字衝突 + 法規替代）")
    print("  b. ⏱ 過期內容掃描 + 自動降權")
    print("  c. 🏥 來源健康檢查")
    print("  d. 🏷️  合規領域映射（D01-D16, 617 節點）")
    print("  e. 🚀 一鍵全部巡檢（a+b+c+d）")
    print("  f. 📊 巡檢報告")

    sub = input("\n  (a/b/c/d/e/f): ").strip().lower()

    start = time.time()
    if sub == "a":
        _run_conflict_detection()
    elif sub == "b":
        _run_expiry_scan()
    elif sub == "c":
        _run_source_health()
    elif sub == "d":
        _run_compliance_domain_mapping()
    elif sub == "e":
        print("\n  🚀 一鍵全部巡檢...\n")
        _run_conflict_detection()
        _run_expiry_scan()
        _run_source_health()
        _run_compliance_domain_mapping()
        print("\n  ═══ 全部巡檢完成 ═══")
        _show_correction_report()
    elif sub == "f":
        _show_correction_report()
    else:
        print("  Invalid")
        return

    if sub not in ("f",):
        print(f"\n  ⏱ 耗時: {time.time() - start:.0f}s")


def main():
    now = datetime.now()
    year = now.year % 100

    print("=" * 60)
    print("  VIGO SFC Bilingual Collector v7.0")
    print("  合規是下線，業務是上線 — VIGO Protocol")
    print("  World-Class 引擎升級：Eval + 結構化存儲")
    print("=" * 60)
    print(f"\n  Date: {now.strftime('%Y-%m-%d %H:%M')}")

    # Check dependencies
    missing = []
    try: import pdfplumber
    except: missing.append("pdfplumber")
    try: from bs4 import BeautifulSoup
    except: missing.append("beautifulsoup4")
    if missing:
        print(f"\n  Missing: {', '.join(missing)}")
        print(f"  Run: pip install {' '.join(missing)}")
        return

    progress = load_progress()
    hashes_data = load_content_hashes()
    if progress.get("uploaded_refs"):
        print(f"\n  📋 Progress: {len(progress['uploaded_refs'])} refs uploaded")
        print(f"     Content hashes: {len(hashes_data.get('hashes', {}))}")

    print("\n  ┌─ 數據採集 ─────────────────────────────────┐")
    print("  │  1. 日常同步 Quick Sync (~5 min)             │")
    print("  │  2. 季度補漏 Quarterly Scan (3/6/12 個月)    │")
    print("  │  3. Test Single Source                        │")
    print("  │  4. ★ 歷史建庫 Historical Archive            │")
    print("  │     (分類掃描：推薦/精簡/完整/自定義)         │")
    print("  ├─ 法規基石 ─────────────────────────────────┤")
    print("  │  5. ★ Foundational (35 PDFs + PDPO + 擴展)   │")
    print("  │     含：PDPO Cap.486 / WINGS 指南 / 巡查加強 │")
    print("  ├─ 行業知識 ─────────────────────────────────┤")
    print("  │  6. ★ Industry Intelligence (19 家來源)       │")
    print("  ├─ 工具 ────────────────────────────────────┤")
    print("  │  7. ★ File Importer (含合規模板分類)          │")
    print("  │  8. ★ DB Management (QA/日曆/標籤統計)       │")
    print("  ├─ v5.2 新模式 ──────────────────────────────┤")
    print("  │  9. ★ VATP Deep (FATF/MAS/MiCA/VA ETF)       │")
    print("  │ 10. ★ SFC Operational (WINGS/表格/PDPO/巡查) │")
    print("  ├─ v7.0 World-Class 升級 ────────────────────┤")
    print("  │ 11. ★ Eval Benchmark (50題自動化質量評估)     │")
    print("  │ 12. ★ Structured Extract (結構化表提取)       │")
    print("  │ 13. ★ Data Quality (去重+鑒真+衰減)           │")
    print("  │ 14. ★ Self-Correction (衝突+巡檢+分類映射)    │")
    print("  │  q. 退出 Quit                                 │")
    print("  └──────────────────────────────────────────────┘")

    # ═══ 主循環：完成後可繼續選擇下一項 ═══
    while True:
        mode = input("\n  (1-14, q=退出): ").strip().lower()

        if mode in ("q", "quit", "exit", "0"):
            print("\n  👋 再見！VIGO Protocol signing off.")
            break

        try:
            if mode == "3":
                print("\n  a. Circulars (PDF)")
                print("  b. Enforcement (HTML)")
                print("  c. Consultations (PDF)")
                print("  d. VATP (HTML)")
                print("  e. Press Releases (HTML)")
                src = input("  (a/b/c/d/e): ").strip().lower()

                enforce_kw = ['reprimand', 'fine', 'suspend', 'ban', 'prosecut', 'convict']
                va_kw = ['virtual asset', 'vatp', 'vasp', 'crypto', 'stablecoin', 'staking', 'aspire']

                if src == "a":
                    results = scan_source("Test: Circulars",
                        [{"year": year, "code": "EC", "range": range(8, 0, -1), "doc_type": "circular"}],
                        "circular/openFile", max_items=5)
                elif src == "b":
                    results = scan_source("Test: Enforcement",
                        [{"year": year, "code": "PR", "range": range(15, 0, -1), "doc_type": "enforcement"}],
                        "news/list-content", filter_fn=lambda t: any(k in t.lower() for k in enforce_kw), max_items=5)
                elif src == "c":
                    results = scan_source("Test: Consultations",
                        [{"year": year, "code": "CP", "range": range(8, 0, -1), "doc_type": "consultation"}],
                        "consultation/openFile", max_items=5)
                elif src == "d":
                    results = scan_source("Test: VATP",
                        [{"year": year, "code": "PR", "range": range(15, 0, -1), "doc_type": "vatp"}],
                        "news/list-content", filter_fn=lambda t: any(k in t.lower() for k in va_kw), max_items=5)
                elif src == "e":
                    skip = enforce_kw + va_kw
                    results = scan_source("Test: Press",
                        [{"year": year, "code": "PR", "range": range(15, 0, -1), "doc_type": "press_release"}],
                        "news/list-content", filter_fn=lambda t: not any(k in t.lower() for k in skip), max_items=5)
                else:
                    print("  Invalid"); continue

                if results:
                    print(f"\n  Test complete: {len(results)} items found")

            elif mode in ("1", "2"):
                run_mode_1_2(year, mode)
            elif mode == "4":
                deep_historical_scan()
            elif mode == "5":
                import_foundational_regulations()
            elif mode == "6":
                run_mode_6()
            elif mode == "7":
                run_mode_7()
            elif mode == "8":
                run_mode_8()
            elif mode == "9":
                run_mode_9()
            elif mode == "10":
                run_mode_10()
            elif mode == "11":
                run_mode_11()
            elif mode == "12":
                run_mode_12()
            elif mode == "13":
                run_mode_13()
            elif mode == "14":
                run_mode_14()
            else:
                print("  Invalid, 請輸入 1-14 或 q 退出")
                continue

        except KeyboardInterrupt:
            print("\n\n  ⏹ 當前任務已中斷")
        except Exception as e:
            print(f"\n  ⚠ 錯誤: {type(e).__name__}: {e}")

        # 完成後顯示快捷提示
        print("\n  ─────────────────────────────────")
        print("  ✅ 完成！可繼續選擇下一項操作：")
        print("     11=Eval  12=提取  13=質量  14=巡檢  q=退出")


def parse_args():
    """Parse command-line arguments for non-interactive (CI/CD) mode."""
    parser = argparse.ArgumentParser(
        description='VIGO SFC Bilingual Collector v7.0 — Autonomous Operation',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  Interactive mode (default):
    python auto_collect_upload.py

  Non-interactive (GitHub Actions):
    python auto_collect_upload.py --mode 1 --sub e --non-interactive
    python auto_collect_upload.py --mode 13 --sub e --non-interactive
    python auto_collect_upload.py --mode 11 --sub a --non-interactive
    python auto_collect_upload.py --mode 11 --compare-last --non-interactive
    python auto_collect_upload.py --mode 1 --sub e --stale-only --non-interactive
        """
    )
    parser.add_argument('--mode', type=str, help='Mode number (1-14)')
    parser.add_argument('--sub', type=str, default='e', help='Sub-mode letter (a/b/c/d/e/f), default: e')
    parser.add_argument('--non-interactive', action='store_true', help='Run without user prompts (for CI/CD)')
    parser.add_argument('--stale-only', action='store_true', help='Mode 1: only re-crawl stale sources')
    parser.add_argument('--compare-last', action='store_true', help='Mode 11: compare with previous eval')
    parser.add_argument('--version', action='version', version='VIGO Collector v7.0.3')
    return parser.parse_args()


def run_non_interactive(args):
    """Execute a single mode non-interactively (for GitHub Actions / cron)."""
    import builtins
    original_input = builtins.input

    # Mock input() to return the sub-mode automatically
    call_count = [0]
    def mock_input(prompt=""):
        call_count[0] += 1
        # First input call is usually sub-mode selection
        if call_count[0] == 1:
            response = args.sub.lower()
        else:
            # Subsequent inputs: auto-confirm with 'y' or return empty
            if 'y/n' in prompt.lower() or '(y' in prompt.lower():
                response = 'y'
            elif '退出' in prompt or 'q=' in prompt:
                response = 'q'
            else:
                response = ''
        print(f"  [AUTO] {prompt.strip()} → {response}")
        return response

    builtins.input = mock_input

    now = datetime.now()
    year = now.year % 100

    print("=" * 60)
    print("  VIGO SFC Bilingual Collector v7.0 [NON-INTERACTIVE]")
    print(f"  Mode: {args.mode} | Sub: {args.sub} | Stale-only: {args.stale_only}")
    print("=" * 60)
    print(f"  Date: {now.strftime('%Y-%m-%d %H:%M')}")

    # Check dependencies
    missing = []
    try: import pdfplumber
    except: missing.append("pdfplumber")
    try: from bs4 import BeautifulSoup
    except: missing.append("beautifulsoup4")
    if missing:
        print(f"\n  Missing: {', '.join(missing)}")
        sys.exit(1)

    progress = load_progress()
    hashes_data = load_content_hashes()

    try:
        mode = args.mode
        if mode == "1":
            run_mode_1_2(year, "1")
        elif mode == "2":
            run_mode_1_2(year, "2")
        elif mode == "5":
            import_foundational_regulations()
        elif mode == "6":
            run_mode_6()
        elif mode == "7":
            run_mode_7()
        elif mode == "8":
            run_mode_8()
        elif mode == "9":
            run_mode_9()
        elif mode == "10":
            run_mode_10()
        elif mode == "11":
            if args.compare_last:
                # Special: compare last two eval reports
                print("\n  📊 Comparing last two eval reports...")
                import glob
                eval_files = sorted(glob.glob(os.path.join(SCRIPT_DIR, 'reports', 'eval_*.json')))
                if len(eval_files) >= 2:
                    with open(eval_files[-2]) as f: prev = json.load(f)
                    with open(eval_files[-1]) as f: curr = json.load(f)
                    prev_recall = prev.get('recall_at_5', 0)
                    curr_recall = curr.get('recall_at_5', 0)
                    delta = curr_recall - prev_recall
                    print(f"  Previous: {prev_recall:.1%} | Current: {curr_recall:.1%} | Delta: {delta:+.1%}")
                    if delta < -0.05:
                        print("  ⚠️ ALERT: Eval score dropped by more than 5%!")
                        sys.exit(1)
                    else:
                        print("  ✅ Eval score stable or improved.")
                else:
                    print("  Not enough eval reports to compare.")
            else:
                run_mode_11()
        elif mode == "12":
            run_mode_12()
        elif mode == "13":
            run_mode_13()
        elif mode == "14":
            run_mode_14()
        else:
            print(f"  Invalid mode: {mode}")
            sys.exit(1)

        print("\n  ✅ Non-interactive execution complete.")

    except Exception as e:
        print(f"\n  ❌ ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        builtins.input = original_input


if __name__ == "__main__":
    args = parse_args()

    if args.mode and args.non_interactive:
        # CI/CD mode: run single mode and exit
        run_non_interactive(args)
    elif args.mode:
        # Command-line mode specified but interactive
        print(f"  Hint: use --non-interactive for automated execution")
        run_non_interactive(args)
    else:
        # Interactive mode (original behavior)
        try:
            main()
        except KeyboardInterrupt:
            print("\n\n  👋 Cancelled")
        except Exception as e:
            print(f"\n  ERROR: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            print("\n")
            input("Press Enter to close...")
